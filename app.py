# -*- coding: utf-8 -*-
"""ToFanari — Main GUI application."""

import csv
import os
import re
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import List, Tuple
from urllib.parse import quote


def _get_app_base_path() -> str:
    """Return a base path safe for packaged (frozen) execution. Use for any app-bundled resources."""
    if getattr(sys, "frozen", False):
        return getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def _get_book_registry_path() -> str:
    """Fixed path for book_registry.xlsx. Uses get_settings_dir()."""
    from config import get_settings_dir
    return os.path.join(get_settings_dir(), "book_registry.xlsx")


try:
    from PIL import ImageTk
except ImportError:
    ImageTk = None

from config import (
    APP_TITLE,
    APP_VERSION,
    VERSION,
    CREAM,
    DARK_RED,
    DEFAULT_BUNNY_BASE_URL,
    BUNNY_ROOT_FOLDER,
    GOLD,
    GREY_LIGHT,
    GREY_MED,
    LIGHT_RED,
    MARKER,
    MID_RED,
    TEXT_DARK,
    WHITE,
    GREEN,
)
from database import (
    build_database_xlsx,
    build_database_xlsx_from_marker_matches,
    preview_lines,
    preview_lines_from_marker_matches,
    validate_mp3_files,
    get_mp3_code,
    get_mp3_file,
    get_mp3_url,
    get_mp3_url_for_source_file,
)
from pdf_ops import (
    Marker,
    apply_markers,
    detect_markers,
    detect_numbered_markers_from_pdf,
    detect_tofanari_markers_from_pdf,
    embed_numbered_markers_pdf,
    extract_preview_text,
    extract_hymn_preview_lines,
    render_page_thumbnail,
)
from validators import (
    validate_duplicate_positions,
    validate_page_numbers,
)
from validation_preflight import (
    run_full_validation,
    validate_mp3_numbering,
    validate_mp3_filename_pattern,
)
from book_registry import (
    load_book_registry,
    find_book_for_chapter,
    append_book_to_registry,
    get_expected_chapters_for_book,
    compare_imported_vs_expected,
    validate_chapter_in_catalog,
    ChapterDef,
)
from imported_chapters import (
    load_imported_chapters,
    add_imported_chapter,
    get_imported_chapter_codes,
)
from bunny_preparation import (
    run_chapter_preparation,
    format_preparation_report,
    create_manifest_template,
)
from marker_matching import (
    match_markers_to_mp3,
    format_matching_report,
    MarkerMatchingResult,
    parse_mp3_folder,
    extract_mp3_id,
    build_mp3_prefix_map,
    sort_mp3_filenames_by_numeric_prefix,
    POSTUPLOAD_AUDIO_MAPPING_JSON,
    PREUPLOAD_AUDIO_MAPPING_JSON,
    run_stage1_preupload_save,
    run_stage2_postupload_save,
)
from ready_pdf import validate_ready_pdf, ReadyPdfValidationResult
from tofanari_log import log_import
from parameters import (
    load_catalog_from_parameters,
    load_parameters,
    save_parameters,
    get_bookshelf_structure,
    get_collection_completeness,
    DEFAULT_PATH as PARAMETERS_DEFAULT_PATH,
)


# ─────────────────────────── UI helpers ──────────────────────────────

def _btn(
    parent,
    text,
    command,
    bg=DARK_RED,
    fg=WHITE,
    fs=11,
    pady=8,
    width=None,
    big=False,
    **kw,
):
    font_size = 13 if big else fs
    b = tk.Button(
        parent,
        text=text,
        command=command,
        bg=bg,
        fg=fg,
        font=("Arial", font_size, "bold"),
        relief="flat",
        cursor="hand2",
        pady=pady,
        activebackground=MID_RED,
        activeforeground=WHITE,
        **kw,
    )
    if width:
        b.config(width=width)
    return b


def _clean_preview_for_title(preview: str) -> str:
    """Clean extracted preview text for use as hymn title: remove bullets/symbols, normalize spaces."""
    if not preview or not isinstance(preview, str):
        return ""
    s = preview.strip()
    for ch in "\u25a0\u25aa\u2022\u00b7\u2023\u2043\u2219\u2024\u2027\u2022\u2024":
        s = s.replace(ch, " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s[:200] if len(s) > 200 else s


def song_title_from_mp3_file(mp3_file: str) -> str:
    """Return MP3 filename without .mp3 extension. No cleaning or transformation."""
    if not mp3_file:
        return ""
    return os.path.splitext(os.path.basename(mp3_file))[0]


def is_internal_generated_mp3(filename: str, book_code: str) -> bool:
    """
    True only if filename matches BOOKCODE-NNN.mp3 (already-generated internal format).
    Excludes: AN01-001.mp3, BOOK-001.mp3, <BOOKCODE>-<number>.mp3
    Includes: 001 Title.mp3, 010 Another.mp3 (source files)
    """
    if not filename or not (book_code or "").strip():
        return False
    code = (book_code or "").strip()
    # Must match: BOOKCODE- then digits only, then .mp3 (case-insensitive)
    pattern = rf"^{re.escape(code)}-\d+\.mp3$"
    return re.match(pattern, filename, re.IGNORECASE) is not None


def split_mp3_files(mp3_folder: str, book_code: str) -> tuple[List[str], List[str]]:
    """
    Split MP3 files in folder into source and generated.
    Mixed folders (source + generated) are supported and expected.
    Returns: (source_files, generated_files)
    - source_files: original files (001 Title.mp3, etc.)
    - generated_files: app-created BOOKCODE-NNN.mp3
    """
    source_files: List[str] = []
    generated_files: List[str] = []
    if not mp3_folder or not os.path.isdir(mp3_folder):
        return (source_files, generated_files)
    code = (book_code or "").strip()
    try:
        for f in os.listdir(mp3_folder):
            full = os.path.join(mp3_folder, f)
            if not os.path.isfile(full):
                continue
            if not f.lower().endswith(".mp3"):
                continue
            if is_internal_generated_mp3(f, code):
                generated_files.append(f)
            else:
                source_files.append(f)
    except OSError:
        pass
    return (sorted(source_files), sorted(generated_files))


def get_source_mp3_files(mp3_folder: str, book_code: str) -> List[str]:
    """Return source MP3 files (convenience wrapper around split_mp3_files for mixed folder)."""
    source, _ = split_mp3_files(mp3_folder, book_code)
    return source


def list_mp3_files_in_folder(folder: str) -> List[str]:
    """List all .mp3 files in folder (case-insensitive), ordered by leading hymn number (001…010…100)."""
    if not folder or not os.path.isdir(folder):
        return []
    try:
        names = [
            f for f in os.listdir(folder)
            if os.path.isfile(os.path.join(folder, f)) and f.lower().endswith(".mp3")
        ]
        return sort_mp3_filenames_by_numeric_prefix(names)
    except OSError:
        return []


def _enable_clipboard_paste(root, widget):
    """Enable Ctrl+V, Shift+Insert, and right-click paste for Entry or Text widgets.
    Works with Greek and other keyboard layouts (keycode-based fallback).
    """
    def do_paste(event=None):
        try:
            widget.event_generate("<<Paste>>")
        except Exception:
            try:
                text = root.clipboard_get()
                if text:
                    if isinstance(widget, tk.Text):
                        if widget.selection_present():
                            widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
                        widget.insert(tk.INSERT, text)
                    else:
                        if widget.selection_present():
                            widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
                        widget.insert(tk.INSERT, text)
            except Exception:
                pass
        return "break"

    def on_key(event):
        # keycode 86 = physical V key; Ctrl mask = 0x4
        if event.keycode == 86 and (event.state & 0x4):
            if event.keysym.lower() != "v":
                do_paste(event)
                return "break"

    def show_context_menu(event):
        menu = tk.Menu(widget, tearoff=0)
        menu.add_command(label="Επικόλληση", command=lambda: do_paste())
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    widget.bind("<Control-v>", lambda e: do_paste(e), add="+")
    widget.bind("<Control-V>", lambda e: do_paste(e), add="+")
    widget.bind("<Shift-Insert>", lambda e: do_paste(e), add="+")
    widget.bind("<Key>", on_key, add="+")
    widget.bind("<Button-3>", show_context_menu, add="+")
    if sys.platform == "darwin":
        widget.bind("<Command-v>", lambda e: do_paste(e), add="+")


# ─────────────────────────── Main Application ────────────────────────

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{APP_TITLE}  {VERSION}")
        self.root.geometry("1000x760")
        self.root.configure(bg=CREAM)
        self.root.resizable(True, True)

        self.pdf_path = tk.StringVar()  # Original PDF (selected in Step 1)
        self.marked_pdf_path = tk.StringVar()  # Marked PDF (created in Step 2)
        self.original_pdf_path = ""  # Selected PDF path (set by _pick_pdf)
        self.work_folder = ""  # dirname(original_pdf_path), computed after PDF selection
        self.status = tk.StringVar(value=f"Καλωσήρθατε στο {APP_TITLE} {VERSION}")
        self.mrks: List[Marker] = []
        self.assignments: List[dict] = []  # one per marker: song_title, echos, section (StringVars)
        # Mapping from internal MP3 filename (e.g. AN01-001.mp3) to original source name
        # e.g. "AN01-001.mp3" -> "001 Κύριε Ἐκέκραξα.mp3"
        self.mp3_original_name_by_new_file: dict[str, str] = {}

        self.source_mp3_folder = ""  # Exact path selected by user; never append input_mp3/mp3
        self.html_chapter_folder = ""  # Chapter folder with index.html (for marker-based matching)
        self.marker_matching_result: MarkerMatchingResult | None = None
        self.ready_pdf_path: str = ""  # Validated ready PDF (alternative to full workflow)
        self.ready_pdf_validation: ReadyPdfValidationResult | None = None
        self.ready_pipeline_mp3_folder: str = ""  # MP3 folder for step-by-step import
        self.ready_pipeline_mp3_passed: bool = False  # Step 2 validation result
        self.ready_pipeline_book_slug: str = ""  # Step 3 selection
        self.ready_pipeline_chapter_code: str = ""
        self.dv = {
            "code": tk.StringVar(value="BOOK"),
            "book_slug": tk.StringVar(),  # For catalog-based selection; used with code (chapter_code)
            "bunny": tk.StringVar(value=DEFAULT_BUNNY_BASE_URL),
            "mp3_folder": tk.StringVar(),  # Mirrors source_mp3_folder for UI
            "html_folder": tk.StringVar(),  # Chapter folder with HTML (for marker matching)
        }
        self.bunny_prep = {
            "base_cdn_url": tk.StringVar(value=DEFAULT_BUNNY_BASE_URL),
            "root_remote_folder": tk.StringVar(value=BUNNY_ROOT_FOLDER),
            "book_slug": tk.StringVar(),
            "chapter_code": tk.StringVar(),
            "thinkific_course_name": tk.StringVar(),
            "thinkific_chapter_name": tk.StringVar(),
            "flipbuilder_book_name": tk.StringVar(),
            "api_key": tk.StringVar(),
        }

        self._build_ui()

        # Step 3 gating state (strict Step 2 completion requirement)
        self._markers_read_from_pdf: bool = False
        self._step3_ready: bool = False

    def _update_step3_ready_state(self) -> None:
        """Step 3 is allowed only when markers exist and numbering exists."""
        markers_exist = bool(self.mrks)
        numbering_ok = bool(
            markers_exist
            and all(
                (m.number is not None and isinstance(m.number, int) and m.number > 0)
                for m in self.mrks
                if m.keep
            )
        )
        self._step3_ready = bool(markers_exist and numbering_ok)

    def _ensure_step3_ready_or_block(self) -> bool:
        self._update_step3_ready_state()
        if (not self._step3_ready) and bool(self.mrks):
            # Markers exist but numbering may be missing: generate numbering automatically.
            self._tab2_renumber_markers()
            self._update_step3_ready_state()
        if self._step3_ready:
            return True
        messagebox.showerror(
            "Step 3 blocked",
            "PDF is not marked. Please complete Step 2 (Σήμανση).",
        )
        return False

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg=DARK_RED, height=70)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="♪", bg=DARK_RED, fg=GOLD, font=("Arial", 28)).pack(
            side="left", padx=15
        )
        tf = tk.Frame(hdr, bg=DARK_RED)
        tf.pack(side="left")
        tk.Label(tf, text="ToFanari", bg=DARK_RED, fg=WHITE, font=("Georgia", 22, "bold")).pack(
            anchor="w"
        )
        tk.Label(
            tf,
            text="Βυζαντινή Μουσική — Σύστημα Ψηφιακών Βιβλίων",
            bg=DARK_RED,
            fg="#DDDDDD",
            font=("Arial", 10),
        ).pack(anchor="w")
        tk.Label(hdr, text=VERSION, bg=DARK_RED, fg=GOLD, font=("Arial", 10)).pack(
            side="right", padx=15
        )

        # Notebook
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=CREAM, borderwidth=0)
        style.configure(
            "TNotebook.Tab",
            background=GREY_LIGHT,
            foreground=TEXT_DARK,
            padding=[12, 6],
            font=("Arial", 10, "bold"),
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", DARK_RED)],
            foreground=[("selected", WHITE)],
        )
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True)

        self.tab1 = tk.Frame(nb, bg=CREAM)
        self.tab2 = tk.Frame(nb, bg=CREAM)
        self.tab3 = tk.Frame(nb, bg=CREAM)
        self.tab4 = tk.Frame(nb, bg=CREAM)
        self.tab5 = tk.Frame(nb, bg=CREAM)
        self.tab6 = tk.Frame(nb, bg=CREAM)
        self.tab7 = tk.Frame(nb, bg=CREAM)
        self.tab8 = tk.Frame(nb, bg=CREAM)
        nb.add(self.tab8, text="  0. Parameters / Master Catalog  ")
        nb.add(self.tab1, text="  1. Επιλογή Βιβλίου  ")
        nb.add(self.tab2, text="  2. Σήμανση PDF  ")
        nb.add(self.tab3, text="  3. Έλεγχος Κουμπιών  ")
        nb.add(self.tab4, text="  4. Βάση Δεδομένων  ")
        nb.add(self.tab5, text="  5. Αντιστοίχιση Ύμνων  ")
        nb.add(self.tab6, text="  6. Έλεγχος Βάσης  ")
        nb.add(self.tab7, text="  7. Bunny.net Preparation  ")
        nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.book_registry_books: List[dict] = []
        self.book_registry_path: str = ""
        self.parameters_path: str = ""  # settings/parameters.json when loaded from Parameters
        self.catalog_source: str = ""  # "parameters" | "excel" | ""
        self.book_registry_add_vars: dict = {}

        self._build_tab1()
        self._build_tab2()
        self._build_tab3()
        self._build_tab4()
        self._build_tab5()
        self._build_tab6()
        self._build_tab7()
        self._build_tab8()

        # Auto-load Parameters on startup if available
        self._load_parameters_on_startup()

        # Status bar
        sb = tk.Frame(self.root, bg=DARK_RED, height=28)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        tk.Label(
            sb,
            textvariable=self.status,
            bg=DARK_RED,
            fg=WHITE,
            font=("Arial", 9),
            anchor="w",
        ).pack(fill="x", padx=10, pady=4)

    def _section_header(self, parent, text: str):
        fr = tk.Frame(parent, bg=DARK_RED, height=42)
        fr.pack(fill="x")
        fr.pack_propagate(False)
        tk.Label(
            fr, text=text, bg=DARK_RED, fg=WHITE, font=("Arial", 13, "bold")
        ).pack(side="left", padx=15, pady=8)

    def _card(self, parent):
        c = tk.Frame(parent, bg=WHITE, bd=0)
        c.pack(fill="both", expand=True, padx=20, pady=16)
        return c

    def get_mp3_folder(self) -> str:
        """
        Canonical MP3 folder path. Set in Tab 4 (picker) or Tab 6 Step 2.
        Returns source_mp3_folder, or ready_pipeline_mp3_folder if set (Step 2 before confirm).
        """
        path = (self.source_mp3_folder or self.ready_pipeline_mp3_folder or "").strip()
        return path

    def get_current_mp3_files(self) -> List[str]:
        """
        Return sorted list of MP3 filenames in the active MP3 folder only.
        Uses os.listdir(mp3_folder) only — no recursive, no subfolders, no cache.
        Excludes directories via os.path.isfile.
        """
        mp3_folder = self.get_mp3_folder()
        if not mp3_folder or not os.path.isdir(mp3_folder):
            return []
        try:
            return sorted(
                f for f in os.listdir(mp3_folder)
                if os.path.isfile(os.path.join(mp3_folder, f)) and f.lower().endswith(".mp3")
            )
        except OSError:
            return []

    def reset_mp3_runtime_state(self) -> None:
        """
        Clear MP3-related runtime state. Call at start of Έλεγχος MP3 and Validate Database.
        """
        self.mp3_original_name_by_new_file.clear()
        if hasattr(self, "validation_report_text") and self.validation_report_text.winfo_exists():
            try:
                self.validation_report_text.config(state="normal")
                self.validation_report_text.delete("1.0", tk.END)
                self.validation_report_text.insert("1.0", "Εκτελέστε «Έλεγχος Βάσης» για νέα αναφορά.")
                self.validation_report_text.config(state="disabled")
            except tk.TclError:
                pass

    def _label(self, parent, text: str):
        tk.Label(
            parent,
            text=text,
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w", pady=(10, 2))

    def _s(self, msg: str):
        self.status.set(msg)

    def _build_tab1(self):
        f = self.tab1
        self._section_header(f, "📚  Επιλογή Βιβλίου")
        card = self._card(f)

        self._label(card, "Original PDF / Αρχείο PDF πρωτότυπο:")
        row = tk.Frame(card, bg=WHITE)
        row.pack(fill="x", pady=3)
        tk.Entry(
            row,
            textvariable=self.pdf_path,
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(side="left", fill="x", expand=True)
        _btn(
            row,
            "Νέο Βιβλίο / Reset",
            self._confirm_reset_project,
            bg=MID_RED,
            width=18,
        ).pack(side="right", padx=(6, 4))
        _btn(row, "Επιλογή PDF…", self._pick_pdf, width=14).pack(side="right", padx=(6, 0))

        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=15)

        info = tk.Frame(card, bg=LIGHT_RED, bd=0)
        info.pack(fill="x", pady=6)
        tk.Label(
            info,
            text="Αντιγράψτε τον δείκτη ■ και επικολλήστε τον στο Melodos:",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w", padx=10, pady=(8, 2))
        copy_row = tk.Frame(info, bg=LIGHT_RED)
        copy_row.pack(anchor="w", padx=10, pady=(0, 8))
        tk.Label(
            copy_row, text="■", bg=LIGHT_RED, fg=DARK_RED, font=("Arial", 24)
        ).pack(side="left", padx=(0, 12))
        _btn(
            copy_row,
            "📋  Αντιγραφή ■",
            lambda: (
                self.root.clipboard_clear(),
                self.root.clipboard_append(MARKER),
                self._s("✅ Ο δείκτης ■ αντιγράφηκε!"),
            ),
            bg=GREEN,
            pady=6,
        ).pack(side="left")

    def _confirm_reset_project(self):
        """Ask confirmation then reset project state."""
        if messagebox.askyesno(
            "Νέο Βιβλίο / Reset",
            "Do you want to clear the current project and start a new one?",
        ):
            self.reset_project_state()

    def reset_project_state(self):
        """Clear all current project state; app behaves like a fresh launch. Does not close the app."""
        self.pdf_path.set("")
        self.marked_pdf_path.set("")
        self.original_pdf_path = ""
        self.work_folder = ""
        self.source_mp3_folder = ""
        self.html_chapter_folder = ""
        self.marker_matching_result = None
        self.ready_pdf_path = ""
        self.ready_pdf_validation = None
        self.ready_pipeline_mp3_folder = ""
        self.ready_pipeline_mp3_passed = False
        self.ready_pipeline_book_slug = ""
        self.ready_pipeline_chapter_code = ""
        self.dv["mp3_folder"].set("")
        self.dv["html_folder"].set("")
        self.dv["code"].set("BOOK")
        self.dv["bunny"].set(DEFAULT_BUNNY_BASE_URL)
        self.mrks = []
        self._markers_read_from_pdf = False
        self._step3_ready = False
        self.assignments = []
        self.mp3_original_name_by_new_file.clear()
        self.status.set(f"Καλωσήρθατε στο {APP_TITLE} {VERSION}")
        if hasattr(self, "tab2_original_pdf_lbl") and self.tab2_original_pdf_lbl.winfo_exists():
            self.tab2_original_pdf_lbl.config(text="—")
        if hasattr(self, "tab2_marked_pdf_lbl") and self.tab2_marked_pdf_lbl.winfo_exists():
            self.tab2_marked_pdf_lbl.config(text="—")
        if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
            self.cnt_lbl.config(text="—", fg=TEXT_DARK)
        if hasattr(self, "lst") and self.lst.winfo_exists():
            self.lst.delete(0, "end")
        if hasattr(self, "mp3_files_count_lbl") and self.mp3_files_count_lbl.winfo_exists():
            self.mp3_files_count_lbl.config(text="MP3 Files Found: 0")
        if hasattr(self, "marker_tree") and self.marker_tree.winfo_exists():
            for item in self.marker_tree.get_children():
                self.marker_tree.delete(item)
        if hasattr(self, "marker_matching_lbl") and self.marker_matching_lbl.winfo_exists():
            self.marker_matching_lbl.config(text="")
        if hasattr(self, "mp3_validation_report") and self.mp3_validation_report.winfo_exists():
            self.mp3_validation_report.config(state="normal")
            self.mp3_validation_report.delete("1.0", tk.END)
            self.mp3_validation_report.insert("1.0", "Επιλέξτε φάκελο MP3 και πατήστε «Έλεγχος MP3» για αναφορά.")
            self.mp3_validation_report.config(state="disabled")
        if hasattr(self, "btn_validate_mp3") and self.btn_validate_mp3.winfo_exists():
            self.btn_validate_mp3.config(state="disabled")
        if hasattr(self, "btn_gen_db") and self.btn_gen_db.winfo_exists():
            self.btn_gen_db.config(state="disabled")
        if hasattr(self, "assign_inner") and self.assign_inner.winfo_exists():
            for w in self.assign_inner.winfo_children():
                w.destroy()
            tk.Label(
                self.assign_inner,
                text="Tab 2: «Φόρτωση markers από PDF (001…)» ή «Εντοπισμός Markers ■».",
                bg=WHITE,
                fg=TEXT_DARK,
                font=("Arial", 10),
            ).pack(pady=20)
        if hasattr(self, "preview_code_lbl") and self.preview_code_lbl.winfo_exists():
            self.preview_code_lbl.config(text="—")
        if hasattr(self, "preview_text") and self.preview_text.winfo_exists():
            self.preview_text.config(state="normal")
            self.preview_text.delete("1.0", tk.END)
            self.preview_text.insert("1.0", "Κάντε κλικ σε μια γραμμή για να δείτε την αρχή του ύμνου.")
            self.preview_text.config(state="disabled")
        if hasattr(self, "preview_thumbnail_lbl") and self.preview_thumbnail_lbl.winfo_exists():
            self.preview_thumbnail_lbl.config(image="", text="Κάντε κλικ σε γραμμή", compound="center")
        self._preview_photo = None
        self._preview_pdf_path = ""
        self._preview_page = 0
        self._tab5_preview_row_index = -1
        if hasattr(self, "validation_report_text") and self.validation_report_text.winfo_exists():
            self.validation_report_text.config(state="normal")
            self.validation_report_text.delete("1.0", tk.END)
            self.validation_report_text.insert("1.0", "Εκτελέστε «Έλεγχος Βάσης» ή «Select Ready PDF» για νέα αναφορά.")
            self.validation_report_text.config(state="disabled")
        if hasattr(self, "step1_ind") and self.step1_ind.winfo_exists():
            self._pipeline_sync_ui_from_state()
        if hasattr(self, "dbprev") and self.dbprev.winfo_exists():
            self.dbprev.config(state="normal")
            self.dbprev.delete("1.0", tk.END)
            self.dbprev.config(state="disabled")
        self._s("Έγινε επαναφορά. Ξεκινήστε νέο βιβλίο.")

    def _build_tab2(self):
        f = self.tab2
        self._section_header(f, "🔍  Σήμανση PDF — Εντοπισμός & Εφαρμογή Markers")
        card = self._card(f)
        self._label(card, "Original PDF:")
        self.tab2_original_pdf_lbl = tk.Label(card, text="—", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9), anchor="w", justify="left")
        self.tab2_original_pdf_lbl.pack(anchor="w", fill="x", pady=(0, 2))
        self._label(card, "Marked PDF (δημιουργείται μετά τη Σήμανση):")
        self.tab2_marked_pdf_lbl = tk.Label(card, text="—", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9), anchor="w", justify="left")
        self.tab2_marked_pdf_lbl.pack(anchor="w", fill="x", pady=(0, 6))
        # Manual placement viewer (single-app workflow) — recommended path for PDFs without markers
        self._label(card, "✅ Προτεινόμενη σειρά (χωρίς να φύγετε από το app):")
        seq = (
            "1) Φόρτωση/Εμφάνιση PDF στο viewer → 2) Χειροκίνητη τοποθέτηση (κλικ) → "
            "3) Αποθήκευση markers στο PDF (001…) → 4) Φόρτωση markers από PDF (001…) → "
            "5) Δημιουργία PDF με Κουμπιά"
        )
        tk.Label(card, text=seq, bg=WHITE, fg=TEXT_DARK, font=("Arial", 9), wraplength=880, justify="left").pack(
            anchor="w", fill="x", pady=(0, 6)
        )

        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=10)
        self._label(card, "Χειροκίνητη τοποθέτηση markers (001…):")
        hint = (
            "Scroll στο PDF κάτω. Αριστερό κλικ = προσθήκη marker. Δεξί κλικ = διαγραφή. "
            "Ctrl+Z = αναίρεση τελευταίου. Μετά πατήστε «Αποθήκευση markers στο PDF (001…)»."
        )
        tk.Label(card, text=hint, bg=WHITE, fg=TEXT_DARK, font=("Arial", 9), wraplength=880, justify="left").pack(
            anchor="w", fill="x", pady=(0, 6)
        )

        tools = tk.Frame(card, bg=WHITE)
        tools.pack(fill="x", pady=(0, 6))
        _btn(
            tools,
            "💾  Αποθήκευση markers στο PDF (001…)",
            self._tab2_viewer_save_markers_pdf,
            bg=GREEN,
            fg=WHITE,
            width=34,
            pady=6,
        ).pack(side="left")
        self.tab2_manual_status = tk.Label(tools, text="Manual: —", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9, "bold"))
        self.tab2_manual_status.pack(side="left", padx=(10, 0))

        # Keep the viewer height bounded so the next-step buttons below stay visible.
        viewer_outer = tk.Frame(card, bg=WHITE)
        viewer_outer.pack(fill="x", expand=False, pady=(0, 6))
        viewer_outer.grid_rowconfigure(0, weight=0)
        viewer_outer.grid_columnconfigure(0, weight=1)
        self.tab2_canvas = tk.Canvas(
            viewer_outer,
            bg=GREY_LIGHT,
            highlightthickness=0,
            cursor="crosshair",
            height=430,
        )
        self.tab2_vbar = ttk.Scrollbar(viewer_outer, orient="vertical", command=self.tab2_canvas.yview)
        self.tab2_hbar = ttk.Scrollbar(viewer_outer, orient="horizontal", command=self.tab2_canvas.xview)
        self.tab2_canvas.configure(yscrollcommand=self.tab2_vbar.set, xscrollcommand=self.tab2_hbar.set)
        self.tab2_canvas.grid(row=0, column=0, sticky="nsew")
        self.tab2_vbar.grid(row=0, column=1, sticky="ns")
        self.tab2_hbar.grid(row=1, column=0, sticky="ew")

        self._tab2_doc = None
        self._tab2_page_photos = []
        self._tab2_page_layout = {}  # page_no -> (top_y_canvas, height_canvas, sx, sy, width_canvas)
        self._tab2_pix_width = 1
        self._tab2_pix_height = 1
        self._tab2_zoom = 1.0
        self._tab2_selected_idx = -1
        self._tab2_undo_stack = []  # indices (in mrks) of adds

        self.tab2_canvas.bind("<Button-1>", self._tab2_on_left_click)
        self.tab2_canvas.bind("<Button-3>", self._tab2_on_right_click)
        self.tab2_canvas.bind("<MouseWheel>", self._tab2_on_mousewheel)
        self.root.bind_all("<Control-z>", self._tab2_on_ctrl_z, add="+")
        self.tab2_canvas.bind("<Configure>", self._tab2_on_canvas_resize, add="+")

        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=10)
        self._label(card, "Εντοπισμός / Φόρτωση markers (από υπάρχον PDF):")
        btn_row = tk.Frame(card, bg=WHITE)
        btn_row.pack(fill="x", pady=(0, 4))
        _btn(
            btn_row,
            "🔍  Εντοπισμός Markers ■ στο PDF",
            self._detect,
            bg=DARK_RED,
            fg=WHITE,
            width=28,
            pady=6,
        ).pack(side="left", padx=(0, 8))
        _btn(
            btn_row,
            "📥  Φόρτωση markers από PDF (001, 002, …)",
            self._load_markers_from_pdf_numbers,
            bg=GOLD,
            fg=TEXT_DARK,
            width=32,
            pady=6,
        ).pack(side="left")
        self.cnt_lbl = tk.Label(card, text="—", bg=WHITE, fg=DARK_RED, font=("Arial", 14, "bold"))
        self.cnt_lbl.pack(pady=6)

        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=10)
        self._label(card, "Δημιουργία PDF με κουμπιά (από τους markers που έχουν φορτωθεί):")
        _btn(card, "🖨️  Δημιουργία PDF με Κουμπιά", self._apply, big=True).pack(fill="x", pady=6)
        _btn(
            card,
            "✅  Τελικό PDF με κουμπιά (Final)",
            self._generate_final_pdf_with_buttons,
            bg=GREEN,
            fg=WHITE,
            big=True,
        ).pack(fill="x", pady=4)
        self.prog = ttk.Progressbar(card, orient="horizontal", length=400, mode="determinate")
        self.prog.pack(pady=8)
        self.prog_lbl = tk.Label(card, text="", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9))
        self.prog_lbl.pack()

    def _build_tab3(self):
        f = self.tab3
        self._section_header(f, "✅  Έλεγχος & Διαχείριση Κουμπιών")
        card = self._card(f)
        _btn(card, "🔄  Ανανέωση Λίστας", self._refresh_list).pack(anchor="w", pady=4)
        lf = tk.Frame(card, bg=WHITE)
        lf.pack(fill="both", expand=True, pady=6)
        sb = tk.Scrollbar(lf)
        sb.pack(side="right", fill="y")
        self.lst = tk.Listbox(
            lf,
            yscrollcommand=sb.set,
            font=("Courier", 10),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            selectbackground=DARK_RED,
            selectforeground=WHITE,
            relief="flat",
            bd=0,
        )
        self.lst.pack(fill="both", expand=True)
        sb.config(command=self.lst.yview)
        _btn(
            card,
            "⛔  Απενεργοποίηση / Ενεργοποίηση επιλεγμένου",
            self._toggle_keep,
            bg=MID_RED,
        ).pack(fill="x", pady=4)

    def _build_tab4(self):
        f = self.tab4
        self._section_header(f, "📊  Δημιουργία Βάσης Δεδομένων")
        card = self._card(f)
        # Catalog-based book/chapter selection (no arbitrary codes)
        self.tab4_catalog_frame = tk.Frame(card, bg=WHITE)
        self.tab4_catalog_frame.pack(anchor="w", pady=3)
        self._label(self.tab4_catalog_frame, "Βιβλίο & Κεφάλαιο (μόνο από Parameters):")
        cat_row = tk.Frame(self.tab4_catalog_frame, bg=WHITE)
        cat_row.pack(anchor="w", pady=2)
        self.tab4_book_combo = ttk.Combobox(cat_row, font=("Arial", 10), width=35, state="readonly")
        self.tab4_book_combo.pack(side="left", padx=(0, 8))
        self.tab4_book_combo.bind("<<ComboboxSelected>>", self._on_tab4_book_selected)
        self.tab4_chapter_combo = ttk.Combobox(cat_row, font=("Arial", 10), width=20, state="readonly")
        self.tab4_chapter_combo.pack(side="left", padx=(0, 8))
        self.tab4_chapter_combo.bind("<<ComboboxSelected>>", self._on_tab4_chapter_selected)
        self.tab4_catalog_hint = tk.Label(
            self.tab4_catalog_frame,
            text="Φορτώστε το Parameters (Tab 0) πρώτα.",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 9),
        )
        self.tab4_catalog_hint.pack(anchor="w", pady=(2, 0))
        self._label(card, "Bunny Base URL (π.χ. https://fanari.b-cdn.net):")
        tk.Entry(
            card,
            textvariable=self.dv["bunny"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=3)
        self._label(card, "Φάκελος κεφαλαίου (HTML) — για marker matching:")
        html_row = tk.Frame(card, bg=WHITE)
        html_row.pack(fill="x", pady=3)
        self.html_folder_entry = tk.Entry(
            html_row,
            textvariable=self.dv["html_folder"],
            font=("Arial", 9),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        )
        self.html_folder_entry.pack(side="left", fill="x", expand=True)
        _btn(html_row, "Επιλογή φακέλου κεφαλαίου", self._pick_html_folder, width=24).pack(side="right", padx=(6, 0))
        self._label(card, "Φάκελος MP3:")
        mp3_row = tk.Frame(card, bg=WHITE)
        mp3_row.pack(fill="x", pady=3)
        self.mp3_folder_entry = tk.Entry(
            mp3_row,
            textvariable=self.dv["mp3_folder"],
            font=("Arial", 9),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        )
        self.mp3_folder_entry.pack(side="left", fill="x", expand=True)
        self.mp3_folder_entry.bind(
            "<FocusOut>",
            lambda e: self._on_mp3_folder_entry_changed(),
        )
        _btn(mp3_row, "Επιλογή Φακέλου MP3", self._pick_mp3_folder, width=22).pack(side="right", padx=(6, 0))
        self.mp3_files_count_lbl = tk.Label(
            card, text="MP3 Files Found: 0", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9)
        )
        self.mp3_files_count_lbl.pack(anchor="w", pady=(2, 0))
        self._label(card, "Έλεγχος MP3 — Αποτελέσματα:")
        mp3_report_frame = tk.Frame(card, bg=WHITE)
        mp3_report_frame.pack(fill="both", expand=True, pady=3)
        scrollbar_mp3 = ttk.Scrollbar(mp3_report_frame)
        self.mp3_validation_report = tk.Text(
            mp3_report_frame,
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            height=10,
            wrap="word",
            state="disabled",
            yscrollcommand=scrollbar_mp3.set,
        )
        scrollbar_mp3.config(command=self.mp3_validation_report.yview)
        scrollbar_mp3.pack(side="right", fill="y")
        self.mp3_validation_report.pack(side="left", fill="both", expand=True)
        self.mp3_validation_report.insert("1.0", "Επιλέξτε φάκελο MP3 και πατήστε «Έλεγχος MP3» για αναφορά.")
        self.btn_validate_mp3 = _btn(
            card,
            "🔊  Έλεγχος MP3",
            self._validate_mp3,
            bg=GREEN,
            width=18,
        )
        self.btn_validate_mp3.pack(anchor="w", pady=4)
        self.btn_validate_mp3.config(state="disabled")
        self.btn_analyze_markers = _btn(
            card,
            "📎  Ανάλυση Markers (HTML + MP3)",
            self._analyze_markers,
            bg=GREEN,
            width=28,
        )
        self.btn_analyze_markers.pack(anchor="w", pady=4)
        self.btn_stage1_audio = _btn(
            card,
            "✓  Stage 1: PDF markers ↔ 001.mp3 (pre-upload JSON)",
            self._stage1_preupload_audio_validate,
            bg=DARK_RED,
            width=36,
        )
        self.btn_stage1_audio.pack(anchor="w", pady=2)
        self.btn_stage2_audio = _btn(
            card,
            "☁  Stage 2: Bunny URLs (post-upload JSON)",
            self._stage2_postupload_audio_urls,
            bg=DARK_RED,
            width=36,
        )
        self.btn_stage2_audio.pack(anchor="w", pady=2)
        self._label(card, "Detected markers | Matched audio:")
        marker_table_frame = tk.Frame(card, bg=WHITE)
        marker_table_frame.pack(fill="both", expand=True, pady=3)
        marker_scroll_y = ttk.Scrollbar(marker_table_frame)
        marker_scroll_x = ttk.Scrollbar(marker_table_frame, orient="horizontal")
        self.marker_tree = ttk.Treeview(
            marker_table_frame,
            columns=("marker", "mp3", "status"),
            show="headings",
            height=6,
            yscrollcommand=marker_scroll_y.set,
            xscrollcommand=marker_scroll_x.set,
        )
        self.marker_tree.heading("marker", text="Marker")
        self.marker_tree.heading("mp3", text="MP3")
        self.marker_tree.heading("status", text="Status")
        self.marker_tree.column("marker", width=80)
        self.marker_tree.column("mp3", width=180)
        self.marker_tree.column("status", width=100)
        marker_scroll_y.config(command=self.marker_tree.yview)
        marker_scroll_x.config(command=self.marker_tree.xview)
        marker_scroll_y.pack(side="right", fill="y")
        marker_scroll_x.pack(side="bottom", fill="x")
        self.marker_tree.pack(side="left", fill="both", expand=True)
        self.marker_matching_lbl = tk.Label(
            card, text="", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9)
        )
        self.marker_matching_lbl.pack(anchor="w", pady=(2, 0))
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=12)
        self.btn_gen_db = _btn(
            card,
            "📊  Δημιουργία database.xlsx",
            self._gen_db,
            big=True,
        )
        self.btn_gen_db.pack(fill="x", pady=6)
        self.btn_gen_db.config(state="disabled")
        self.dbprev = tk.Text(
            card,
            font=("Courier", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            height=14,
            relief="flat",
            state="disabled",
        )
        self.dbprev.pack(fill="both", expand=True, pady=6)

    def _build_tab5(self):
        f = self.tab5
        self._section_header(f, "🎵  5. Αντιστοίχιση Ύμνων")
        card = self._card(f)
        tk.Label(
            card,
            text="Ορίστε τον τίτλο και προαιρετικά Ήχο/Τμήμα. MP3 Code, File και URL δημιουργούνται αυτόματα.",
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(0, 6))
        _btn(card, "🔄  Ανανέωση", self._refresh_tab5).pack(anchor="w", pady=4)
        # Scrollable area (song buttons / rows)
        canvas_frame = tk.Frame(card, bg=WHITE)
        canvas_frame.pack(fill="both", expand=True, pady=6)
        self.assign_canvas = tk.Canvas(canvas_frame, bg=WHITE, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.assign_canvas.yview)
        self.assign_inner = tk.Frame(self.assign_canvas, bg=WHITE)
        self.assign_inner.bind(
            "<Configure>",
            lambda e: self.assign_canvas.configure(scrollregion=self.assign_canvas.bbox("all")),
        )
        self.assign_canvas.create_window((0, 0), window=self.assign_inner, anchor="nw")
        self.assign_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.assign_canvas.pack(side="left", fill="both", expand=True)
        # Hymn preview panel (below song buttons)
        preview_frame = tk.Frame(card, bg=LIGHT_RED, bd=1, relief="solid")
        preview_frame.pack(fill="x", pady=(8, 4))
        tk.Label(
            preview_frame,
            text="Προεπισκόπηση ύμνου",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w", padx=8, pady=(6, 2))
        # Content row: large thumbnail on left, hymn code + first 1–2 lines on right
        preview_content = tk.Frame(preview_frame, bg=LIGHT_RED)
        preview_content.pack(fill="x", padx=8, pady=(0, 8))
        # Left: large clickable thumbnail
        thumbnail_frame = tk.Frame(preview_content, bg=GREY_LIGHT, bd=1, relief="solid", cursor="hand2")
        thumbnail_frame.pack(side="left", padx=(0, 12), pady=0)
        self.preview_thumbnail_lbl = tk.Label(
            thumbnail_frame,
            text="Κάντε κλικ σε γραμμή",
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            font=("Arial", 9),
            cursor="hand2",
            anchor="center",
        )
        self.preview_thumbnail_lbl.pack(padx=6, pady=6)
        self.preview_thumbnail_lbl.bind("<Button-1>", lambda e: self._open_preview_pdf())
        thumbnail_frame.bind("<Button-1>", lambda e: self._open_preview_pdf())
        self._preview_photo = None
        self._preview_pdf_path = ""
        self._preview_page = 0
        # Right: hymn code + first 1–2 lines (horizontal text)
        right_frame = tk.Frame(preview_content, bg=LIGHT_RED)
        right_frame.pack(side="left", fill="both", expand=True)
        self.preview_code_lbl = tk.Label(
            right_frame,
            text="—",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 12, "bold"),
            anchor="w",
        )
        self.preview_code_lbl.pack(anchor="w", pady=(0, 4))
        self.preview_text = tk.Text(
            right_frame,
            wrap="word",
            height=3,
            font=("Arial", 10),
            bg=LIGHT_RED,
            fg=TEXT_DARK,
            relief="flat",
            state="disabled",
        )
        self.preview_text.pack(anchor="w", fill="x", expand=True)
        self.preview_text.config(state="normal")
        self.preview_text.insert("1.0", "Κάντε κλικ σε μια γραμμή για να δείτε την αρχή του ύμνου.")
        self.preview_text.config(state="disabled")
        self._tab5_preview_row_index = -1

    def _build_tab6(self):
        """Ready PDF + MP3 import — strict 4-step pipeline."""
        f = self.tab6
        self._section_header(f, "📋  6. Εισαγωγή Έτοιμου PDF (Βήμα-Βήμα)")
        card = self._card(f)
        tk.Label(
            card,
            text="Ακολουθήστε τα βήματα με τη σειρά. Κάθε βήμα ενεργοποιείται μόνο όταν το προηγούμενο πετύχει.",
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(0, 8))
        # Step indicators
        step_row = tk.Frame(card, bg=WHITE)
        step_row.pack(fill="x", pady=4)
        self.step1_ind = tk.Label(step_row, text="[1] PDF", bg=GREY_MED, fg=WHITE, font=("Arial", 10, "bold"), padx=8, pady=4)
        self.step1_ind.pack(side="left", padx=2)
        tk.Label(step_row, text="→", bg=WHITE, fg=TEXT_DARK).pack(side="left", padx=2)
        self.step2_ind = tk.Label(step_row, text="[2] MP3", bg=GREY_MED, fg=WHITE, font=("Arial", 10, "bold"), padx=8, pady=4)
        self.step2_ind.pack(side="left", padx=2)
        tk.Label(step_row, text="→", bg=WHITE, fg=TEXT_DARK).pack(side="left", padx=2)
        self.step3_ind = tk.Label(step_row, text="[3] Βιβλίο", bg=GREY_MED, fg=WHITE, font=("Arial", 10, "bold"), padx=8, pady=4)
        self.step3_ind.pack(side="left", padx=2)
        tk.Label(step_row, text="→", bg=WHITE, fg=TEXT_DARK).pack(side="left", padx=2)
        self.step4_ind = tk.Label(step_row, text="[4] Επιβεβαίωση", bg=GREY_MED, fg=WHITE, font=("Arial", 10, "bold"), padx=8, pady=4)
        self.step4_ind.pack(side="left", padx=2)
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=8)
        # Step 1: PDF
        s1 = tk.Frame(card, bg=CREAM, bd=1, relief="solid")
        s1.pack(fill="x", pady=4)
        s1r = tk.Frame(s1, bg=CREAM)
        s1r.pack(fill="x", padx=10, pady=8)
        self.btn_step1 = _btn(s1r, "1. Select Ready PDF", self._pipeline_step1_pdf, bg=GREEN, width=20)
        self.btn_step1.pack(side="left", padx=(0, 12))
        self.step1_status = tk.Label(s1r, text="—", bg=CREAM, fg=TEXT_DARK, font=("Arial", 10))
        self.step1_status.pack(side="left")
        # Step 2: MP3
        s2 = tk.Frame(card, bg=CREAM, bd=1, relief="solid")
        s2.pack(fill="x", pady=4)
        s2r = tk.Frame(s2, bg=CREAM)
        s2r.pack(fill="x", padx=10, pady=8)
        self.btn_step2 = _btn(s2r, "2. Select MP3 Folder", self._pipeline_step2_mp3, width=20)
        self.btn_step2.pack(side="left", padx=(0, 12))
        self.btn_step2.config(state="disabled")
        self.step2_status = tk.Label(s2r, text="— (απαιτείται πρώτα PDF PASS)", bg=CREAM, fg=GREY_MED, font=("Arial", 9))
        self.step2_status.pack(side="left")
        # Step 3: Book + Chapter
        s3 = tk.Frame(card, bg=CREAM, bd=1, relief="solid")
        s3.pack(fill="x", pady=4)
        s3r = tk.Frame(s3, bg=CREAM)
        s3r.pack(fill="x", padx=10, pady=8)
        tk.Label(s3r, text="3. Βιβλίο:", bg=CREAM, fg=TEXT_DARK, font=("Arial", 9)).pack(side="left", padx=(0, 4))
        self.pipeline_book_combo = ttk.Combobox(s3r, font=("Arial", 9), width=30, state="readonly")
        self.pipeline_book_combo.pack(side="left", padx=(0, 12))
        self.pipeline_book_combo.bind("<<ComboboxSelected>>", self._pipeline_on_book_selected)
        tk.Label(s3r, text="Κεφάλαιο:", bg=CREAM, fg=TEXT_DARK, font=("Arial", 9)).pack(side="left", padx=(12, 4))
        self.pipeline_chapter_combo = ttk.Combobox(s3r, font=("Arial", 9), width=22, state="readonly")
        self.pipeline_chapter_combo.pack(side="left", padx=(0, 12))
        self.pipeline_chapter_combo.bind("<<ComboboxSelected>>", self._pipeline_on_chapter_selected)
        self.step3_status = tk.Label(s3r, text="— (απαιτείται πρώτα MP3 PASS)", bg=CREAM, fg=GREY_MED, font=("Arial", 9))
        self.step3_status.pack(side="left", padx=(12, 0))
        # Step 4: Confirm
        s4 = tk.Frame(card, bg=CREAM, bd=1, relief="solid")
        s4.pack(fill="x", pady=4)
        s4r = tk.Frame(s4, bg=CREAM)
        s4r.pack(fill="x", padx=10, pady=8)
        self.btn_step4 = _btn(s4r, "4. Confirm Import", self._pipeline_step4_confirm, bg=GREEN, width=20)
        self.btn_step4.pack(side="left", padx=(0, 12))
        self.btn_step4.config(state="disabled")
        self.step4_status = tk.Label(s4r, text="— (επιλέξτε βιβλίο και κεφάλαιο)", bg=CREAM, fg=GREY_MED, font=("Arial", 9))
        self.step4_status.pack(side="left")
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=8)
        # Report area
        tk.Label(card, text="Αναφορά ελέγχου:", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9, "bold")).pack(anchor="w")
        report_frame = tk.Frame(card, bg=WHITE)
        report_frame.pack(fill="both", expand=True, pady=4)
        scrollbar = ttk.Scrollbar(report_frame)
        self.validation_report_text = tk.Text(
            report_frame,
            wrap="word",
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            relief="flat",
            state="disabled",
            height=12,
            yscrollcommand=scrollbar.set,
        )
        scrollbar.config(command=self.validation_report_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.validation_report_text.pack(side="left", fill="both", expand=True)
        self.validation_report_text.tag_configure("status_ok", foreground=GREEN, font=("Consolas", 9, "bold"))
        self.validation_report_text.tag_configure("status_errors", foreground="#c03030", font=("Consolas", 9, "bold"))
        # Legacy: Database validation button (for full workflow)
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=8)
        _btn(card, "Έλεγχος Βάσης (για πλήρη workflow)", self._run_preflight_validation, width=35).pack(anchor="w", pady=4)

    def _build_tab7(self):
        """Bunny.net Preparation — no upload; prepare and display data for future upload."""
        f = self.tab7
        self._section_header(f, "☁️  7. Bunny.net Preparation")
        card = self._card(f)
        tk.Label(
            card,
            text="Ρυθμίστε τα πεδία για μελλοντική μεταφόρτωση. Δεν γίνεται ακόμα upload.",
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(0, 8))
        self._label(card, "Base CDN URL (e.g. https://fanari.b-cdn.net):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["base_cdn_url"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "Book slug (e.g. anastasimatarion):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["book_slug"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "Chapter code (e.g. AN01):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["chapter_code"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "Root remote folder (default: books):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["root_remote_folder"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "Thinkific Course Name:")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["thinkific_course_name"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "Thinkific Chapter Name:")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["thinkific_chapter_name"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "FlipBuilder Book Name (optional):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["flipbuilder_book_name"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
        ).pack(fill="x", pady=2)
        self._label(card, "API Key (optional, not used yet):")
        tk.Entry(
            card,
            textvariable=self.bunny_prep["api_key"],
            font=("Arial", 10),
            bg=GREY_LIGHT,
            relief="flat",
            bd=5,
            show="*",
        ).pack(fill="x", pady=2)
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=10)
        btn_row = tk.Frame(card, bg=WHITE)
        btn_row.pack(fill="x", pady=4)
        _btn(btn_row, "Preview Bunny URLs", self._preview_bunny_urls, bg=GREEN, width=20).pack(side="left", padx=(0, 8))
        _btn(btn_row, "Export Bunny Upload Manifest", self._export_bunny_manifest, width=26).pack(side="left")
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=12)
        self._section_header(card, "Chapter Preparation (Batch)")
        tk.Label(
            card,
            text="Validate chapters from manifest before upload. Each chapter = work folder (with database.xlsx) + MP3 folder.",
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(0, 4))
        prep_row = tk.Frame(card, bg=WHITE)
        prep_row.pack(fill="x", pady=4)
        _btn(prep_row, "Create Manifest Template", self._create_chapters_manifest_template, bg=GREEN, width=22).pack(side="left", padx=(0, 8))
        _btn(prep_row, "Load Manifest & Run Preparation", self._run_chapter_preparation, width=28).pack(side="left", padx=(0, 8))
        _btn(prep_row, "Export Report", self._export_chapter_preparation_report, width=14).pack(side="left")
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=12)
        self._section_header(card, "Publishing Exports")
        pub_btn_row = tk.Frame(card, bg=WHITE)
        pub_btn_row.pack(fill="x", pady=4)
        _btn(pub_btn_row, "Preview Publishing Paths", self._preview_publishing_paths, bg=GREEN, width=22).pack(side="left", padx=(0, 8))
        _btn(pub_btn_row, "Export Bunny Manifest", self._export_publishing_bunny_manifest, width=20).pack(side="left", padx=(0, 8))
        _btn(pub_btn_row, "Export Thinkific Lesson List", self._export_thinkific_lessons, width=24).pack(side="left", padx=(0, 8))
        _btn(pub_btn_row, "Export FlipBuilder Audio Links", self._export_flipbuilder_links, width=26).pack(side="left")
        self._label(card, "Preview — row number, title, mp3 filename, bunny folder, bunny url, thinkific lesson title:")
        report_frame = tk.Frame(card, bg=WHITE)
        report_frame.pack(fill="both", expand=True, pady=4)
        scrollbar = ttk.Scrollbar(report_frame)
        self.bunny_preview_text = tk.Text(
            report_frame,
            wrap="word",
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            relief="flat",
            state="disabled",
            yscrollcommand=scrollbar.set,
        )
        scrollbar.config(command=self.bunny_preview_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.bunny_preview_text.pack(side="left", fill="both", expand=True)
        self.bunny_preview_text.insert("1.0", "Πατήστε «Preview Bunny URLs» ή «Preview Publishing Paths».")

    def _build_tab8(self):
        """Parameters / Master Catalog — official book definitions. MUST be loaded first."""
        f = self.tab8
        self._section_header(f, "📋  0. Parameters / Master Catalog")
        card = self._card(f)
        tk.Label(
            card,
            text="FIRST STEP: Load the official book catalog. All import operations require book/chapter selection from this catalog. No arbitrary codes.",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w", padx=8, pady=8)
        tk.Label(
            card,
            text="Load Parameters (JSON) or Import from Excel. Parameters = collections + books. No merge, no upload.",
            bg=WHITE,
            fg=TEXT_DARK,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(0, 8))
        load_row = tk.Frame(card, bg=WHITE)
        load_row.pack(fill="x", pady=4)
        _btn(load_row, "Load Parameters", self._load_parameters, bg=GREEN, width=18).pack(side="left", padx=(0, 8))
        _btn(load_row, "Import from Excel", self._load_book_registry, width=18).pack(side="left", padx=(0, 8))
        _btn(load_row, "Build Bookshelf", self._build_bookshelf_from_parameters, width=16).pack(side="left", padx=(0, 8))
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=12)
        self._section_header(card, "Add New Book")
        add_fields = [
            "Book_Code",
            "Book_Title",
            "Book_Slug",
            "Thinkific_Course_Name",
            "Subscription_Group",
            "FlipBuilder_Book_Name",
            "Bookshelf_Name",
            "Bookshelf_Order",
            "Bunny_Root_Folder",
            "Chapter_List",
            "Expected_Chapters_Count",
            "Notes",
        ]
        self.book_registry_add_vars = {k: tk.StringVar() for k in add_fields}
        add_grid = tk.Frame(card, bg=WHITE)
        add_grid.pack(fill="x", pady=4)
        for i, name in enumerate(add_fields):
            lbl_text = name + ":"
            if name == "Chapter_List":
                lbl_text = "Chapter_List (AN01,AN02 or per line):"
            elif name == "Expected_Chapters_Count":
                lbl_text = "Expected_Chapters_Count (optional):"
            lbl = tk.Label(add_grid, text=lbl_text, bg=WHITE, fg=TEXT_DARK, font=("Arial", 9), width=28, anchor="e")
            lbl.grid(row=i, column=0, sticky="e", padx=(0, 4), pady=2)
            ent = tk.Entry(add_grid, textvariable=self.book_registry_add_vars[name], font=("Arial", 10), bg=GREY_LIGHT, width=50)
            ent.grid(row=i, column=1, sticky="ew", padx=4, pady=2)
        add_grid.columnconfigure(1, weight=1)
        add_btn_row = tk.Frame(card, bg=WHITE)
        add_btn_row.pack(fill="x", pady=4)
        _btn(add_btn_row, "Add Book to Registry", self._add_book_to_registry, bg=GREEN, width=20).pack(side="left")
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=12)
        self.book_registry_summary_lbl = tk.Label(
            card,
            text="Catalog: missing",
            bg=LIGHT_RED,
            fg=DARK_RED,
            font=("Arial", 9),
        )
        self.book_registry_summary_lbl.pack(anchor="w", pady=(0, 4))
        self._label(card, "Loaded books (select one):")
        list_frame = tk.Frame(card, bg=WHITE)
        list_frame.pack(fill="both", expand=True, pady=4)
        scrollbar = ttk.Scrollbar(list_frame)
        self.book_registry_listbox = tk.Listbox(
            list_frame,
            font=("Consolas", 10),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            selectmode="single",
            height=8,
            yscrollcommand=scrollbar.set,
        )
        self.book_registry_listbox.bind("<<ListboxSelect>>", self._on_book_registry_select)
        scrollbar.config(command=self.book_registry_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.book_registry_listbox.pack(side="left", fill="both", expand=True)
        self._label(card, "Selected book metadata:")
        self.book_registry_meta_text = tk.Text(
            card,
            wrap="word",
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            height=6,
        )
        self.book_registry_meta_text.pack(fill="x", pady=2)
        self._label(card, "Chapter lookup — enter chapter code (e.g. AN01):")
        lookup_row = tk.Frame(card, bg=WHITE)
        lookup_row.pack(fill="x", pady=2)
        self.book_registry_chapter_var = tk.StringVar()
        tk.Entry(
            lookup_row,
            textvariable=self.book_registry_chapter_var,
            font=("Arial", 10),
            width=12,
            bg=GREY_LIGHT,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            lookup_row,
            text="Lookup",
            command=self._lookup_chapter_in_registry,
            bg=GREEN,
            fg=WHITE,
            font=("Arial", 9, "bold"),
            relief="flat",
            cursor="hand2",
        ).pack(side="left")
        self._label(card, "Preview — matching book for chapter:")
        self.book_registry_preview_text = tk.Text(
            card,
            wrap="word",
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            height=6,
            state="disabled",
        )
        self.book_registry_preview_text.pack(fill="x", pady=2)
        tk.Frame(card, bg=GREY_MED, height=1).pack(fill="x", pady=8)
        self._label(card, "Chapter status (expected vs imported):")
        cmp_btn_row = tk.Frame(card, bg=WHITE)
        cmp_btn_row.pack(anchor="w", pady=4)
        _btn(cmp_btn_row, "Compare imported vs expected", self._compare_book_chapters, bg=GREEN, width=26).pack(side="left", padx=(0, 8))
        self.btn_merge = _btn(cmp_btn_row, "Merge (when complete)", lambda: None, width=20)
        self.btn_merge.pack(side="left")
        self.btn_merge.config(state="disabled")
        self.merge_reason_lbl = tk.Label(cmp_btn_row, text="Select book, Compare, then Merge if allowed.", bg=WHITE, fg=TEXT_DARK, font=("Arial", 9))
        self.merge_reason_lbl.pack(side="left", padx=(8, 0))
        self.book_registry_compare_text = tk.Text(
            card,
            wrap="word",
            font=("Consolas", 9),
            bg=GREY_LIGHT,
            fg=TEXT_DARK,
            height=12,
            state="disabled",
        )
        self.book_registry_compare_text.tag_configure("status_complete", foreground=GREEN, font=("Consolas", 9, "bold"))
        self.book_registry_compare_text.tag_configure("status_incomplete", foreground="#b8860b", font=("Consolas", 9, "bold"))
        self.book_registry_compare_text.tag_configure("status_error", foreground="#c03030", font=("Consolas", 9, "bold"))
        self.book_registry_compare_text.pack(fill="both", expand=True, pady=2)

    def _load_parameters_on_startup(self):
        """Auto-load Parameters (JSON) on startup if file exists."""
        if os.path.isfile(PARAMETERS_DEFAULT_PATH):
            ok, books, _ = load_catalog_from_parameters(PARAMETERS_DEFAULT_PATH)
            if ok and books:
                self.book_registry_books = books
                self.parameters_path = PARAMETERS_DEFAULT_PATH
                self.catalog_source = "parameters"
                self._reload_book_registry_list_from_books()
                self._refresh_tab4_catalog()
                self._s(f"Parameters loaded: {len(books)} books")

    def _load_parameters(self):
        """Load catalog from fixed settings/parameters.json. No file dialog."""
        path = PARAMETERS_DEFAULT_PATH
        print(f"[ToFanari] Load Parameters: resolved path = {path}")  # temporary debug
        if not os.path.isfile(path):
            messagebox.showwarning("Parameters", "parameters.json not found in settings folder")
            self._update_catalog_status_label(missing=True)
            return
        ok, books, errors = load_catalog_from_parameters(path)
        if not ok:
            messagebox.showerror("Parameters", "\n".join(errors))
            return
        coll_count = len(set((b.get("collection") or "").strip() for b in books if (b.get("collection") or "").strip()))
        self.book_registry_books = books
        self.parameters_path = path
        self.book_registry_path = ""
        self.catalog_source = "parameters"
        self._reload_book_registry_list_from_books()
        self._refresh_tab4_catalog()
        self.book_registry_meta_text.delete("1.0", tk.END)
        self.book_registry_preview_text.config(state="normal")
        self.book_registry_preview_text.delete("1.0", tk.END)
        self.book_registry_preview_text.insert("1.0", f"Loaded {len(books)} book(s) from Parameters: {path}")
        self.book_registry_preview_text.config(state="disabled")
        messagebox.showinfo("Parameters", f"Parameters loaded: {len(books)} books, {coll_count} collection(s)")
        self._s(f"Parameters loaded: {len(books)} books")

    def _build_bookshelf_from_parameters(self):
        """Regenerate bookshelf (books-data.js, index.html) from parameters.json."""
        p_path = (self.parameters_path or "").strip() or PARAMETERS_DEFAULT_PATH
        if not os.path.isfile(p_path):
            messagebox.showwarning(
                "Build Bookshelf",
                "Load Parameters first or ensure settings/parameters.json exists.",
            )
            return
        try:
            import subprocess
            r = subprocess.run(
                [sys.executable, os.path.join(os.path.dirname(__file__), "build_bookshelf.py")],
                cwd=os.path.dirname(os.path.abspath(__file__)),
                capture_output=True,
                text=True,
            )
            if r.returncode == 0:
                messagebox.showinfo("Build Bookshelf", "Bookshelf built successfully.")
                self._s("Bookshelf built from Parameters")
            else:
                messagebox.showerror("Build Bookshelf", r.stderr or r.stdout or "Build failed")
        except Exception as e:
            messagebox.showerror("Build Bookshelf", str(e))

    def _update_catalog_status_label(self, missing: bool = False, book_count: int = 0):
        """Update catalog status: 'Catalog: missing' or 'Catalog: loaded — X books'."""
        if not hasattr(self, "book_registry_summary_lbl") or not self.book_registry_summary_lbl.winfo_exists():
            return
        if missing or book_count == 0:
            self.book_registry_summary_lbl.config(text="Catalog: missing", bg=LIGHT_RED, fg=DARK_RED)
        else:
            self.book_registry_summary_lbl.config(
                text=f"Catalog: loaded — {book_count} books",
                bg=CREAM,
                fg=TEXT_DARK,
            )

    def _reload_book_registry_list_from_books(self):
        """Refresh the book listbox from self.book_registry_books (collection-aware display)."""
        self.book_registry_listbox.delete(0, tk.END)
        if hasattr(self, "book_registry_summary_lbl") and self.book_registry_summary_lbl.winfo_exists():
            if not self.book_registry_books:
                self._update_catalog_status_label(missing=True)
            else:
                self._update_catalog_status_label(book_count=len(self.book_registry_books))
        if not self.book_registry_books:
            return
        for b in self.book_registry_books:
            slug = (b.get("Book_Slug") or "").strip()
            title = ((b.get("Book_Title") or "").strip())[:35]
            coll = (b.get("collection") or "").strip()
            if coll:
                line = f"[{coll}] {slug} — {title}"
            else:
                line = f"{slug} — {title}"
            self.book_registry_listbox.insert(tk.END, line)

    def _reload_book_registry_list(self, path: str):
        """Reload the books list from Excel path and refresh listbox."""
        ok, books, errors = load_book_registry(path)
        if not ok:
            return
        self.book_registry_books = books
        self.book_registry_path = path
        self.parameters_path = ""
        self.catalog_source = "excel"
        self._update_catalog_status_label(book_count=len(books))
        self.book_registry_listbox.delete(0, tk.END)
        for b in books:
            code = (b.get("Book_Code") or "").strip().upper()
            title = (b.get("Book_Title", "") or "")[:40]
            slug = (b.get("Book_Slug") or "").strip().upper()
            imported = get_imported_chapter_codes(slug)
            exp_count = b.get("expected_chapters_count", 0) or len(b.get("chapters") or [])
            cmp_result = compare_imported_vs_expected(
                code, b.get("chapters") or [], imported,
                expected_count=exp_count if exp_count else None,
            )
            status = cmp_result.get("status", "—")
            imp = cmp_result.get("imported_count", 0)
            exp = cmp_result.get("expected_count", exp_count)
            line = f"{code} | {slug} | {imp}/{exp} → {status} | {title}"
            self.book_registry_listbox.insert(tk.END, line)

    def _add_book_to_registry(self):
        """Add new book from form to book_registry.xlsx. Always saves to settings/book_registry.xlsx."""
        path = _get_book_registry_path()
        settings_dir = os.path.dirname(path)
        if settings_dir and not os.path.isdir(settings_dir):
            os.makedirs(settings_dir, exist_ok=True)
        book = {
            "Book_Code": (self.book_registry_add_vars["Book_Code"].get() or "").strip().upper(),
            "Book_Title": (self.book_registry_add_vars["Book_Title"].get() or "").strip(),
            "Book_Slug": (self.book_registry_add_vars["Book_Slug"].get() or "").strip().upper(),
            "Thinkific_Course_Name": (self.book_registry_add_vars["Thinkific_Course_Name"].get() or "").strip(),
            "Subscription_Group": (self.book_registry_add_vars["Subscription_Group"].get() or "").strip(),
            "FlipBuilder_Book_Name": (self.book_registry_add_vars["FlipBuilder_Book_Name"].get() or "").strip(),
            "Bookshelf_Name": (self.book_registry_add_vars["Bookshelf_Name"].get() or "").strip(),
            "Bookshelf_Order": (self.book_registry_add_vars["Bookshelf_Order"].get() or "").strip(),
            "Bunny_Root_Folder": (self.book_registry_add_vars["Bunny_Root_Folder"].get() or "").strip(),
            "Chapter_List": (self.book_registry_add_vars["Chapter_List"].get() or "").strip(),
            "Expected_Chapters_Count": (self.book_registry_add_vars["Expected_Chapters_Count"].get() or "").strip(),
            "Is_Active": "",
            "Notes": (self.book_registry_add_vars["Notes"].get() or "").strip(),
        }
        ok, errors = append_book_to_registry(path, book)
        if not ok:
            messagebox.showerror("Add Book", "\n".join(errors))
            return
        self.book_registry_path = path
        self._reload_book_registry_list(path)
        self._refresh_tab4_catalog()
        messagebox.showinfo("Add Book", "Book saved successfully.")
        self._s("Book saved to registry.")

    def _load_book_registry(self):
        """Load book_registry.xlsx via file dialog."""
        path = filedialog.askopenfilename(
            title="Load Book Registry",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        ok, books, errors = load_book_registry(path)
        if not ok:
            messagebox.showerror("Book Registry", "\n".join(errors))
            return
        self._reload_book_registry_list(path)
        self.book_registry_meta_text.delete("1.0", tk.END)
        self.book_registry_preview_text.config(state="normal")
        self.book_registry_preview_text.delete("1.0", tk.END)
        self.book_registry_preview_text.insert("1.0", f"Loaded {len(self.book_registry_books)} book(s) from {path}")
        self.book_registry_preview_text.config(state="disabled")
        self._refresh_tab4_catalog()
        self._s(f"Book registry loaded: {path}")

    def _on_book_registry_select(self, event):
        """Show selected book metadata."""
        sel = self.book_registry_listbox.curselection()
        if not sel or not self.book_registry_books:
            return
        idx = sel[0]
        if idx >= len(self.book_registry_books):
            return
        book = self.book_registry_books[idx]
        lines = []
        for k in ["Book_Code", "Book_Title", "Book_Slug", "Thinkific_Course_Name", "FlipBuilder_Book_Name", "Bunny_Root_Folder", "Chapter_List"]:
            v = book.get(k, "")
            lines.append(f"{k}: {v}")
        self.book_registry_meta_text.delete("1.0", tk.END)
        self.book_registry_meta_text.insert("1.0", "\n".join(lines))

    def _lookup_chapter_in_registry(self):
        """Look up chapter code in registry and show matching book."""
        ch = (self.book_registry_chapter_var.get() or "").strip()
        self.book_registry_preview_text.config(state="normal")
        self.book_registry_preview_text.delete("1.0", tk.END)
        if not ch:
            self.book_registry_preview_text.insert("1.0", "Enter a chapter code (e.g. AN01) and click Lookup.")
            self.book_registry_preview_text.config(state="disabled")
            return
        if not self.book_registry_books:
            self.book_registry_preview_text.insert("1.0", "Load Book Registry first.")
            self.book_registry_preview_text.config(state="disabled")
            return
        book = find_book_for_chapter(ch, self.book_registry_books)
        if not book:
            self.book_registry_preview_text.insert("1.0", f"No book found for chapter code: {ch}")
        else:
            lines = [
                f"Chapter code: {ch}",
                f"Book_Code: {book.get('Book_Code', '')}",
                f"Book_Title: {book.get('Book_Title', '')}",
                f"Book_Slug: {book.get('Book_Slug', '')}",
                f"Thinkific_Course_Name: {book.get('Thinkific_Course_Name', '')}",
                f"FlipBuilder_Book_Name: {book.get('FlipBuilder_Book_Name', '')}",
                f"Bunny_Root_Folder: {book.get('Bunny_Root_Folder', '')}",
            ]
            self.book_registry_preview_text.insert("1.0", "\n".join(lines))
        self.book_registry_preview_text.config(state="disabled")

    def _merge_book_chapters(self):
        """Merge validated chapters (placeholder — merge step not yet implemented)."""
        messagebox.showinfo(
            "Merge",
            "Merge step is not yet implemented.\n\nWhen all expected chapters are imported and validated, this button will merge them for FlipBuilder.",
        )

    def _compare_book_chapters(self):
        """Compare imported vs expected for selected book. Show completeness, missing, status."""
        sel = self.book_registry_listbox.curselection()
        if not sel or not self.book_registry_books:
            self.book_registry_compare_text.config(state="normal")
            self.book_registry_compare_text.delete("1.0", tk.END)
            self.book_registry_compare_text.insert("1.0", "Select a book first, then click Compare.")
            self.book_registry_compare_text.config(state="disabled")
            if hasattr(self, "btn_merge") and self.btn_merge.winfo_exists():
                self.btn_merge.config(state="disabled")
                if hasattr(self, "merge_reason_lbl"):
                    self.merge_reason_lbl.config(text="Select book, Compare first")
            return
        book = self.book_registry_books[sel[0]]
        book_slug = (book.get("Book_Slug") or "").strip()
        book_code = (book.get("Book_Code") or "").strip()
        book_title = (book.get("Book_Title") or "").strip()
        expected = book.get("chapters") or []
        expected_active = [ch for ch in expected if isinstance(ch, ChapterDef) and ch.active]
        exp_count = book.get("expected_chapters_count") or len(expected_active)
        imported = get_imported_chapter_codes(book_slug)
        cmp_result = compare_imported_vs_expected(
            book_code, expected_active, imported, expected_count=exp_count,
        )
        status = cmp_result.get("status", "—")
        imp = cmp_result.get("imported_count", 0)
        exp = cmp_result.get("expected_count", exp_count)
        miss = cmp_result.get("missing", [])
        imp_list = cmp_result.get("imported", [])
        lines = [
            f"BOOK: {book_code} — {book_title}",
            "",
            "Expected chapters: " + ", ".join(c.code for c in expected_active) if expected_active else "Expected chapters: (none)",
            "Imported chapters: " + ", ".join(imp_list) if imp_list else "Imported chapters: (none)",
            "Missing chapters: " + ", ".join(miss) if miss else "Missing chapters: (none)",
            "",
            f"Status: {status}  ({imp}/{exp} imported)",
            "",
        ]
        if cmp_result.get("imported"):
            lines.append("IMPORTED: " + ", ".join(cmp_result["imported"]))
        if cmp_result.get("missing"):
            miss = cmp_result["missing"]
            miss_ord = cmp_result.get("missing_orders", [])
            lines.append("MISSING (highlighted): " + ", ".join(miss))
            if miss_ord:
                lines.append("  → Chapter numbers missing: " + ", ".join(str(o) for o in miss_ord))
        if cmp_result.get("duplicate"):
            lines.append("DUPLICATE: " + ", ".join(cmp_result["duplicate"]))
        if cmp_result.get("invalid"):
            lines.append("INVALID: " + ", ".join(cmp_result["invalid"]))
        lines.append("")
        merge_ok = cmp_result.get("merge_allowed")
        if merge_ok:
            lines.append("MERGE: ALLOWED — all expected chapters imported, validated, no duplicates.")
        else:
            reason = []
            if miss:
                reason.append("missing chapters")
            if cmp_result.get("duplicate"):
                reason.append("duplicates")
            if cmp_result.get("invalid"):
                reason.append("invalid entries")
            lines.append("MERGE: BLOCKED — " + (", ".join(reason) if reason else "fix issues above"))
        if hasattr(self, "btn_merge") and self.btn_merge.winfo_exists():
            if merge_ok:
                self.btn_merge.config(state="normal")
                self.btn_merge.config(command=self._merge_book_chapters)
                if hasattr(self, "merge_reason_lbl"):
                    self.merge_reason_lbl.config(text="Merge allowed (ready for merge step)")
            else:
                self.btn_merge.config(state="disabled")
                self.btn_merge.config(command=lambda: None)
                if hasattr(self, "merge_reason_lbl"):
                    self.merge_reason_lbl.config(text="Merge blocked: " + (", ".join(reason) if reason else "see above"))
        if self.catalog_source == "parameters":
            lines.append("")
            try:
                p_path = (self.parameters_path or "").strip() or PARAMETERS_DEFAULT_PATH
                ok_p, data_p, _ = load_parameters(p_path)
                if ok_p and data_p:
                    colls = get_collection_completeness(data_p, get_imported_chapter_codes)
                    if colls:
                        lines.append("--- COLLECTION COMPLETENESS ---")
                        for c in colls:
                            lines.append(f"  {c['title']}: {c['complete_count']}/{c['total_count']} books → {c['status']}")
            except Exception:
                pass
        content = "\n".join(lines)
        self.book_registry_compare_text.config(state="normal")
        self.book_registry_compare_text.delete("1.0", tk.END)
        self.book_registry_compare_text.insert("1.0", content)
        idx = content.find(f"→ {status}")
        if idx >= 0:
            end = content.find("\n", idx)
            if end < 0:
                end = len(content)
            start_pos = f"1.0+{idx}c"
            end_pos = f"1.0+{end}c"
            tag = "status_complete" if status == "COMPLETE" else "status_incomplete" if status == "INCOMPLETE" else "status_error"
            self.book_registry_compare_text.tag_add(tag, start_pos, end_pos)
        self.book_registry_compare_text.config(state="disabled")

    def _build_bunny_public_url(
        self, base_cdn_url: str, book_slug: str, chapter_code: str, filename: str
    ) -> str:
        """Build future public URL: BaseURL + / + BookSlug + / + ChapterCode + / + filename (URL-encoded)."""
        base = (base_cdn_url or "").strip().rstrip("/")
        slug = (book_slug or "").strip()
        code = (chapter_code or "").strip()
        if not base or not slug or not code or not filename:
            return ""
        return f"{base}/{quote(slug, safe='')}/{quote(code, safe='')}/{quote(filename, safe='')}"

    def _build_bunny_folder_path(
        self, root_remote_folder: str, book_slug: str, chapter_code: str
    ) -> str:
        """Build Bunny folder path: root/book_slug/chapter_code/ (e.g. books/anastasimatarion/AN01/)."""
        root = (root_remote_folder or "").strip().strip("/")
        slug = (book_slug or "").strip().strip("/")
        code = (chapter_code or "").strip().strip("/")
        if not root or not slug or not code:
            return ""
        return f"{root}/{slug}/{code}/"

    def _build_bunny_public_url_publishing(
        self,
        base_cdn_url: str,
        root_remote_folder: str,
        book_slug: str,
        chapter_code: str,
        mp3_filename: str,
    ) -> str:
        """Build Bunny public URL: BaseCDN + / + RootRemoteFolder + / + BookSlug + / + ChapterCode + / + mp3_filename."""
        base = (base_cdn_url or "").strip().rstrip("/")
        root = (root_remote_folder or "").strip().strip("/")
        slug = (book_slug or "").strip().strip("/")
        code = (chapter_code or "").strip().strip("/")
        if not base or not root or not slug or not code or not mp3_filename:
            return ""
        return f"{base}/{quote(root, safe='')}/{quote(slug, safe='')}/{quote(code, safe='')}/{quote(mp3_filename, safe='')}"

    def _validate_bunny_config(
        self,
        base_cdn_url: str,
        book_slug: str,
        chapter_code: str,
        source_mp3_files: List[str],
    ) -> Tuple[bool, List[str]]:
        """Validate Bunny configuration. Returns (ok, list of error messages). No network calls."""
        errors: List[str] = []
        base = (base_cdn_url or "").strip()
        if not base:
            errors.append("Base CDN URL is required.")
        elif not base.lower().startswith("https://"):
            errors.append("Base CDN URL must start with https://")
        slug = (book_slug or "").strip()
        if not slug:
            errors.append("Book slug is required.")
        elif " " in slug:
            errors.append("Book slug must not contain spaces.")
        if not (chapter_code or "").strip():
            errors.append("Chapter code is required.")
        if self.book_registry_books:
            ok, _, cat_errs = validate_chapter_in_catalog(chapter_code, slug, self.book_registry_books)
            if not ok:
                errors.extend(cat_errs or ["Chapter not in catalog or wrong book."])
        if not source_mp3_files:
            errors.append("No source MP3 files. Select MP3 folder in Tab 4.")
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y)) if self.mrks else []
        if ordered and len(source_mp3_files) < len(ordered):
            errors.append(f"Fewer source MP3 files ({len(source_mp3_files)}) than hymn rows ({len(ordered)}).")
        seen = set()
        for f in source_mp3_files:
            if f in seen:
                errors.append(f"Duplicate filename: {f}")
            seen.add(f)
        pattern_res = validate_mp3_filename_pattern(source_mp3_files)
        for w in pattern_res.get("warnings", []):
            errors.append(w)
        return (len(errors) == 0, errors)

    def _get_bunny_manifest_rows(self) -> List[dict]:
        """Build list of dicts: row_number, title, source_mp3_filename, filename (canonical), future_public_url."""
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        book_slug = (self.bunny_prep["book_slug"].get() or "").strip()
        chapter_code = (self.bunny_prep["chapter_code"].get() or "").strip()
        mp3_folder = self.get_mp3_folder()
        source_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y)) if self.mrks else []
        rows = []
        for i in range(max(len(ordered), len(source_files))):
            row_num = i + 1
            title = ""
            if i < len(self.assignments):
                title = (self.assignments[i]["song_title"].get() or "").strip()
            source_mp3_filename = source_files[i] if i < len(source_files) else ""
            canonical_filename = get_mp3_file(chapter_code, row_num) if chapter_code else ""
            future_url = (
                self._build_bunny_public_url(base, book_slug, chapter_code, canonical_filename)
                if canonical_filename
                else ""
            )
            rows.append({
                "row_number": row_num,
                "title": title,
                "source_mp3_filename": source_mp3_filename,
                "filename": canonical_filename,
                "future_public_url": future_url,
            })
        return rows

    def _preview_bunny_urls(self):
        """Validate config, build URLs, display in Tab 7 preview. No upload."""
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        book_slug = (self.bunny_prep["book_slug"].get() or "").strip()
        chapter_code = (self.bunny_prep["chapter_code"].get() or "").strip()
        mp3_folder = self.get_mp3_folder()
        source_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        ok, errs = self._validate_bunny_config(base, book_slug, chapter_code, source_files)
        if not ok:
            messagebox.showerror("Bunny Configuration", "\n".join(errs))
            return
        rows = self._get_bunny_manifest_rows()
        lines = ["Row\tTitle\tSource MP3 Filename\tFuture Public URL", ""]
        for r in rows:
            lines.append(f"{r['row_number']}\t{r['title']}\t{r['source_mp3_filename']}\t{r['future_public_url']}")
        self.bunny_preview_text.config(state="normal")
        self.bunny_preview_text.delete("1.0", tk.END)
        self.bunny_preview_text.insert("1.0", "\n".join(lines))
        self.bunny_preview_text.config(state="disabled")
        self._s("Preview Bunny URLs ready.")

    def _export_bunny_manifest(self):
        """Export manifest with row number, title, filename, future public URL. No upload."""
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        book_slug = (self.bunny_prep["book_slug"].get() or "").strip()
        chapter_code = (self.bunny_prep["chapter_code"].get() or "").strip()
        mp3_folder = self.get_mp3_folder()
        source_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        ok, errs = self._validate_bunny_config(base, book_slug, chapter_code, source_files)
        if not ok:
            messagebox.showerror("Bunny Configuration", "\n".join(errs))
            return
        path = filedialog.asksaveasfilename(
            title="Export Bunny Upload Manifest",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Excel (XLSX)", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        rows = self._get_bunny_manifest_rows()
        fieldnames = ["row_number", "title", "filename", "future_public_url"]
        export_rows = [{"row_number": r["row_number"], "title": r["title"], "filename": r["filename"], "future_public_url": r["future_public_url"]} for r in rows]
        try:
            if path.lower().endswith(".xlsx"):
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror("Export Error", "openpyxl is required for XLSX. Install with: pip install openpyxl")
                    return
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Bunny Upload Manifest"
                for col, name in enumerate(fieldnames, 1):
                    ws.cell(row=1, column=col, value=name)
                for row_idx, r in enumerate(export_rows, 2):
                    for col, key in enumerate(fieldnames, 1):
                        ws.cell(row=row_idx, column=col, value=r.get(key, ""))
                wb.save(path)
            else:
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=fieldnames)
                    w.writeheader()
                    w.writerows(export_rows)
        except OSError as e:
            messagebox.showerror("Export Error", str(e))
            return
        self._s(f"Manifest saved: {path}")
        messagebox.showinfo("Export", f"Bunny upload manifest saved:\n{path}")

    def _create_chapters_manifest_template(self):
        """Create a CSV template for chapters manifest."""
        path = filedialog.asksaveasfilename(
            title="Create Chapters Manifest Template",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        if create_manifest_template(path):
            self._s(f"Template saved: {path}")
            messagebox.showinfo("Template", f"Chapters manifest template saved:\n{path}\n\nEdit the paths and add more rows.")
        else:
            messagebox.showerror("Error", f"Failed to create template at: {path}")

    def _run_chapter_preparation(self):
        """Load chapters manifest, validate each chapter, show report."""
        path = filedialog.askopenfilename(
            title="Load Chapters Manifest",
            filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        root = (self.bunny_prep["root_remote_folder"].get() or "").strip() or BUNNY_ROOT_FOLDER
        manifest_ok, report, manifest_errors = run_chapter_preparation(
            manifest_path=path,
            base_cdn_url=base,
            root_remote_folder=root,
        )
        if not manifest_ok:
            messagebox.showerror("Chapters Manifest", "\n".join(manifest_errors))
            return
        if not report:
            self.bunny_preview_text.config(state="normal")
            self.bunny_preview_text.delete("1.0", tk.END)
            self.bunny_preview_text.insert("1.0", "No chapters in manifest.")
            self.bunny_preview_text.config(state="disabled")
            self._s("Manifest loaded but no chapters.")
            return
        text = format_preparation_report(report, base_cdn_url=base, manifest_warnings=manifest_errors if manifest_errors else None)
        self.bunny_preview_text.config(state="normal")
        self.bunny_preview_text.delete("1.0", tk.END)
        self.bunny_preview_text.insert("1.0", text)
        self.bunny_preview_text.config(state="disabled")
        ready = sum(1 for r in report if r.get("status") == "READY")
        self._s(f"Chapter preparation: {ready}/{len(report)} READY")
        messagebox.showinfo(
            "Chapter Preparation",
            f"Validated {len(report)} chapter(s).\nREADY: {ready}  |  NOT_READY: {len(report) - ready}\n\nReport shown in preview area.",
        )

    def _export_chapter_preparation_report(self):
        """Export the current preview content (chapter preparation report) to file."""
        content = self.bunny_preview_text.get("1.0", tk.END)
        if not content.strip():
            messagebox.showinfo("Export", "No report to export. Run Chapter Preparation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Export Chapter Preparation Report",
            defaultextension=".txt",
            filetypes=[("Text", "*.txt"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            self._s(f"Report exported: {path}")
            messagebox.showinfo("Export", f"Report saved:\n{path}")
        except OSError as e:
            messagebox.showerror("Export Error", str(e))

    def _build_database_rows_for_validation(self) -> List[dict]:
        """Build rows from current state (source of truth) for validation and publishing."""
        code = (self.dv["code"].get() or "").strip() or "BOOK"
        bunny = (self.dv["bunny"].get() or "").strip() or ""
        pdf = self.pdf_path.get().strip()
        mp3_folder = self.get_mp3_folder()
        mp3_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y)) if self.mrks else []
        rows = []
        for i, m in enumerate(ordered):
            if i >= len(self.assignments):
                break
            song_title = (self.assignments[i]["song_title"].get() or "").strip()
            mp3_code = get_mp3_code(code, i + 1)
            mp3_file = mp3_files[i] if i < len(mp3_files) else ""
            url = get_mp3_url(bunny, code, i + 1) if bunny else ""
            preview = ""
            if pdf and os.path.isfile(pdf):
                preview = extract_preview_text(pdf, m)
            rows.append({
                "song_title": song_title,
                "mp3_code": mp3_code,
                "mp3_file": mp3_file,
                "url": url,
                "page": getattr(m, "page", None),
                "y": getattr(m, "y", None),
                "preview_text": preview or "",
            })
        return rows

    def _validate_publishing_export(self) -> Tuple[bool, List[str]]:
        """Validate before publishing export. Returns (ok, errors). Blocks if validation has errors."""
        errors: List[str] = []
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        if not base:
            errors.append("Base CDN URL is required.")
        elif not base.lower().startswith("https://"):
            errors.append("Base CDN URL must start with https://")
        book_slug = (self.bunny_prep["book_slug"].get() or "").strip()
        if not book_slug:
            errors.append("Book slug is required.")
        chapter_code = (self.bunny_prep["chapter_code"].get() or "").strip()
        if not chapter_code:
            errors.append("Chapter code is required.")
        rows = self._build_database_rows_for_validation()
        mp3_folder = self.get_mp3_folder()
        mp3_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        result = run_full_validation(
            rows,
            mp3_folder=mp3_folder,
            mp3_files=mp3_files,
            mp3_count=len(mp3_files),
            pdf_path=self.pdf_path.get().strip(),
        )
        status_kind = result.get("status_kind", "errors")
        if status_kind == "errors":
            errors.append("Database validation has errors. Run Tab 6 Έλεγχος Βάσης and fix errors first.")
        for i, r in enumerate(rows):
            title = (r.get("song_title") or "").strip()
            mp3_file = (r.get("mp3_file") or "").strip()
            if not title:
                errors.append(f"Row {i + 1}: empty lesson title.")
            if not mp3_file:
                errors.append(f"Row {i + 1}: empty MP3 filename.")
        seen = set()
        for r in rows:
            f = (r.get("mp3_file") or "").strip()
            if f and f in seen:
                errors.append(f"Duplicate MP3 filename: {f}")
            if f:
                seen.add(f)
        return (len(errors) == 0, errors)

    def _get_publishing_rows(self) -> List[dict]:
        """Build publishing rows from validated database. One row per hymn."""
        base = (self.bunny_prep["base_cdn_url"].get() or "").strip()
        root = (self.bunny_prep["root_remote_folder"].get() or "").strip() or "books"
        book_slug = (self.bunny_prep["book_slug"].get() or "").strip()
        chapter_code = (self.bunny_prep["chapter_code"].get() or "").strip()
        thinkific_course = (self.bunny_prep["thinkific_course_name"].get() or "").strip()
        thinkific_chapter = (self.bunny_prep["thinkific_chapter_name"].get() or "").strip()
        rows = self._build_database_rows_for_validation()
        out = []
        for i, r in enumerate(rows):
            row_num = i + 1
            title = (r.get("song_title") or "").strip()
            mp3_filename = get_mp3_file(chapter_code, row_num) if chapter_code else ""
            bunny_folder = self._build_bunny_folder_path(root, book_slug, chapter_code)
            bunny_url = self._build_bunny_public_url_publishing(
                base, root, book_slug, chapter_code, mp3_filename
            )
            thinkific_lesson_title = f"{chapter_code} – {title}" if chapter_code and title else title
            out.append({
                "row_number": row_num,
                "chapter_code": chapter_code,
                "title": title,
                "mp3_filename": mp3_filename,
                "bunny_folder": bunny_folder,
                "bunny_public_url": bunny_url,
                "thinkific_lesson_title": thinkific_lesson_title,
                "thinkific_course_name": thinkific_course,
                "thinkific_chapter_name": thinkific_chapter,
                "audio_url": bunny_url,
            })
        return out

    def _preview_publishing_paths(self):
        """Preview publishing paths. Uses validated database. No network."""
        ok, errs = self._validate_publishing_export()
        if not ok:
            messagebox.showerror("Publishing Export", "\n".join(errs))
            return
        rows = self._get_publishing_rows()
        lines = [
            "Row\tTitle\tMP3 Filename\tBunny Folder\tBunny URL\tThinkific Lesson Title",
            "",
        ]
        for r in rows:
            lines.append(
                f"{r['row_number']}\t{r['title']}\t{r['mp3_filename']}\t{r['bunny_folder']}\t{r['bunny_public_url']}\t{r['thinkific_lesson_title']}"
            )
        self.bunny_preview_text.config(state="normal")
        self.bunny_preview_text.delete("1.0", tk.END)
        self.bunny_preview_text.insert("1.0", "\n".join(lines))
        self.bunny_preview_text.config(state="disabled")
        self._s("Preview Publishing Paths ready.")

    def _save_export_file(
        self,
        rows: List[dict],
        fieldnames: List[str],
        default_title: str,
        default_name: str,
    ) -> bool:
        """Save export to CSV or XLSX. Returns True if saved."""
        path = filedialog.asksaveasfilename(
            title=default_title,
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Excel (XLSX)", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return False
        export_rows = [{k: r.get(k, "") for k in fieldnames} for r in rows]
        try:
            if path.lower().endswith(".xlsx"):
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror("Export Error", "openpyxl required for XLSX. pip install openpyxl")
                    return False
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = default_name[:31]
                for col, name in enumerate(fieldnames, 1):
                    ws.cell(row=1, column=col, value=name)
                for row_idx, r in enumerate(export_rows, 2):
                    for col, key in enumerate(fieldnames, 1):
                        ws.cell(row=row_idx, column=col, value=r.get(key, ""))
                wb.save(path)
            else:
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=fieldnames)
                    w.writeheader()
                    w.writerows(export_rows)
        except OSError as e:
            messagebox.showerror("Export Error", str(e))
            return False
        self._s(f"Exported: {path}")
        messagebox.showinfo("Export", f"Saved:\n{path}")
        return True

    def _export_publishing_bunny_manifest(self):
        """Export Bunny manifest: row_number, title, mp3_filename, bunny_folder, bunny_public_url."""
        ok, errs = self._validate_publishing_export()
        if not ok:
            messagebox.showerror("Publishing Export", "\n".join(errs))
            return
        rows = self._get_publishing_rows()
        fieldnames = ["row_number", "title", "mp3_filename", "bunny_folder", "bunny_public_url"]
        self._save_export_file(rows, fieldnames, "Export Bunny Manifest", "Bunny Manifest")

    def _export_thinkific_lessons(self):
        """Export Thinkific lesson list."""
        ok, errs = self._validate_publishing_export()
        if not ok:
            messagebox.showerror("Publishing Export", "\n".join(errs))
            return
        rows = self._get_publishing_rows()
        fieldnames = [
            "row_number",
            "chapter_code",
            "lesson_title",
            "thinkific_course_name",
            "thinkific_chapter_name",
            "audio_url",
            "mp3_filename",
        ]
        thinkific_rows = [
            {
                "row_number": r["row_number"],
                "chapter_code": r["chapter_code"],
                "lesson_title": r["thinkific_lesson_title"],
                "thinkific_course_name": r["thinkific_course_name"],
                "thinkific_chapter_name": r["thinkific_chapter_name"],
                "audio_url": r["audio_url"],
                "mp3_filename": r["mp3_filename"],
            }
            for r in rows
        ]
        self._save_export_file(
            thinkific_rows,
            fieldnames,
            "Export Thinkific Lesson List",
            "Thinkific Lessons",
        )

    def _export_flipbuilder_links(self):
        """Export FlipBuilder audio links."""
        ok, errs = self._validate_publishing_export()
        if not ok:
            messagebox.showerror("Publishing Export", "\n".join(errs))
            return
        rows = self._get_publishing_rows()
        fieldnames = ["row_number", "title", "audio_url", "mp3_filename"]
        flip_rows = [
            {
                "row_number": r["row_number"],
                "title": r["title"],
                "audio_url": r["audio_url"],
                "mp3_filename": r["mp3_filename"],
            }
            for r in rows
        ]
        self._save_export_file(
            flip_rows,
            fieldnames,
            "Export FlipBuilder Audio Links",
            "FlipBuilder Links",
        )

    def _run_preflight_validation(self):
        """Build rows from current state, run full validation, show report in UI."""
        self.reset_mp3_runtime_state()
        rows = self._build_database_rows_for_validation()
        mp3_folder = self.get_mp3_folder()
        mp3_files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        result = run_full_validation(
            rows,
            mp3_folder=mp3_folder,
            mp3_files=mp3_files,
            mp3_count=len(mp3_files),
            pdf_path=self.pdf_path.get().strip(),
        )
        self.validation_report_text.config(state="normal")
        self.validation_report_text.delete("1.0", tk.END)
        report = result.get("report_lines") or []
        status_text = result.get("status_text") or ""
        status_kind = result.get("status_kind") or "errors"
        content = "\n".join(report)
        self.validation_report_text.insert("1.0", content)
        # Color the STATUS line (find it and tag it)
        start = content.find("STATUS:")
        if start >= 0:
            end = content.find("\n", start)
            if end < 0:
                end = len(content)
            line_start = "1.0"
            line_end = "1.0"
            if start > 0:
                line_start = self.validation_report_text.index(f"1.0+{start}c")
            if end > start:
                line_end = self.validation_report_text.index(f"1.0+{end}c")
            tag = "status_ok" if status_kind == "ok" else "status_warnings" if status_kind == "warnings" else "status_errors"
            self.validation_report_text.tag_add(tag, line_start, line_end)
        self.validation_report_text.config(state="disabled")
        self._s("Έλεγχος βάσης ολοκληρώθηκε.")

    def _pipeline_step1_pdf(self):
        """Step 1: Select Ready PDF only. Run marker detection. PASS/FAIL."""
        path = filedialog.askopenfilename(
            title="1. Select Ready PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path or not path.strip():
            return
        self.ready_pdf_validation = validate_ready_pdf(path)
        r = self.ready_pdf_validation
        self.ready_pdf_path = path if r.passed else ""
        self.ready_pipeline_mp3_folder = ""
        self.ready_pipeline_mp3_passed = False
        self._pipeline_update_report(list(r.report_lines))
        if r.passed:
            if hasattr(self, "step1_ind"):
                self.step1_ind.config(bg=GREEN, text="[1] PDF PASS")
            if hasattr(self, "step1_status"):
                self.step1_status.config(text=f"PASS — {r.marker_count} markers", fg=GREEN)
            if hasattr(self, "btn_step2"):
                self.btn_step2.config(state="normal")
            if hasattr(self, "step2_status"):
                self.step2_status.config(text="Επιλέξτε φάκελο MP3", fg=TEXT_DARK)
            self._pipeline_disable_steps_2_3_4()
            if hasattr(self, "btn_step2"):
                self.btn_step2.config(state="normal")  # Step 2 enabled after Step 1 PASS
            self._s("Βήμα 1 OK. Συνεχίστε με Βήμα 2 (MP3).")
        else:
            if hasattr(self, "step1_ind"):
                self.step1_ind.config(bg="#c03030", text="[1] PDF FAIL")
            if hasattr(self, "step1_status"):
                self.step1_status.config(text="FAIL — διορθώστε το PDF", fg="#c03030")
            self._pipeline_disable_steps_2_3_4()
            self._s("Βήμα 1 απέτυχε.")
            messagebox.showerror("Βήμα 1 — FAIL", "Το PDF δεν περνά τον έλεγχο:\n\n" + "\n".join(r.errors))

    def _pipeline_reload_catalog_if_empty(self):
        """Reload catalog from parameters.json if book_registry_books empty."""
        if self.book_registry_books:
            return
        path = (self.parameters_path or "").strip() or PARAMETERS_DEFAULT_PATH
        if not path or not os.path.isfile(path):
            return
        try:
            ok, books, _ = load_catalog_from_parameters(path)
            if ok and books:
                self.book_registry_books = books
                self.parameters_path = path
                self.catalog_source = "parameters"
                self._reload_book_registry_list_from_books()
                self._refresh_tab4_catalog()
                print("[Step3] Reloaded catalog from parameters (was empty)")
        except Exception as e:
            print(f"[Step3] Could not reload catalog: {e}")

    def _pipeline_disable_steps_2_3_4(self):
        """Disable steps 2–4 and reset their UI. Does NOT clear ready_pipeline_book_slug/chapter_code."""
        if hasattr(self, "btn_step2"):
            self.btn_step2.config(state="disabled")
        if hasattr(self, "btn_step4"):
            self.btn_step4.config(state="disabled")
        for lbl in ["step2_ind", "step3_ind", "step4_ind"]:
            if hasattr(self, lbl):
                w = getattr(self, lbl)
                w.config(bg=GREY_MED)
                if lbl == "step2_ind":
                    w.config(text="[2] MP3")
                elif lbl == "step3_ind":
                    w.config(text="[3] Βιβλίο")
                else:
                    w.config(text="[4] Επιβεβαίωση")
        if hasattr(self, "step2_status"):
            self.step2_status.config(text="—", fg=GREY_MED)
        if hasattr(self, "step3_status"):
            self.step3_status.config(text="—", fg=GREY_MED)
        if hasattr(self, "step4_status"):
            self.step4_status.config(text="—", fg=GREY_MED)
        if hasattr(self, "pipeline_book_combo"):
            self.pipeline_book_combo.set("")
            self.pipeline_chapter_combo.set("")
            self.pipeline_book_combo["values"] = []
            self.pipeline_chapter_combo["values"] = []
            self.pipeline_book_combo.config(state="disabled")
            self.pipeline_chapter_combo.config(state="disabled")

    def _pipeline_step2_mp3(self):
        """Step 2: Select MP3 folder. Validate vs markers. PASS/FAIL."""
        if not self.ready_pdf_path or not self.ready_pdf_validation or not self.ready_pdf_validation.passed:
            messagebox.showerror("Βήμα 2", "Ολοκληρώστε πρώτα το Βήμα 1 (PDF PASS).")
            return
        folder = filedialog.askdirectory(title="2. Select MP3 Folder")
        if not folder or not os.path.isdir(folder):
            return
        self.ready_pipeline_mp3_folder = folder
        self.source_mp3_folder = folder
        self.dv["mp3_folder"].set(folder)
        marker_count = len(self.ready_pdf_validation.markers)
        mp3_id_to_file, _ = parse_mp3_folder(folder)
        expected_ids = {f"{(i + 1):03d}" for i in range(marker_count)}
        found_ids = set(mp3_id_to_file.keys())
        matched = expected_ids & found_ids
        missing = sorted(expected_ids - found_ids)
        extra = sorted(found_ids - expected_ids)
        passed = len(missing) == 0
        self.ready_pipeline_mp3_passed = passed
        lines = [
            "--- ΒΗΜΑ 2: MP3 VALIDATION ---",
            f"Markers expected: {marker_count}",
            f"MP3 count: {len(mp3_id_to_file)}",
            f"Matched: {len(matched)}/{marker_count}",
        ]
        if missing:
            lines.append(f"Missing: {', '.join(missing[:15])}{'...' if len(missing) > 15 else ''}")
        if extra:
            lines.append(f"Extra: {', '.join(extra[:15])}{'...' if len(extra) > 15 else ''}")
        lines.append("")
        lines.append(f"RESULT: {'PASS' if passed else 'FAIL'}")
        self._pipeline_update_report(lines)
        if passed:
            self.step2_ind.config(bg=GREEN, text="[2] MP3 PASS")
            self.step2_status.config(text=f"PASS — {len(matched)} matched", fg=GREEN)
            self._pipeline_enable_step3()
            self._s("Βήμα 2 OK. Επιλέξτε βιβλίο και κεφάλαιο.")
        else:
            self.step2_ind.config(bg="#c03030", text="[2] MP3 FAIL")
            self.step2_status.config(text="FAIL — missing or extra files", fg="#c03030")
            self.btn_step4.config(state="disabled")
            self.step3_ind.config(bg=GREY_MED, text="[3] Βιβλίο")
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            self._s("Βήμα 2 απέτυχε.")

    def _pipeline_enable_step3(self):
        """Enable Step 3 book/chapter selection after Step 2 PASS. Restores persisted state."""
        self.step3_ind.config(bg=GOLD, text="[3] Βιβλίο")
        if not self.book_registry_books:
            self._pipeline_reload_catalog_if_empty()
        if not self.book_registry_books:
            self.step3_status.config(text="Φορτώστε Parameters (Tab 0)", fg="#c03030")
            return
        books_display = []
        for b in self.book_registry_books:
            slug = (b.get("Book_Slug") or "").strip().upper()
            title = ((b.get("Book_Title") or "").strip())[:30]
            if slug:
                books_display.append(f"{slug} — {title}")
        self.pipeline_book_combo["values"] = books_display
        self.pipeline_chapter_combo.config(state="readonly")
        self.pipeline_book_combo.config(state="readonly")
        if not books_display:
            self.step3_status.config(text="Επιλέξτε βιβλίο και κεφάλαιο", fg=TEXT_DARK)
            return
        # Restore persisted book/chapter if valid
        slug_norm = (self.ready_pipeline_book_slug or "").strip().upper()
        ch_norm = (self.ready_pipeline_chapter_code or "").strip().upper()
        book_idx = 0
        if slug_norm:
            for i, b in enumerate(self.book_registry_books):
                if ((b.get("Book_Slug") or "").strip().upper()) == slug_norm:
                    book_idx = i
                    break
        self.pipeline_book_combo.current(book_idx)
        self._pipeline_on_book_selected(None, restore_chapter=bool(ch_norm))
        self.step3_status.config(text="Επιλέξτε βιβλίο και κεφάλαιο", fg=TEXT_DARK)

    def _pipeline_on_book_selected(self, event, restore_chapter=False):
        """Update chapter combo when book selected. restore_chapter=True preserves selected chapter from state."""
        if not self.book_registry_books or not hasattr(self, "pipeline_book_combo"):
            return
        sel = self.pipeline_book_combo.current()
        if sel < 0:
            return
        book = self.book_registry_books[sel]
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
        book_code = (book.get("Book_Code") or book.get("Book_Slug") or "").strip().upper()
        ch_codes = [(c.code or "").strip().upper() for c in chapters if (c.code or "").strip()]
        print(f"[Step3] Selected book: {book_code} | Available chapter codes: {ch_codes} | restore_chapter={restore_chapter}")
        ch_display = [f"{(c.code or '').upper()} — {c.title or ''}" for c in chapters]
        self.pipeline_chapter_combo["values"] = ch_display
        self.ready_pipeline_book_slug = (book.get("Book_Slug") or "").strip().upper()
        if restore_chapter and self.ready_pipeline_chapter_code:
            ch_norm = (self.ready_pipeline_chapter_code or "").strip().upper()
            ch_idx = next((i for i, c in enumerate(chapters) if (c.code or "").strip().upper() == ch_norm), -1)
            if ch_idx >= 0:
                self.pipeline_chapter_combo.current(ch_idx)
                self.pipeline_chapter_combo.set(ch_display[ch_idx])
                print(f"[Step3] Restored chapter: {self.ready_pipeline_chapter_code}")
            else:
                self.pipeline_chapter_combo.set("")
                self.ready_pipeline_chapter_code = ""
        else:
            self.pipeline_chapter_combo.set("")
            self.ready_pipeline_chapter_code = ""
        self._pipeline_update_step4_state()

    def _pipeline_on_chapter_selected(self, event):
        """Set chapter when selected."""
        if not self.book_registry_books:
            return
        book_sel = self.pipeline_book_combo.current()
        ch_sel = self.pipeline_chapter_combo.current()
        if book_sel < 0 or ch_sel < 0:
            return
        book = self.book_registry_books[book_sel]
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
        if ch_sel < len(chapters):
            self.ready_pipeline_chapter_code = (chapters[ch_sel].code or "").strip().upper()
            self.ready_pipeline_book_slug = (book.get("Book_Slug") or "").strip().upper()
            print(f"[Step3] Selected chapter: {self.ready_pipeline_chapter_code}")
        self._pipeline_update_step4_state()

    def _pipeline_update_step4_state(self):
        """Enable Step 4 if book+chapter selected and no duplicate."""
        ch = self.ready_pipeline_chapter_code
        slug = self.ready_pipeline_book_slug
        if not slug:
            self.btn_step4.config(state="disabled")
            self.step4_status.config(text="Επιλέξτε βιβλίο και κεφάλαιο", fg=GREY_MED)
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            return
        # Get current book's chapters for validation
        book = next((b for b in (self.book_registry_books or []) if (b.get("Book_Slug") or "").strip().upper() == (slug or "").strip().upper()), None)
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active] if book else []
        if not chapters:
            self.btn_step4.config(state="disabled")
            self.step4_status.config(text="Δεν βρέθηκαν ορισμοί κεφαλαίων για το επιλεγμένο βιβλίο", fg="#c03030")
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            print("[Step3] Duplicate check: no chapter definitions for book")
            return
        if not ch:
            self.btn_step4.config(state="disabled")
            self.step4_status.config(text="Επιλέξτε βιβλίο και κεφάλαιο", fg=GREY_MED)
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            return
        # Duplicate check only for valid chapter that exists in catalog
        ch_norm = (ch or "").strip().upper()
        valid_codes = {(c.code or "").strip().upper() for c in chapters}
        if ch_norm not in valid_codes:
            self.btn_step4.config(state="disabled")
            self.step4_status.config(text="Μη έγκυρη επιλογή κεφαλαίου", fg="#c03030")
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            print(f"[Step3] Duplicate check: invalid chapter '{ch_norm}' not in {valid_codes}")
            return
        already = get_imported_chapter_codes((slug or "").strip().upper())
        is_duplicate = ch_norm in already
        print(f"[Step3] Duplicate check: slug={slug} ch={ch_norm} already={already} -> duplicate={is_duplicate}")
        if is_duplicate:
            self.btn_step4.config(state="disabled")
            self.step4_status.config(text="Διπλότυπο κεφάλαιο", fg="#c03030")
            self.step4_ind.config(bg=GREY_MED, text="[4] Επιβεβαίωση")
            return
        self.btn_step4.config(state="normal")
        self.step4_status.config(text="Έτοιμο για εισαγωγή", fg=GREEN)
        self.step4_ind.config(bg=GOLD, text="[4] Επιβεβαίωση")

    def _pipeline_step4_confirm(self):
        """Step 4: Confirm Import. Save association, update registry, log."""
        if not all([
            self.ready_pdf_path,
            self.ready_pdf_validation and self.ready_pdf_validation.passed,
            self.ready_pipeline_mp3_passed,
            self.ready_pipeline_mp3_folder,
            self.ready_pipeline_book_slug,
            self.ready_pipeline_chapter_code,
        ]):
            messagebox.showerror("Επιβεβαίωση", "Ολοκληρώστε όλα τα προηγούμενα βήματα.")
            return
        book_slug = self.ready_pipeline_book_slug
        chapter_code = self.ready_pipeline_chapter_code
        slug_norm = (book_slug or "").strip().upper()
        ch_norm = (chapter_code or "").strip().upper()
        print(f"BOOK CODE USED: {slug_norm} / {ch_norm}")
        book = next((b for b in self.book_registry_books if (b.get("Book_Slug") or "").strip().upper() == slug_norm), None)
        if not book:
            messagebox.showerror("Σφάλμα", "Βιβλίο δεν βρέθηκε.")
            return
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
        ch_def = next((c for c in chapters if (c.code or "").upper() == ch_norm), None)
        if not ch_def:
            messagebox.showerror("Σφάλμα", "Κεφάλαιο δεν βρέθηκε.")
            return
        chapter_order = ch_def.order
        self.source_mp3_folder = self.ready_pipeline_mp3_folder
        self.dv["mp3_folder"].set(self.ready_pipeline_mp3_folder)
        self._register_ready_pdf(
            self.ready_pdf_path,
            self.ready_pdf_validation,
            ch_norm,
            slug_norm,
        )
        add_imported_chapter(slug_norm, ch_norm, chapter_order, self.ready_pdf_path)
        n_markers = len(self.ready_pdf_validation.markers)
        mp3_files = list_mp3_files_in_folder(self.ready_pipeline_mp3_folder)
        try:
            log_import(slug_norm, ch_norm, n_markers, len(mp3_files))
        except Exception:
            pass
        self.step4_ind.config(bg=GREEN, text="[4] DONE")
        self.step4_status.config(text="Εισαγωγή ολοκληρώθηκε", fg=GREEN)
        self._s("Εισαγωγή ολοκληρώθηκε.")
        messagebox.showinfo(
            "Εισαγωγή ολοκληρώθηκε",
            f"Κεφάλαιο {ch_norm} εισήχθη επιτυχώς.\nMarkers: {n_markers}\nMP3: {len(mp3_files)} αρχεία",
        )
        self._pipeline_reset_for_next()

    def _pipeline_reset_for_next(self):
        """Reset pipeline for next import (keep Step 1 if PDF still valid)."""
        self.ready_pipeline_mp3_folder = ""
        self.ready_pipeline_mp3_passed = False
        self.ready_pipeline_book_slug = ""
        self.ready_pipeline_chapter_code = ""
        self._pipeline_disable_steps_2_3_4()
        if self.ready_pdf_path:
            if hasattr(self, "btn_step2"):
                self.btn_step2.config(state="normal")
            if hasattr(self, "step2_status"):
                self.step2_status.config(text="Επιλέξτε φάκελο MP3", fg=TEXT_DARK)

    def _pipeline_update_report(self, lines):
        """Update validation report text."""
        self.validation_report_text.config(state="normal")
        self.validation_report_text.delete("1.0", tk.END)
        self.validation_report_text.insert("1.0", "\n".join(lines))
        self.validation_report_text.config(state="disabled")

    def _pick_ready_pdf(self):
        """Legacy alias for pipeline step 1."""
        self._pipeline_step1_pdf()

    def _associate_ready_pdf(self):
        """Open associate dialog when user has a validated ready PDF but has not yet registered it."""
        if not self.ready_pdf_path or not self.ready_pdf_validation or not self.ready_pdf_validation.passed:
            messagebox.showinfo(
                "Σύνδεση με βιβλίο",
                "Επιλέξτε πρώτα ένα έτοιμο PDF με το κουμπί «Επιλογή Έτοιμου PDF» και βεβαιωθείτε ότι περνά τον έλεγχο.",
            )
            return
        self._show_associate_book_dialog(self.ready_pdf_path, self.ready_pdf_validation)

    def _show_associate_book_dialog(self, pdf_path: str, validation: ReadyPdfValidationResult):
        """Dialog to associate ready PDF with predefined book and chapter from Parameters."""
        if not self.book_registry_books:
            messagebox.showerror(
                "Κατάλογος απαιτείται",
                "Φορτώστε πρώτα το Parameters (Tab 0).\n\nΔεν επιτρέπεται εισαγωγή χωρίς προκαθορισμένο κατάλογο βιβλίων.",
            )
            return
        dlg = tk.Toplevel(self.root)
        dlg.title("Σύνδεση με βιβλίο / κεφάλαιο")
        dlg.geometry("500x400")
        dlg.transient(self.root)
        dlg.grab_set()
        frame = tk.Frame(dlg, bg=CREAM, padx=20, pady=16)
        frame.pack(fill="both", expand=True)
        tk.Label(
            frame,
            text="Επιλέξτε βιβλίο και κεφάλαιο από τις προκαθορισμένες παραμέτρους:",
            bg=CREAM,
            fg=TEXT_DARK,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w", pady=(0, 8))
        tk.Label(frame, text="1. Βιβλίο:", bg=CREAM, fg=TEXT_DARK, font=("Arial", 9, "bold")).pack(anchor="w")
        book_list_frame = tk.Frame(frame, bg=CREAM)
        book_list_frame.pack(fill="x", pady=2)
        book_lb = tk.Listbox(book_list_frame, height=5, font=("Arial", 10), bg=GREY_LIGHT, selectmode="single")
        book_lb.pack(side="left", fill="x", expand=True)
        book_scroll = ttk.Scrollbar(book_list_frame, command=book_lb.yview)
        book_scroll.pack(side="right", fill="y")
        book_lb.config(yscrollcommand=book_scroll.set)
        for b in self.book_registry_books:
            slug = ((b.get("Book_Slug") or "").strip()).upper()
            title = (b.get("Book_Title", "") or "")[:40]
            book_lb.insert(tk.END, f"{slug}  —  {title}")
        tk.Label(frame, text="2. Κεφάλαιο:", bg=CREAM, fg=TEXT_DARK, font=("Arial", 9, "bold")).pack(anchor="w", pady=(10, 0))
        chapter_list_frame = tk.Frame(frame, bg=CREAM)
        chapter_list_frame.pack(fill="x", pady=2)
        chapter_lb = tk.Listbox(chapter_list_frame, height=5, font=("Arial", 10), bg=GREY_LIGHT, selectmode="single")
        chapter_lb.pack(side="left", fill="x", expand=True)
        chapter_scroll = ttk.Scrollbar(chapter_list_frame, command=chapter_lb.yview)
        chapter_scroll.pack(side="right", fill="y")
        chapter_lb.config(yscrollcommand=chapter_scroll.set)

        def on_book_select(event):
            chapter_lb.delete(0, tk.END)
            sel = book_lb.curselection()
            if not sel:
                return
            book = self.book_registry_books[sel[0]]
            chapters = book.get("chapters") or []
            for ch in chapters:
                if isinstance(ch, ChapterDef):
                    if ch.active:
                        chapter_lb.insert(tk.END, f"{(ch.code or '').upper()}  —  {ch.title or '(χωρίς τίτλο)'}")
                else:
                    chapter_lb.insert(tk.END, str(ch))

        book_lb.bind("<<ListboxSelect>>", on_book_select)
        if self.book_registry_books:
            book_lb.selection_set(0)
            on_book_select(None)

        def on_ok():
            book_sel = book_lb.curselection()
            ch_sel = chapter_lb.curselection()
            if not book_sel or not ch_sel:
                messagebox.showerror("Σφάλμα", "Επιλέξτε βιβλίο και κεφάλαιο.", parent=dlg)
                return
            book = self.book_registry_books[book_sel[0]]
            book_slug = (book.get("Book_Slug") or "").strip().upper()
            book_code = (book.get("Book_Code") or "").strip().upper()
            chapters = book.get("chapters") or []
            active_chapters = [c for c in chapters if isinstance(c, ChapterDef) and c.active]
            if ch_sel[0] >= len(active_chapters):
                messagebox.showerror("Σφάλμα", "Μη έγκυρη επιλογή κεφαλαίου.", parent=dlg)
                return
            ch_def = active_chapters[ch_sel[0]]
            chapter_code = (ch_def.code or "").strip().upper()
            chapter_order = ch_def.order
            already = get_imported_chapter_codes(book_slug)
            if chapter_code in already:
                messagebox.showerror(
                    "Διπλότυπο",
                    f"Το κεφάλαιο {chapter_code} είναι ήδη εισαγμένο για αυτό το βιβλίο.\nΔεν επιτρέπονται διπλότυπα.",
                    parent=dlg,
                )
                return
            ok_val, book_val, errs = validate_chapter_in_catalog(chapter_code, book_slug, self.book_registry_books)
            if not ok_val:
                messagebox.showerror(
                    "Σφάλμα",
                    "Το κεφάλαιο δεν ανήκει στο επιλεγμένο βιβλίο:\n" + "\n".join(errs or []),
                    parent=dlg,
                )
                return
            warns = []
            mc = len(validation.markers)
            if mc < 2:
                warns.append(f"Πολύ λίγα markers ({mc}). Ελέγξτε το PDF.")
            elif mc > 400:
                warns.append(f"Πολλά markers ({mc}). Βεβαιωθείτε ότι είναι σωστό κεφάλαιο.")
            mp3_folder = self.get_mp3_folder()
            html_folder = self.html_chapter_folder or (self.dv["html_folder"].get() or "").strip()
            if mp3_folder and html_folder and validation.markers:
                try:
                    from marker_matching import match_markers_to_mp3
                    r = match_markers_to_mp3(html_folder, mp3_folder)
                    if r.markers_without_audio or r.audio_without_marker:
                        warns.append("MP3 ↔ Markers: υπάρχουν markers χωρίς ήχο ή αρχεία χωρίς marker.")
                except Exception:
                    pass
            if warns:
                if not messagebox.askyesno("Προειδοποίηση", "Προειδοποιήσεις:\n\n" + "\n".join(warns) + "\n\nΣυνεχίζετε;", parent=dlg):
                    return
            self._register_ready_pdf(pdf_path, validation, chapter_code, book_slug)
            add_imported_chapter(book_slug, chapter_code, chapter_order, pdf_path)
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btn_frame = tk.Frame(frame, bg=CREAM)
        btn_frame.pack(fill="x", pady=(16, 0))
        _btn(btn_frame, "OK", on_ok, bg=GREEN).pack(side="left", padx=(0, 8))
        _btn(btn_frame, "Ακύρωση", on_cancel).pack(side="left")
        dlg.protocol("WM_DELETE_WINDOW", on_cancel)

    def _register_ready_pdf(
        self,
        pdf_path: str,
        validation: ReadyPdfValidationResult,
        chapter_code: str,
        book_slug: str,
    ):
        """Register validated ready PDF into the pipeline: set pdf_path, mrks, work_folder, etc."""
        self.original_pdf_path = pdf_path
        self.work_folder = os.path.dirname(pdf_path)
        self.pdf_path.set(pdf_path)
        self.mrks = validation.markers
        self._init_assignments()
        code_norm = (chapter_code or "").strip().upper()
        slug_norm = (book_slug or "").strip().upper()
        self.dv["code"].set(code_norm)
        self.dv["book_slug"].set(slug_norm or "")
        self.bunny_prep["chapter_code"].set(code_norm)
        if slug_norm:
            self.bunny_prep["book_slug"].set(slug_norm)
        if hasattr(self, "tab2_original_pdf_lbl") and self.tab2_original_pdf_lbl.winfo_exists():
            self.tab2_original_pdf_lbl.config(text=pdf_path)
        if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
            n = len(self.mrks)
            self.cnt_lbl.config(text=f"✅ {n} markers (από έτοιμο PDF)", fg=GREEN)
        if hasattr(self, "lst") and self.lst.winfo_exists():
            self._refresh_list()
        self._update_mp3_buttons_state(len(self.get_current_mp3_files()) > 0)
        n_markers = len(self.mrks)
        mp3_files = self.get_current_mp3_files()
        mp3_matched = len([f for f in mp3_files if f]) if mp3_files else 0
        try:
            log_import(book_slug or chapter_code, chapter_code, n_markers, mp3_matched)
        except Exception:
            pass
        self._s(f"Έτοιμο PDF εγγράφηκε: {chapter_code} ({n_markers} markers)")
        messagebox.showinfo(
            "Έτοιμο PDF εγγράφηκε",
            f"Το PDF εγγράφηκε ως κεφάλαιο {chapter_code}.\n"
            f"Markers: {len(self.mrks)}\n\n"
            "Μπορείτε να συνεχίσετε με:\n"
            "• Tab 4: Επιλογή MP3, Δημιουργία database.xlsx\n"
            "• Tab 6: Έλεγχος Βάσης\n"
            "• Tab 7: Bunny Preparation",
        )

    def _refresh_tab4_catalog(self):
        """Populate Tab 4 book/chapter dropdowns from Parameters catalog."""
        if not hasattr(self, "tab4_book_combo") or not self.tab4_book_combo.winfo_exists():
            return
        if not self.book_registry_books:
            self.tab4_book_combo["values"] = []
            self.tab4_book_combo.set("")
            self.tab4_chapter_combo["values"] = []
            self.tab4_chapter_combo.set("")
            self.tab4_catalog_hint.config(text="Φορτώστε το Parameters (Tab 0) πρώτα.", bg=LIGHT_RED, fg=DARK_RED)
            return
        books_display = []
        for b in self.book_registry_books:
            slug = (b.get("Book_Slug") or "").strip().upper()
            title = ((b.get("Book_Title") or "").strip())[:35]
            if slug:
                books_display.append(f"{slug} — {title}")
        self.tab4_book_combo["values"] = books_display
        self.tab4_catalog_hint.config(text="Επιλέξτε βιβλίο και κεφάλαιο από την λίστα.", bg=CREAM, fg=TEXT_DARK)
        # Preserve selection if valid (uppercase comparison)
        cur_slug = (self.dv["book_slug"].get() or "").strip().upper()
        cur_code = (self.dv["code"].get() or "").strip().upper()
        if cur_slug and cur_code:
            for i, b in enumerate(self.book_registry_books):
                if ((b.get("Book_Slug") or "").strip().upper()) == cur_slug:
                    chs = [c for c in (b.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
                    if any((c.code or "").upper() == cur_code for c in chs):
                        self.tab4_book_combo.current(i)
                        self._on_tab4_book_selected(None)
                        for j, c in enumerate(chs):
                            if (c.code or "").upper() == cur_code:
                                self.tab4_chapter_combo.current(j)
                                break
                        return
        if books_display:
            self.tab4_book_combo.current(0)
            self._on_tab4_book_selected(None)

    def _on_tab4_book_selected(self, event):
        """Update chapter combo when book changes."""
        if not self.book_registry_books or not hasattr(self, "tab4_book_combo"):
            return
        sel = self.tab4_book_combo.current()
        if sel < 0:
            return
        book = self.book_registry_books[sel]
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
        ch_display = [f"{(c.code or '').upper()} — {c.title or ''}" for c in chapters]
        self.tab4_chapter_combo["values"] = ch_display
        if ch_display:
            self.tab4_chapter_combo.current(0)
            self._on_tab4_chapter_selected(None)
        else:
            self.tab4_chapter_combo.set("")
            self.dv["code"].set("")
            self.dv["book_slug"].set(((book.get("Book_Slug") or "").strip()).upper())

    def _on_tab4_chapter_selected(self, event):
        """Set dv code and book_slug when chapter selected."""
        if not self.book_registry_books or not hasattr(self, "tab4_chapter_combo"):
            return
        book_sel = self.tab4_book_combo.current()
        ch_sel = self.tab4_chapter_combo.current()
        if book_sel < 0 or ch_sel < 0:
            return
        book = self.book_registry_books[book_sel]
        chapters = [c for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
        if ch_sel >= len(chapters):
            return
        ch_def = chapters[ch_sel]
        code_norm = (ch_def.code or "").strip().upper()
        slug_norm = (book.get("Book_Slug") or "").strip().upper()
        self.dv["code"].set(code_norm)
        self.dv["book_slug"].set(slug_norm)
        self.bunny_prep["chapter_code"].set(code_norm)
        self.bunny_prep["book_slug"].set(slug_norm)

    def _on_tab_changed(self, event):
        """Refresh content when tab selected."""
        try:
            nb = event.widget
            idx = nb.index(nb.select())
            if idx == 2:  # Tab 2 Σήμανση PDF
                if hasattr(self, "tab2_original_pdf_lbl") and self.tab2_original_pdf_lbl.winfo_exists():
                    self.tab2_original_pdf_lbl.config(text=self.pdf_path.get().strip() or "—")
                if hasattr(self, "tab2_marked_pdf_lbl") and self.tab2_marked_pdf_lbl.winfo_exists():
                    self.tab2_marked_pdf_lbl.config(text=self.marked_pdf_path.get().strip() or "—")
            elif idx == 3:  # Tab 3 Έλεγχος Κουμπιών
                if not self._ensure_step3_ready_or_block():
                    nb.select(self.tab2)
            elif idx == 4:  # Tab 4 Βάση Δεδομένων
                self._sync_mp3_folder_ui()
                self._refresh_tab4_catalog()
                self._update_mp3_folder_status()
            elif idx == 5:  # Tab 5 Αντιστοίχιση Ύμνων
                self._refresh_tab5()
            elif idx == 6:  # Tab 6 Έλεγχος Βάσης
                self._update_ready_pdf_status_display()
        except Exception:
            pass

    def _update_ready_pdf_status_display(self):
        """Update ready PDF / pipeline status in Tab 6 when switching to tab."""
        if hasattr(self, "step1_ind") and self.step1_ind.winfo_exists():
            self._pipeline_sync_ui_from_state()

    def _pipeline_sync_ui_from_state(self):
        """Sync pipeline step indicators from current state (on tab switch)."""
        r = self.ready_pdf_validation
        if not r:
            if hasattr(self, "step1_status"):
                self.step1_status.config(text="—", fg=GREY_MED)
            if hasattr(self, "step1_ind"):
                self.step1_ind.config(bg=GREY_MED, text="[1] PDF")
            self._pipeline_disable_steps_2_3_4()
            return
        if r.passed:
            if hasattr(self, "step1_ind"):
                self.step1_ind.config(bg=GREEN, text="[1] PDF PASS")
            if hasattr(self, "step1_status"):
                self.step1_status.config(text=f"PASS — {r.marker_count} markers", fg=GREEN)
            if hasattr(self, "btn_step2"):
                self.btn_step2.config(state="normal")
            if hasattr(self, "step2_status"):
                self.step2_status.config(text="Επιλέξτε φάκελο MP3", fg=TEXT_DARK)
            if self.ready_pipeline_mp3_passed and self.ready_pipeline_mp3_folder:
                if hasattr(self, "step2_ind"):
                    self.step2_ind.config(bg=GREEN, text="[2] MP3 PASS")
                if hasattr(self, "step2_status"):
                    self.step2_status.config(text="PASS", fg=GREEN)
                self._pipeline_enable_step3()
                self._pipeline_update_step4_state()
            else:
                self._pipeline_disable_steps_2_3_4()
                if hasattr(self, "btn_step2"):
                    self.btn_step2.config(state="normal")
        else:
            if hasattr(self, "step1_ind"):
                self.step1_ind.config(bg="#c03030", text="[1] PDF FAIL")
            if hasattr(self, "step1_status"):
                self.step1_status.config(text="FAIL", fg="#c03030")
            self._pipeline_disable_steps_2_3_4()

    def _init_assignments(self):
        """Reset assignments to match current markers (call after detect)."""
        self.assignments = []
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y))
        for _ in ordered:
            self.assignments.append({
                "song_title": tk.StringVar(),
                "echos": tk.StringVar(),
                "section": tk.StringVar(),
            })

    def _on_mp3_folder_entry_changed(self) -> None:
        """When user edits MP3 folder Entry (Tab 4), persist to state so Step 5 sees it."""
        path = (self.dv["mp3_folder"].get() or "").strip()
        if path and os.path.isdir(path):
            self.source_mp3_folder = path

    def _sync_mp3_folder_ui(self) -> None:
        """Sync Tab 4 Entry from stored mp3_folder so UI reflects current state."""
        mp3_folder = self.get_mp3_folder()
        if mp3_folder and hasattr(self, "dv") and "mp3_folder" in self.dv:
            current = (self.dv["mp3_folder"].get() or "").strip()
            if current != mp3_folder:
                self.dv["mp3_folder"].set(mp3_folder)

    def _refresh_tab5(self):
        """Rebuild assignment rows (preserves edits). Uses stored MP3 folder from Step 4 / Tab 4."""
        for w in self.assign_inner.winfo_children():
            w.destroy()
        if not self.mrks:
            tk.Label(
                self.assign_inner,
                text="Tab 2: «Φόρτωση markers από PDF (001…)» ή «Εντοπισμός Markers ■».",
                bg=WHITE,
                fg=TEXT_DARK,
                font=("Arial", 10),
            ).pack(pady=20)
            return
        mp3_folder = self.get_mp3_folder()
        if not mp3_folder:
            tk.Label(
                self.assign_inner,
                text="MP3 folder not selected in Step 4.",
                bg=WHITE,
                fg="#c03030",
                font=("Arial", 11, "bold"),
            ).pack(pady=20)
            tk.Label(
                self.assign_inner,
                text="Επιλέξτε φάκελο MP3 στο Βήμα 4 (Βάση Δεδομένων) ή στο Βήμα 2 της Ready PDF.",
                bg=WHITE,
                fg=TEXT_DARK,
                font=("Arial", 9),
            ).pack(pady=(0, 20))
            return
        if len(self.assignments) != len(self.mrks):
            self._init_assignments()
        code = (self.dv["code"].get() or "").strip() or "BOOK"
        bunny = (self.dv["bunny"].get() or "").strip() or DEFAULT_BUNNY_BASE_URL
        pdf = self.pdf_path.get().strip()
        mp3_folder = self.get_mp3_folder()
        # Use only source files (001 Title.mp3) — exclude generated (MIKP_001-001.mp3)
        source_files = get_source_mp3_files(mp3_folder, code) if mp3_folder and code else []
        mp3_filenames_sorted = sort_mp3_filenames_by_numeric_prefix(source_files)
        prefix_map = build_mp3_prefix_map(mp3_filenames_sorted)
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y))
        for i, m in enumerate(ordered):
            # Use PDF marker id when set (001…) so Tab 5 matches real MP3 by numeric prefix, not row index
            seq = (
                m.number
                if m.number is not None and isinstance(m.number, int) and m.number > 0
                else (i + 1)
            )
            want_id = f"{seq:03d}"
            actual_mp3 = prefix_map.get(want_id, "")
            # File column: real filename only — no fallback to generated (BOOK-001.mp3)
            mp3_file_name = actual_mp3
            final_title = ""
            if actual_mp3:
                final_title = song_title_from_mp3_file(actual_mp3)
            if not final_title:
                for fn in mp3_filenames_sorted:
                    if extract_mp3_id(fn) == want_id:
                        final_title = song_title_from_mp3_file(fn)
                        break
            if not final_title and i < len(mp3_filenames_sorted):
                final_title = song_title_from_mp3_file(mp3_filenames_sorted[i])
            self.assignments[i]["song_title"].set(final_title)
            row = tk.Frame(self.assign_inner, bg=GREY_LIGHT, relief="flat", bd=0, cursor="hand2")
            row.pack(fill="x", pady=1)
            row.bind("<Button-1>", lambda e, idx=i: self._on_preview_row(idx))
            tk.Label(row, text=f"#{seq:03d}", width=5, bg=GREY_LIGHT, font=("Courier", 9)).pack(side="left", padx=2, pady=2)
            tk.Label(row, text=f"Σελ {m.page}", width=5, bg=GREY_LIGHT, font=("Courier", 9)).pack(side="left", padx=2, pady=2)
            tk.Label(row, text=f"Y{m.y:.0f}", width=5, bg=GREY_LIGHT, font=("Courier", 9)).pack(side="left", padx=2, pady=2)
            e_title = tk.Entry(row, textvariable=self.assignments[i]["song_title"], width=28, font=("Arial", 9))
            e_title.pack(side="left", padx=2, pady=2)
            _enable_clipboard_paste(self.root, e_title)
            e_echos = tk.Entry(row, textvariable=self.assignments[i]["echos"], width=8, font=("Arial", 9))
            e_echos.pack(side="left", padx=2, pady=2)
            _enable_clipboard_paste(self.root, e_echos)
            e_section = tk.Entry(row, textvariable=self.assignments[i]["section"], width=10, font=("Arial", 9))
            e_section.pack(side="left", padx=2, pady=2)
            _enable_clipboard_paste(self.root, e_section)
            mp3_code = get_mp3_code(code, seq)
            tk.Label(row, text=mp3_code, width=10, bg=GREY_LIGHT, fg=TEXT_DARK, font=("Courier", 9)).pack(side="left", padx=2, pady=2)
            tk.Label(row, text=mp3_file_name, width=28, bg=GREY_LIGHT, fg=TEXT_DARK, font=("Courier", 8)).pack(side="left", padx=2, pady=2)
            # URL: real filename only — no URL when no real file
            url = get_mp3_url_for_source_file(bunny, code, actual_mp3) if actual_mp3 else ""
            url_short = (url[:45] + "…") if len(url) > 48 else url
            url_lbl = tk.Label(row, text=url_short, width=48, bg=GREY_LIGHT, fg=TEXT_DARK, font=("Courier", 8), anchor="w")
            url_lbl.pack(side="left", padx=2, pady=2, fill="x", expand=True)
            for w in row.winfo_children():
                if isinstance(w, (tk.Label, tk.Entry)):
                    w.bind("<Button-1>", lambda e, idx=i: self._on_preview_row(idx))
        self._s("Ανανέωση Αντιστοίχισης")

    def _on_preview_row(self, row_index: int):
        """Update hymn preview panel when user clicks a song row."""
        if not self.mrks or row_index < 0:
            return
        ordered = sorted(self.mrks, key=lambda m: (m.page, m.y))
        if row_index >= len(ordered):
            return
        m = ordered[row_index]
        # Use actual page from marker (1-based PDF page); never default to 1
        page_number = m.page
        if page_number < 1:
            page_number = 0
        seq = (
            m.number
            if m.number is not None and isinstance(m.number, int) and m.number > 0
            else (row_index + 1)
        )
        hymn_code = f"#{seq:03d}"
        title = (self.assignments[row_index]["song_title"].get() or "").strip()
        display_code = f"{hymn_code}  {title}" if title else hymn_code
        self.preview_code_lbl.config(text=display_code)
        pdf = self.pdf_path.get().strip()
        self._preview_pdf_path = pdf
        self._preview_page = page_number
        self._tab5_preview_row_index = row_index
        # First 1–2 lines of hymn text
        if pdf and os.path.isfile(pdf):
            content = extract_hymn_preview_lines(pdf, m, max_lines=2, max_chars_per_line=90)
            content = content if content else "Preview unavailable for this piece."
        else:
            content = "Preview unavailable for this piece."
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", tk.END)
        self.preview_text.insert("1.0", content)
        self.preview_text.config(state="disabled")
        # Debug: verify page before rendering (remove after fixing)
        if pdf:
            short_pdf = os.path.basename(pdf)
            print(f"Preview row={hymn_code} page={page_number} pdf={short_pdf}")
        self._update_preview_thumbnail(pdf, page_number)

    def _update_preview_thumbnail(self, pdf_path: str, page_number: int):
        """Render and display PDF page thumbnail; show placeholder on failure."""
        if not pdf_path or not os.path.isfile(pdf_path):
            self.preview_thumbnail_lbl.config(image="", text="Thumbnail\nunavailable", compound="center")
            self._preview_photo = None
            return
        if not ImageTk:
            self.preview_thumbnail_lbl.config(image="", text="Thumbnail\nunavailable", compound="center")
            self._preview_photo = None
            return
        if page_number < 1:
            self.preview_thumbnail_lbl.config(image="", text="Thumbnail\nunavailable", compound="center")
            self._preview_photo = None
            return
        img = render_page_thumbnail(pdf_path, page_number, max_width=230)
        if img is None:
            self.preview_thumbnail_lbl.config(image="", text="Thumbnail\nunavailable", compound="center")
            self._preview_photo = None
            return
        photo = ImageTk.PhotoImage(img)
        self._preview_photo = photo
        self.preview_thumbnail_lbl.config(image=photo, text="", compound="image")
        self.preview_thumbnail_lbl.image = photo

    def _find_acrobat_reader(self) -> str | None:
        """Find Adobe Acrobat Reader on Windows. Returns exe path or None."""
        if sys.platform != "win32":
            return None
        common_paths = [
            os.path.join(os.environ.get("ProgramFiles", ""), "Adobe", "Acrobat DC", "Acrobat", "Acrobat.exe"),
            os.path.join(os.environ.get("ProgramFiles", ""), "Adobe", "Acrobat Reader DC", "Reader", "AcroRd64.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Adobe", "Acrobat Reader DC", "Reader", "AcroRd32.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Adobe", "Acrobat Reader 2023", "Reader", "AcroRd32.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Adobe", "Acrobat Reader 2017", "Reader", "AcroRd32.exe"),
        ]
        for p in common_paths:
            if p and os.path.isfile(p):
                return p
        return None

    def _open_preview_pdf(self):
        """Open the selected PDF at the correct page if possible.

        Note: os.startfile() always opens page 1 because the default Windows PDF
        viewer (Edge or built-in) ignores the #page=N fragment in the path.
        We try Adobe Acrobat Reader with /A "page=N" when available.
        """
        path = self._preview_pdf_path or self.pdf_path.get().strip()
        if not path or not os.path.isfile(path):
            self._s("Δεν υπάρχει PDF για άνοιγμα.")
            return
        path = os.path.normpath(path)
        page = self._preview_page or 0
        print(f"Open PDF path={path} page={page}")

        try:
            if sys.platform == "win32":
                acrobat = self._find_acrobat_reader()
                if acrobat and page > 1:
                    subprocess.Popen([acrobat, "/A", f"page={page}", path])
                    self._s(f"Άνοιγμα σελ. {page}: {os.path.basename(path)}")
                else:
                    os.startfile(path)
                    self._s(f"Άνοιγμα: {os.path.basename(path)}")
            else:
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.run([opener, path], check=False)
                self._s(f"Άνοιγμα: {os.path.basename(path)}")
        except (OSError, FileNotFoundError) as e:
            self._s(f"Σφάλμα άνοιγματος: {e}")

    def _pick_pdf(self):
        """Select Original PDF. Compute work_folder = dirname(original_pdf_path)."""
        path = filedialog.askopenfilename(
            title="Επιλογή Original PDF",
            filetypes=[("PDF αρχεία", "*.pdf"), ("Όλα τα αρχεία", "*.*")],
        )
        if not path:
            return
        self.original_pdf_path = path
        self.work_folder = os.path.dirname(path)
        self.pdf_path.set(path)
        self._markers_read_from_pdf = False
        self._step3_ready = False
        if hasattr(self, "tab2_original_pdf_lbl") and self.tab2_original_pdf_lbl.winfo_exists():
            self.tab2_original_pdf_lbl.config(text=path)
        self._s(f"PDF: {os.path.basename(path)}")
        # New PDF: try embedded 001/002… labels first (no ■ pass required); else empty until user runs Tab 2
        self._auto_load_numbered_markers_from_pdf(path)
        if self.mrks:
            self._tab2_renumber_markers()
        else:
            messagebox.showinfo(
                "Step 2 required",
                "PDF is not marked. Please complete Step 2 (Σήμανση).",
            )
        # Load into Tab 2 viewer automatically (manual placement)
        self._tab2_viewer_load_pdf()
        self._refresh_tab5()

    # ─────────────────────────── Tab 2 manual marker viewer ───────────────────────────

    def _tab2_close_doc(self) -> None:
        try:
            if self._tab2_doc is not None:
                self._tab2_doc.close()
        except Exception:
            pass
        self._tab2_doc = None

    def _tab2_viewer_load_pdf(self) -> None:
        pdf = (self.pdf_path.get() or "").strip()
        if not pdf or not os.path.isfile(pdf):
            self._s("Tab 2 viewer: επιλέξτε πρώτα PDF στο Tab 1.")
            return
        if ImageTk is None:
            messagebox.showerror("Viewer", "Λείπει PIL/ImageTk. Εγκαταστήστε Pillow.")
            return
        try:
            import fitz  # type: ignore
            from PIL import Image
        except Exception as e:
            messagebox.showerror("Viewer", f"Αποτυχία φόρτωσης βιβλιοθηκών PDF:\n{e}")
            return

        # Open doc
        self._tab2_close_doc()
        try:
            self._tab2_doc = fitz.open(pdf)
        except Exception as e:
            messagebox.showerror("Viewer", f"Αποτυχία ανοίγματος PDF:\n{e}")
            self._tab2_doc = None
            return

        # Render pages as a continuous stack
        self._tab2_page_photos = []
        self._tab2_page_layout = {}
        self.tab2_canvas.delete("all")

        cw = max(1, int(self.tab2_canvas.winfo_width() or 1))
        max_page_w = max(float(self._tab2_doc[i].rect.width) for i in range(self._tab2_doc.page_count))
        z_fit = float(cw) / max(1.0, max_page_w)
        z = max(0.05, z_fit * float(self._tab2_zoom))
        mat = fitz.Matrix(z, z)

        y_cursor = 0.0
        page_gap = 8.0
        max_w = 1

        for idx in range(self._tab2_doc.page_count):
            page = self._tab2_doc[idx]
            rect = page.rect
            if rect.width <= 0 or rect.height <= 0:
                continue
            pix = page.get_pixmap(matrix=mat, alpha=False)
            if pix.width <= 0 or pix.height <= 0:
                continue
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            photo = ImageTk.PhotoImage(img)
            page_no = idx + 1
            self.tab2_canvas.create_image(0, y_cursor, anchor=tk.NW, image=photo, tags=("pdf",))
            self._tab2_page_photos.append(photo)

            sx = float(pix.width) / float(rect.width)
            sy = float(pix.height) / float(rect.height)
            self._tab2_page_layout[page_no] = (y_cursor, float(pix.height), sx, sy, float(pix.width))
            max_w = max(max_w, pix.width)
            y_cursor += float(pix.height) + page_gap

        self._tab2_pix_width = max_w
        self._tab2_pix_height = max(1, int(y_cursor - page_gap if y_cursor > 0 else 1))
        self.tab2_canvas.configure(scrollregion=(0, 0, self._tab2_pix_width, self._tab2_pix_height))
        self._tab2_redraw_overlay()
        try:
            if hasattr(self, "tab2_manual_status") and self.tab2_manual_status.winfo_exists():
                self.tab2_manual_status.config(text="Manual: click to add • right-click to delete • Ctrl+Z undo")
        except Exception:
            pass
        try:
            self.tab2_canvas.focus_set()
        except Exception:
            pass
        self._s("Tab 2 viewer: έτοιμο για τοποθέτηση markers.")

    def _tab2_on_canvas_resize(self, _event=None) -> None:
        # Best-effort: re-render on resize if a doc is loaded.
        if self._tab2_doc is None:
            return
        self.root.after(120, lambda: self._tab2_viewer_load_pdf())

    def _tab2_on_mousewheel(self, event: tk.Event) -> None:
        try:
            delta = int(getattr(event, "delta", 0))
            if delta == 0:
                return
            units = int(delta / 120) or (1 if delta > 0 else -1)
            self.tab2_canvas.yview_scroll(-units * 3, "units")
        except Exception:
            pass

    def _tab2_pdf_from_canvas(self, canvas_x: float, canvas_y: float, page_no: int) -> tuple[float, float]:
        layout = self._tab2_page_layout.get(page_no)
        if not layout:
            return (0.0, 0.0)
        top, _h, sx, sy, _w = layout
        return (canvas_x / sx, (canvas_y - top) / sy)

    def _tab2_canvas_from_pdf(self, x_pdf: float, y_pdf: float, page_no: int) -> tuple[float, float]:
        layout = self._tab2_page_layout.get(page_no)
        if not layout:
            return (0.0, 0.0)
        top, _h, sx, sy, _w = layout
        return (x_pdf * sx, top + (y_pdf * sy))

    def _tab2_page_at_canvas_point(self, canvas_x: float, canvas_y: float) -> int | None:
        for page_no in sorted(self._tab2_page_layout.keys()):
            top, h, _sx, _sy, w = self._tab2_page_layout[page_no]
            if top <= canvas_y <= (top + h) and 0 <= canvas_x <= w:
                return page_no
        return None

    def _tab2_page_by_y_fallback(self, canvas_y: float) -> int | None:
        """Fallback page hit-test: ignore x, pick page by vertical range/proximity."""
        if not self._tab2_page_layout:
            return None
        for page_no in sorted(self._tab2_page_layout.keys()):
            top, h, _sx, _sy, _w = self._tab2_page_layout[page_no]
            if top <= canvas_y <= (top + h):
                return page_no
        return min(
            self._tab2_page_layout.keys(),
            key=lambda p: abs(self._tab2_page_layout[p][0] - canvas_y),
        )

    def _tab2_renumber_markers(self) -> None:
        active = [m for m in self.mrks if m.keep]
        active.sort(key=lambda m: (m.page, m.y, m.x))
        for i, m in enumerate(active, start=1):
            m.number = i

    def _tab2_find_marker_hit(self, page_no: int, x_pdf: float, y_pdf: float) -> int | None:
        # Hit radius in PDF points, roughly stable across zoom.
        r = 10.0
        best = None
        best_d2 = 1e18
        for i, m in enumerate(self.mrks):
            if not m.keep or int(m.page) != int(page_no):
                continue
            dx = float(m.x) - float(x_pdf)
            dy = float(m.y) - float(y_pdf)
            d2 = dx * dx + dy * dy
            if d2 <= r * r and d2 < best_d2:
                best_d2 = d2
                best = i
        return best

    def _tab2_on_left_click(self, event: tk.Event) -> None:
        if self._tab2_doc is None:
            return
        try:
            canvas_x = float(self.tab2_canvas.canvasx(event.x))
            canvas_y = float(self.tab2_canvas.canvasy(event.y))
            page_no = self._tab2_page_at_canvas_point(canvas_x, canvas_y)
            if page_no is None:
                page_no = self._tab2_page_by_y_fallback(canvas_y)
                if page_no is None:
                    return
            # Clamp x to page visible width so clicks near right-side whitespace still work.
            layout = self._tab2_page_layout.get(page_no)
            if layout is not None:
                _top, _h, _sx, _sy, page_w = layout
                if canvas_x < 0:
                    canvas_x = 0.0
                elif canvas_x > page_w:
                    canvas_x = float(page_w) - 1.0
            x_pdf, y_pdf = self._tab2_pdf_from_canvas(canvas_x, canvas_y, page_no)
            hit = self._tab2_find_marker_hit(page_no, x_pdf, y_pdf)
            if hit is not None:
                self._tab2_selected_idx = hit
                self._tab2_redraw_overlay()
                return

            try:
                import fitz  # type: ignore
            except Exception:
                fitz = None  # type: ignore
            rect = fitz.Rect(x_pdf, y_pdf, x_pdf + 1, y_pdf + 1) if fitz is not None else None
            self.mrks.append(Marker(page=int(page_no), x=float(x_pdf), y=float(y_pdf), rect=rect, keep=True, number=None))
            self._tab2_undo_stack.append(len(self.mrks) - 1)
            self._tab2_selected_idx = len(self.mrks) - 1
            self._tab2_renumber_markers()
            self._refresh_list()
            self._tab2_redraw_overlay()
            try:
                if hasattr(self, "tab2_manual_status") and self.tab2_manual_status.winfo_exists():
                    active = len([m for m in self.mrks if m.keep])
                    self.tab2_manual_status.config(text=f"Manual: {active} markers")
            except Exception:
                pass
        except Exception:
            return

    def _tab2_on_right_click(self, event: tk.Event) -> None:
        if self._tab2_doc is None:
            return
        try:
            canvas_x = float(self.tab2_canvas.canvasx(event.x))
            canvas_y = float(self.tab2_canvas.canvasy(event.y))
            page_no = self._tab2_page_at_canvas_point(canvas_x, canvas_y)
            if page_no is None:
                page_no = self._tab2_page_by_y_fallback(canvas_y)
                if page_no is None:
                    return
            layout = self._tab2_page_layout.get(page_no)
            if layout is not None:
                _top, _h, _sx, _sy, page_w = layout
                if canvas_x < 0:
                    canvas_x = 0.0
                elif canvas_x > page_w:
                    canvas_x = float(page_w) - 1.0
            x_pdf, y_pdf = self._tab2_pdf_from_canvas(canvas_x, canvas_y, page_no)
            hit = self._tab2_find_marker_hit(page_no, x_pdf, y_pdf)
            if hit is None:
                return
            self.mrks.pop(hit)
            self._tab2_selected_idx = -1
            self._tab2_undo_stack = [i for i in self._tab2_undo_stack if i != hit]
            self._tab2_renumber_markers()
            self._refresh_list()
            self._tab2_redraw_overlay()
            try:
                if hasattr(self, "tab2_manual_status") and self.tab2_manual_status.winfo_exists():
                    active = len([m for m in self.mrks if m.keep])
                    self.tab2_manual_status.config(text=f"Manual: {active} markers")
            except Exception:
                pass
        except Exception:
            return

    def _tab2_on_ctrl_z(self, event: tk.Event) -> None:
        # Only act when Tab 2 is visible.
        try:
            w = self.root.focus_get()
            if w is None:
                return
            if hasattr(self, "tab2_canvas") and self.tab2_canvas.winfo_exists():
                if w.winfo_toplevel() != self.root:
                    return
            # If focus is not inside tab2_canvas, still allow Ctrl+Z when user is on Tab 2.
        except Exception:
            return
        if self._tab2_doc is None:
            return
        if not self._tab2_undo_stack:
            return
        idx = self._tab2_undo_stack.pop()
        if idx < 0 or idx >= len(self.mrks):
            return
        try:
            self.mrks.pop(idx)
        except Exception:
            return
        self._tab2_selected_idx = -1
        self._tab2_renumber_markers()
        self._refresh_list()
        self._tab2_redraw_overlay()
        try:
            if hasattr(self, "tab2_manual_status") and self.tab2_manual_status.winfo_exists():
                active = len([m for m in self.mrks if m.keep])
                self.tab2_manual_status.config(text=f"Manual: {active} markers")
        except Exception:
            pass

    def _tab2_redraw_overlay(self) -> None:
        try:
            self.tab2_canvas.delete("overlay")
        except Exception:
            pass
        if not self._tab2_page_layout:
            return
        # Draw markers as small black squares + yellow number
        self._tab2_renumber_markers()
        active = [m for m in self.mrks if m.keep]
        active.sort(key=lambda m: (m.page, m.y, m.x))
        for m in active:
            cx, cy = self._tab2_canvas_from_pdf(float(m.x), float(m.y), int(m.page))
            size = 4
            self.tab2_canvas.create_rectangle(
                cx - size,
                cy - size,
                cx + size,
                cy + size,
                fill="black",
                outline="black",
                tags=("overlay",),
            )
            num = int(m.number) if m.number else 0
            self.tab2_canvas.create_text(
                cx,
                cy + 12,
                text=f"{num:03d}" if num else "",
                fill="#FFD700",
                font=("Arial", 9, "bold"),
                tags=("overlay",),
            )

    def _tab2_viewer_save_markers_pdf(self) -> None:
        pdf = (self.pdf_path.get() or "").strip()
        if not pdf or not os.path.isfile(pdf):
            messagebox.showerror("Σφάλμα", "Επιλέξτε έγκυρο PDF (Tab 1).")
            return
        if not self.mrks:
            messagebox.showerror("Σφάλμα", "Δεν υπάρχουν markers. Κάντε κλικ στο PDF για να προσθέσετε.")
            return
        self._tab2_renumber_markers()
        folder = self.work_folder or os.path.dirname(pdf)
        base = os.path.splitext(os.path.basename(pdf))[0]
        initial = os.path.join(folder, base + "_markers001.pdf")
        out = filedialog.asksaveasfilename(
            title="Αποθήκευση PDF με markers (001…)",
            defaultextension=".pdf",
            filetypes=[("PDF αρχεία", "*.pdf")],
            initialdir=folder,
            initialfile=os.path.basename(initial),
        )
        if not out:
            return
        try:
            n = embed_numbered_markers_pdf(pdf, self.mrks, out)
        except Exception as e:
            messagebox.showerror("Αποτυχία", f"Αποτυχία αποθήκευσης PDF:\n{e}")
            return

        # Switch current PDF to saved one and auto-load markers from it
        self.original_pdf_path = out
        self.work_folder = os.path.dirname(out)
        self.pdf_path.set(out)
        if hasattr(self, "tab2_original_pdf_lbl") and self.tab2_original_pdf_lbl.winfo_exists():
            self.tab2_original_pdf_lbl.config(text=out)
        self.mrks = detect_tofanari_markers_from_pdf(out) or detect_numbered_markers_from_pdf(out)
        # Explicitly regenerate numbering immediately after reload so numbers are visible without extra steps.
        self._tab2_renumber_markers()
        self._markers_read_from_pdf = bool(self.mrks)
        self._init_assignments()
        self._refresh_list()
        self._refresh_tab5()
        self._update_step3_ready_state()
        if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
            self.cnt_lbl.config(text=f"✅ Φορτώθηκαν {len(self.mrks)} markers (001…) από το αποθηκευμένο PDF", fg=GREEN)
        self._tab2_viewer_load_pdf()
        self._s(f"✅ Αποθηκεύτηκε PDF με markers: {os.path.basename(out)} ({n})")

    def _pick_mp3_folder(self):
        """Open folder selection; use selected path exactly. Never append input_mp3 or mp3."""
        selected_folder = filedialog.askdirectory(title="Επιλογή Φακέλου MP3")
        if selected_folder:
            self.source_mp3_folder = selected_folder
            self.dv["mp3_folder"].set(selected_folder)
            files = list_mp3_files_in_folder(selected_folder)
            self._update_mp3_folder_status()
            if len(files) == 0:
                messagebox.showwarning(
                    "No MP3 Files Found",
                    "The selected folder does not contain any .mp3 files.",
                )
            self._s(f"Φάκελος MP3: {selected_folder}")

    def _pick_html_folder(self):
        """Open folder selection for chapter HTML folder (index.html or html/ subfolder)."""
        selected_folder = filedialog.askdirectory(title="Επιλογή φακέλου κεφαλαίου (HTML)")
        if selected_folder:
            self.html_chapter_folder = selected_folder
            self.dv["html_folder"].set(selected_folder)
            self._s(f"Φάκελος κεφαλαίου: {selected_folder}")

    def _analyze_markers(self):
        """Run marker detection from HTML + MP3 matching; update table and report."""
        html_folder = (self.dv["html_folder"].get() or "").strip() or self.html_chapter_folder
        mp3_folder = self.get_mp3_folder()
        if not html_folder:
            messagebox.showwarning(
                "Φάκελος κεφαλαίου",
                "Επιλέξτε φάκελο κεφαλαίου (με index.html ή data-audio markers).",
            )
            return
        if not mp3_folder:
            messagebox.showwarning(
                "Φάκελος MP3",
                "Επιλέξτε φάκελο MP3 με το κουμπί «Επιλογή Φακέλου MP3».",
            )
            return
        self.marker_matching_result = match_markers_to_mp3(html_folder, mp3_folder)
        r = self.marker_matching_result
        # Update tree
        if hasattr(self, "marker_tree") and self.marker_tree.winfo_exists():
            for item in self.marker_tree.get_children():
                self.marker_tree.delete(item)
            for m in r.matches:
                self.marker_tree.insert("", "end", values=(m.marker_id, m.mp3_filename, m.status))
        if hasattr(self, "marker_matching_lbl") and self.marker_matching_lbl.winfo_exists():
            ok = sum(1 for m in r.matches if m.status == "OK")
            self.marker_matching_lbl.config(
                text=f"Detected: {len(r.marker_ids)} markers | Matched: {len(r.mp3_id_to_file)} audio | OK: {ok}"
            )
        # Show marker report in MP3 validation area (prepend)
        report_lines = format_matching_report(r)
        self.mp3_validation_report.config(state="normal")
        self.mp3_validation_report.delete("1.0", tk.END)
        self.mp3_validation_report.insert("1.0", "\n".join(report_lines))
        self.mp3_validation_report.config(state="disabled")
        self._update_mp3_buttons_state(True)
        self._s("Ανάλυση markers ολοκληρώθηκε")

    def _stage1_preupload_audio_validate(self):
        """
        Stage 1: active PDF markers vs plain NNN.mp3 filenames; write preupload JSON + report.
        """
        mp3_folder = self.get_mp3_folder()
        if not self.mrks:
            messagebox.showerror(
                "Stage 1",
                "Δεν υπάρχουν markers. Tab 2: «Φόρτωση markers από PDF» ή «Εντοπισμός Markers ■».",
            )
            return
        if not mp3_folder or not os.path.isdir(mp3_folder):
            messagebox.showwarning(
                "Stage 1",
                "Επιλέξτε φάκελο MP3 (μόνο αρχεία τύπου 001.mp3, 002.mp3, …).",
            )
            return
        wf = (self.work_folder or "").strip()
        if not wf:
            messagebox.showerror(
                "Stage 1",
                "Ορίστε φάκελο εργασίας: επιλέξτε Original PDF (Tab 1) ώστε να υπάρχει work folder.",
            )
            return
        try:
            map_path, rep_path, report = run_stage1_preupload_save(wf, self.mrks, mp3_folder)
        except OSError as e:
            messagebox.showerror("Stage 1", f"Αποτυχία εγγραφής αρχείων:\n{e}")
            return
        except Exception as e:
            messagebox.showerror("Stage 1", str(e))
            return
        ready = bool(report.get("ready_for_upload"))
        lines = [
            "STAGE 1 — PRE-UPLOAD AUDIO",
            "",
            f"ready_for_upload: {ready}",
            f"total_markers: {report.get('total_markers')}",
            f"total_audio_files: {report.get('total_audio_files')}",
            f"matched_count: {report.get('matched_count')}",
            f"missing_audio_ids: {report.get('missing_audio_ids')}",
            f"orphan_audio_ids: {report.get('orphan_audio_ids')}",
            f"duplicate_marker_ids: {report.get('duplicate_marker_ids')}",
            f"duplicate_audio_ids: {report.get('duplicate_audio_ids')}",
            "",
            f"Mapping: {map_path}",
            f"Report:  {rep_path}",
        ]
        if hasattr(self, "mp3_validation_report") and self.mp3_validation_report.winfo_exists():
            self.mp3_validation_report.config(state="normal")
            self.mp3_validation_report.delete("1.0", tk.END)
            self.mp3_validation_report.insert("1.0", "\n".join(lines))
            self.mp3_validation_report.config(state="disabled")
        messagebox.showinfo(
            "Stage 1 — Pre-upload validation",
            ("Έτοιμο για upload.\n\n" if ready else "Δεν είναι έτοιμο για upload.\n\n")
            + f"matched: {report.get('matched_count')} / {report.get('total_markers')}\n\n"
            f"Αποθηκεύτηκαν:\n{map_path}\n{rep_path}",
        )
        self._s("Stage 1: pre-upload audio validation OK" if ready else "Stage 1: validation — not ready for upload")

    def _stage2_postupload_audio_urls(self):
        """Stage 2: read pre-upload mapping, append Bunny URLs, write postupload_audio_mapping.json."""
        wf = (self.work_folder or "").strip()
        if not wf:
            messagebox.showerror(
                "Stage 2",
                "Ορίστε φάκελο εργασίας (επιλογή Original PDF στο Tab 1).",
            )
            return
        pre_path = os.path.join(wf, PREUPLOAD_AUDIO_MAPPING_JSON)
        if not os.path.isfile(pre_path):
            messagebox.showerror(
                "Stage 2",
                f"Δεν βρέθηκε {PREUPLOAD_AUDIO_MAPPING_JSON}.\n"
                "Εκτελέστε πρώτα το Stage 1 στον ίδιο φάκελο εργασίας.",
            )
            return
        bunny = (self.dv["bunny"].get() or "").strip() or DEFAULT_BUNNY_BASE_URL
        try:
            out_path = run_stage2_postupload_save(wf, bunny, uploaded_filenames=None)
        except Exception as e:
            messagebox.showerror("Stage 2", str(e))
            return
        messagebox.showinfo(
            "Stage 2 — Remote URLs",
            f"Δημιουργήθηκε:\n{out_path}\n\n"
            f"(Μόνο γραμμές matched_local → matched_remote, base: {bunny})",
        )
        self._s(f"Stage 2: {POSTUPLOAD_AUDIO_MAPPING_JSON}")

    def _update_mp3_folder_status(self) -> None:
        """Update Tab 4: MP3 folder path, file count, validation report, and button state."""
        mp3_folder = self.get_mp3_folder()
        files = list_mp3_files_in_folder(mp3_folder) if mp3_folder else []
        count = len(files)
        # Update count label
        if hasattr(self, "mp3_files_count_lbl") and self.mp3_files_count_lbl.winfo_exists():
            self.mp3_files_count_lbl.config(text=f"MP3 Files Found: {count}")
        # Update validation report
        if hasattr(self, "mp3_validation_report") and self.mp3_validation_report.winfo_exists():
            self.mp3_validation_report.config(state="normal")
            self.mp3_validation_report.delete("1.0", tk.END)
            if not mp3_folder:
                self.mp3_validation_report.insert("1.0", "Επιλέξτε φάκελο MP3 και πατήστε «Έλεγχος MP3» για αναφορά.")
            else:
                lines = [
                    "MP3 VALIDATION REPORT",
                    "",
                    f"MP3 folder: {mp3_folder}",
                    "",
                ]
                if count == 0:
                    lines.extend([
                        "No MP3 files found in the selected folder.",
                        "Please select a folder that contains MP3 files.",
                    ])
                else:
                    lines.append(f"MP3 Files Found: {count}")
                self.mp3_validation_report.insert("1.0", "\n".join(lines))
            self.mp3_validation_report.config(state="disabled")
        # Enable/disable buttons based on MP3 count
        self._update_mp3_buttons_state(count > 0)

    def _update_mp3_buttons_state(self, enabled: bool) -> None:
        """Enable or disable Έλεγχος MP3 and Δημιουργία database.xlsx buttons.
        gen_db is also enabled when marker_matching_result has OK matches (marker-based path).
        """
        if hasattr(self, "btn_validate_mp3") and self.btn_validate_mp3.winfo_exists():
            self.btn_validate_mp3.config(state="normal" if enabled else "disabled")
        gen_db_ok = enabled or (
            self.marker_matching_result is not None
            and any(m.status == "OK" for m in self.marker_matching_result.matches)
        )
        if hasattr(self, "btn_gen_db") and self.btn_gen_db.winfo_exists():
            self.btn_gen_db.config(state="normal" if gen_db_ok else "disabled")

    def _detect(self):
        pdf = self.pdf_path.get().strip()
        if not pdf or not os.path.isfile(pdf):
            messagebox.showerror("Σφάλμα", "Επιλέξτε έγκυρο PDF (Tab 1).")
            return
        self._s("🔍 Αναζήτηση markers…")
        self.mrks = detect_markers(pdf)
        self._markers_read_from_pdf = bool(self.mrks)
        self._init_assignments()
        n = len(self.mrks)
        self.cnt_lbl.config(
            text=f"✅ Βρέθηκαν {n} markers ■" if n else "⚠️ Δεν βρέθηκαν markers ■",
            fg=GREEN if n else "orange",
        )
        self._refresh_list()
        self._update_step3_ready_state()
        self._s(f"Βρέθηκαν {n} markers.")

    def _auto_load_numbered_markers_from_pdf(self, pdf_path: str) -> None:
        """On PDF select: load visible 001,002,… text labels if present (PyMuPDF span scan)."""
        try:
            found = detect_tofanari_markers_from_pdf(pdf_path) or detect_numbered_markers_from_pdf(pdf_path)
        except Exception:
            found = []
        self.mrks = list(found) if found else []
        self._markers_read_from_pdf = bool(found)
        self._init_assignments()
        self._update_step3_ready_state()
        if found:
            if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
                self.cnt_lbl.config(
                    text=f"✅ Φορτώθηκαν {len(found)} markers (001…) από το PDF",
                    fg=GREEN,
                )
            if hasattr(self, "lst") and self.lst.winfo_exists():
                self._refresh_list()
            self._s(f"Αυτόματη φόρτωση: {len(found)} αριθμημένων markers από PDF.")
        else:
            if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
                self.cnt_lbl.config(text="—", fg=DARK_RED)
            if hasattr(self, "lst") and self.lst.winfo_exists():
                self._refresh_list()

    def _load_markers_from_pdf_numbers(self):
        """Manual reload: visible 3-digit labels in PDF (same as auto-load on pick)."""
        pdf = self.pdf_path.get().strip()
        if not pdf or not os.path.isfile(pdf):
            messagebox.showerror("Σφάλμα", "Επιλέξτε έγκυρο PDF (Tab 1).")
            return
        self._s("🔍 Ανάγνωση ετικετών 001, 002… από PDF…")
        try:
            found = detect_tofanari_markers_from_pdf(pdf) or detect_numbered_markers_from_pdf(pdf)
        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Αποτυχία ανάγνωσης PDF:\n{e}")
            return
        self.mrks = found
        self._markers_read_from_pdf = bool(found)
        self._init_assignments()
        self._update_step3_ready_state()
        n = len(found)
        if hasattr(self, "cnt_lbl") and self.cnt_lbl.winfo_exists():
            self.cnt_lbl.config(
                text=(
                    f"✅ Φορτώθηκαν {n} markers (001…)"
                    if n
                    else "⚠️ Δεν βρέθηκαν ετικέτες 001, 002… (αριθμοί στο κείμενο)"
                ),
                fg=GREEN if n else "orange",
            )
        self._refresh_list()
        self._refresh_tab5()
        self._s(f"Φόρτωση από PDF: {n} markers." if n else "Χωρίς αριθμημένες ετικέτες στο PDF.")

    def _apply(self):
        pdf = self.pdf_path.get().strip()
        if not pdf:
            messagebox.showerror("Σφάλμα", "Επιλέξτε PDF (Tab 1).")
            return
        if not self.mrks:
            messagebox.showerror(
                "Σφάλμα",
                "Εκτελέστε πρώτα «Φόρτωση markers από PDF» ή «Εντοπισμός Markers ■» (Tab 2).",
            )
            return
        folder = self.work_folder or os.path.dirname(pdf)
        base = os.path.splitext(os.path.basename(pdf))[0]
        out = os.path.join(folder, base + "_με_κουμπιά.pdf")
        self.prog["value"] = 0

        def run():
            def cb(done: int, total: int):
                pct = int(done / total * 100)
                self.prog["value"] = pct
                self.prog_lbl.config(text=f"{done}/{total} κουμπιά")
                self.root.update_idletasks()

            n = apply_markers(pdf, self.mrks, out, progress_cb=cb)
            # Show messagebox on main thread
            self.root.after(
                0,
                lambda: self._on_apply_done(out, n),
            )

        threading.Thread(target=run, daemon=True).start()

    def _generate_final_pdf_with_buttons(self) -> None:
        """Dedicated final-generation entry point with explicit checks before calling existing apply flow."""
        pdf = self.pdf_path.get().strip()
        if not pdf:
            messagebox.showerror("Σφάλμα", "Επιλέξτε PDF (Tab 1).")
            return
        if not self.mrks:
            messagebox.showerror(
                "Σφάλμα",
                "PDF is not marked. Please complete Step 2 (Σήμανση).",
            )
            return
        # Ensure numbering exists before final output.
        self._tab2_renumber_markers()
        self._update_step3_ready_state()
        self._apply()

    def _on_apply_done(self, out_path: str, count: int):
        self.marked_pdf_path.set(out_path)
        if hasattr(self, "tab2_marked_pdf_lbl") and self.tab2_marked_pdf_lbl.winfo_exists():
            self.tab2_marked_pdf_lbl.config(text=out_path)
        self._s(f"✅ Αποθηκεύτηκε: {out_path}  ({count} κουμπιά)")
        messagebox.showinfo(
            "Έτοιμο!",
            f"✅ Marked PDF:\n{out_path}\n\n{count} κουμπιά τοποθετήθηκαν.",
        )

    def _refresh_list(self):
        if not self._ensure_step3_ready_or_block():
            return
        self.lst.delete(0, "end")
        for i, m in enumerate(self.mrks):
            state = "✅" if m.keep else "⛔"
            num_disp = (
                m.number
                if m.number is not None and isinstance(m.number, int) and m.number > 0
                else (i + 1)
            )
            self.lst.insert(
                "end",
                f"  {state}  #{num_disp:03d}   Σελ {m.page:3d}   Y={m.y:.0f}",
            )

    def _toggle_keep(self):
        if not self._ensure_step3_ready_or_block():
            return
        sel = self.lst.curselection()
        if not sel:
            messagebox.showinfo(
                "Πληροφορία", "Επιλέξτε ένα στοιχείο από τη λίστα."
            )
            return
        idx = sel[0]
        self.mrks[idx].keep = not self.mrks[idx].keep
        self._refresh_list()
        self.lst.selection_set(idx)

    def _validate_mp3(self):
        """Validate MP3 filenames: numbering (missing, duplicate), pattern. Display in Tab 4 report."""
        mp3_folder = self.get_mp3_folder()
        if not mp3_folder or not mp3_folder.strip():
            messagebox.showwarning(
                "Φάκελος MP3",
                "Επιλέξτε φάκελο MP3 με το κουμπί «Επιλογή Φακέλου MP3».",
            )
            return
        if not os.path.isdir(mp3_folder):
            messagebox.showerror("Φάκελος MP3", f"Ο φάκελος δεν υπάρχει: {mp3_folder}")
            return
        files = list_mp3_files_in_folder(mp3_folder)
        num_res = validate_mp3_numbering(files)
        pattern_res = validate_mp3_filename_pattern(files)
        errs = (num_res.get("errors") or []) + (pattern_res.get("errors") or [])
        warns = (num_res.get("warnings") or []) + (pattern_res.get("warnings") or [])
        lines = [
            "MP3 VALIDATION REPORT",
            "",
            f"MP3 Folder: {mp3_folder}",
            f"MP3 Files Found: {len(files)}",
            "",
            f"Errors: {len(errs)}",
            f"Warnings: {len(warns)}",
            "",
        ]
        if errs:
            lines.append("--- ERRORS ---")
            lines.extend(errs)
            lines.append("")
        if warns:
            lines.append("--- WARNINGS ---")
            lines.extend(warns)
            lines.append("")
        if not errs and not warns and files:
            lines.append("OK — All MP3 filenames valid (NNN Title.mp3).")
            self._s("✅ Έλεγχος MP3 OK")
        elif not errs and warns:
            lines.append("OK with warnings — review above.")
            self._s("⚠️ Έλεγχος MP3 με προειδοποιήσεις")
        else:
            lines.append("Issues detected — fix errors above.")
            self._s("⚠️ Έλεγχος MP3: προβλήματα")
        self.mp3_validation_report.config(state="normal")
        self.mp3_validation_report.delete("1.0", tk.END)
        self.mp3_validation_report.insert("1.0", "\n".join(lines))
        self.mp3_validation_report.config(state="disabled")

    def _gen_db(self):
        code = (self.dv["code"].get() or "").strip().upper()
        book_slug = (self.dv["book_slug"].get() or "").strip().upper()
        if not code:
            messagebox.showerror(
                "Σφάλμα",
                "Επιλέξτε βιβλίο και κεφάλαιο από το Parameters (Tab 0).\nΔεν επιτρέπονται αυθαίρετοι κωδικοί.",
            )
            return
        if not self.book_registry_books:
            messagebox.showerror(
                "Σφάλμα",
                "Φορτώστε πρώτα το Parameters / Master Catalog (Tab 0).\nΌλες οι λειτουργίες απαιτούν προκαθορισμένο κατάλογο βιβλίων.",
            )
            return
        ok, book, errs = validate_chapter_in_catalog(code, book_slug, self.book_registry_books)
        if not ok:
            messagebox.showerror(
                "Σφάλμα",
                "Το κεφάλαιο δεν ανήκει στο επιλεγμένο βιβλίο ή λείπει από το κατάλογο:\n" + "\n".join(errs or ["Unknown error"]),
            )
            return
        bunny = (self.dv["bunny"].get() or "").strip() or DEFAULT_BUNNY_BASE_URL

        # Marker-based path: use when we have marker matches (no PDF required)
        use_marker_path = (
            self.marker_matching_result is not None
            and any(m.status == "OK" for m in self.marker_matching_result.matches)
        )
        if use_marker_path:
            fo = (
                self.html_chapter_folder
                or (self.dv["html_folder"].get() or "").strip()
                or os.path.dirname(self.get_mp3_folder() or "")
            )
            if not fo or not os.path.isdir(fo):
                messagebox.showerror(
                    "Σφάλμα",
                    "Επιλέξτε φάκελο κεφαλαίου (HTML) για αποθήκευση database.xlsx.",
                )
                return
            ok_matches = [m for m in self.marker_matching_result.matches if m.status == "OK"]
            assignments_data = [{"song_title": "", "echos": "", "section": "", "notes": ""} for _ in ok_matches]
            path, count = build_database_xlsx_from_marker_matches(
                fo,
                ok_matches,
                code=code,
                bunny_base_url=bunny,
                assignments=assignments_data,
            )
            self._s(f"✅ {path}")
            messagebox.showinfo(
                "Έτοιμο!",
                f"✅ Βάση δεδομένων (marker-based):\n{path}\n\nΕγγραφές: {count}",
            )
            lines = preview_lines_from_marker_matches(
                ok_matches,
                code=code,
                bunny_base_url=bunny,
            )
        else:
            # PDF-based path (existing)
            if not self.mrks:
                messagebox.showerror(
                    "Σφάλμα",
                    "Tab 2: «Φόρτωση markers από PDF» / «Εντοπισμός ■» ή Tab 4 «Ανάλυση Markers» (HTML + MP3).",
                )
                return
            pdf = self.pdf_path.get().strip()
            fo = self.work_folder or (os.path.dirname(pdf) if pdf else "")
            if not fo:
                messagebox.showerror("Σφάλμα", "Επιλέξτε Original PDF (Tab 1).")
                return
            ok, msg = validate_duplicate_positions(self.mrks)
            if not ok:
                messagebox.showwarning("Προειδοποίηση", msg + "\n\nΜπορείτε να συνεχίσετε ή να διορθώσετε στη λίστα (Tab 3).")
            ok, msg = validate_page_numbers(self.mrks)
            if not ok:
                messagebox.showerror("Σφάλμα", msg)
                return
            # Build DB/preview only from active markers so numbering starts at the first one.
            ordered_active = [m for m in self.mrks if m.keep]
            ordered_active.sort(key=lambda m: (m.page, m.y))
            assignments_data = []
            idx_by_marker = {id(m): i for i, m in enumerate(self.mrks)}
            for m in ordered_active:
                i = idx_by_marker.get(id(m))
                if i is None:
                    continue
                a = self.assignments[i] if i < len(self.assignments) else {}
                sv = lambda v: (v.get() if hasattr(v, "get") else (v or ""))
                song_title = (sv(a.get("song_title", "")) or "").strip()
                echos = (sv(a.get("echos", "")) or "").strip()
                section = (sv(a.get("section", "")) or "").strip()
                assignments_data.append({
                    "song_title": song_title,
                    "echos": echos,
                    "section": section,
                    "status": "TODO",
                    "notes": "",
                })
            source_mp3_files = get_source_mp3_files(self.get_mp3_folder(), code)
            path, count = build_database_xlsx(
                fo,
                self.mrks,
                code=code,
                bunny_base_url=bunny,
                assignments=assignments_data,
                source_mp3_files=source_mp3_files,
            )
            self._s(f"✅ {path}")
            n_source = len(source_mp3_files)
            messagebox.showinfo(
                "Έτοιμο!",
                f"✅ Βάση δεδομένων:\n{path}\n\nΕγγραφές: {count}\n"
                f"Αρχεία MP3 πηγής (χρησιμοποιήθηκαν): {n_source}",
            )
            lines = preview_lines(
                self.mrks,
                code=code,
                bunny_base_url=bunny,
                source_mp3_files=source_mp3_files,
            )
        self.dbprev.config(state="normal")
        self.dbprev.delete("1.0", "end")
        self.dbprev.insert("end", "\n".join(lines))
        self.dbprev.config(state="disabled")
        self._refresh_tab5()
        self._s("Βάση δεδομένων ενημερώθηκε.")


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] == "prepare_bunny":
        from bunny_prepare import main_cli

        raise SystemExit(main_cli(sys.argv[2:]))
    if len(sys.argv) >= 2 and sys.argv[1] == "upload_bunny":
        from bunny_workflow import main_workflow_cli

        raise SystemExit(main_workflow_cli(sys.argv[2:]))
    root = tk.Tk()
    App(root)
    root.mainloop()
