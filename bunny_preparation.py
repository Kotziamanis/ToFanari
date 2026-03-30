# -*- coding: utf-8 -*-
"""
Tofanari Main Tool — Bunny.net Chapter Preparation.

Chapter-level validation and readiness reporting for professional publishing workflow.
Each chapter is the master source; validation runs before upload.
No merge, no FlipBuilder, no Thinkific — preparation only.
"""

import csv
import os
import re
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Expected database.xlsx structure
DB_SHEET = "Buttons"
DB_MP3_CODE_COL = "MP3 Code"
DB_MP3_FILE_COL = "MP3 File"

# Source MP3 pattern: "001 Title.mp3" -> 001
_SOURCE_MP3_NUMBER_RE = re.compile(r"^(\d{3})\s")


def build_bunny_folder_path(root_remote_folder: str, book_slug: str, chapter_code: str) -> str:
    """Bunny target folder: root/book_slug/chapter_code/ (e.g. books/anastasimatarion/AN01/)."""
    root = (root_remote_folder or "").strip().strip("/") or "books"
    slug = (book_slug or "").strip().strip("/")
    code = (chapter_code or "").strip().strip("/")
    if not slug or not code:
        return ""
    return f"{root}/{slug}/{code}/"


def build_bunny_public_url(
    base_cdn_url: str,
    root_remote_folder: str,
    book_slug: str,
    chapter_code: str,
    mp3_filename: str,
) -> str:
    """Bunny public URL: base/root/book_slug/chapter_code/mp3_filename."""
    base = (base_cdn_url or "").strip().rstrip("/")
    root = (root_remote_folder or "").strip().strip("/") or "books"
    slug = (book_slug or "").strip().strip("/")
    code = (chapter_code or "").strip().strip("/")
    if not base or not slug or not code or not mp3_filename:
        return ""
    return f"{base}/{quote(root, safe='')}/{quote(slug, safe='')}/{quote(code, safe='')}/{quote(mp3_filename, safe='')}"


def _parse_chapter_code_from_mp3_code(mp3_code: str) -> Optional[str]:
    """AN01-001 -> AN01, AN02-042 -> AN02."""
    if not mp3_code or "-" not in mp3_code:
        return None
    parts = mp3_code.strip().split("-")
    if len(parts) >= 2 and parts[1].isdigit():
        return parts[0].strip()
    return None


def _list_mp3_files(folder: str) -> List[str]:
    """List .mp3 files in folder (flat, case-insensitive)."""
    if not folder or not os.path.isdir(folder):
        return []
    try:
        return sorted(
            f for f in os.listdir(folder)
            if os.path.isfile(os.path.join(folder, f)) and f.lower().endswith(".mp3")
        )
    except OSError:
        return []


def _read_database_info(work_folder: str) -> Tuple[bool, Optional[str], int, List[str], List[str]]:
    """
    Read database.xlsx from work_folder.
    Returns: (ok, chapter_code, row_count, mp3_codes, mp3_files)
    chapter_code from first MP3 Code; row_count = data rows; mp3_codes/mp3_files from rows.
    """
    if not openpyxl:
        return (False, None, 0, [], [])
    path = os.path.join(work_folder, "database.xlsx")
    if not os.path.isfile(path):
        return (False, None, 0, [], [])
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return (False, None, 0, [], [])
    sheet = wb["Buttons"] if "Buttons" in wb.sheetnames else (wb.active if wb.active else None)
    if not sheet:
        wb.close()
        return (False, None, 0, [], [])
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return (False, None, 0, [], [])
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header) if h}
    mp3_code_col = col_idx.get("MP3 Code", col_idx.get("MP3_Code", -1))
    mp3_file_col = col_idx.get("MP3 File", col_idx.get("MP3_File", -1))
    if mp3_code_col < 0:
        return (False, None, 0, [], [])
    chapter_code = None
    mp3_codes: List[str] = []
    mp3_files: List[str] = []
    for row in rows[1:]:
        vals = list(row) + [None] * (len(header) - len(row)) if row else []
        mc = str(vals[mp3_code_col]).strip() if mp3_code_col < len(vals) and vals[mp3_code_col] else ""
        mf = str(vals[mp3_file_col]).strip() if mp3_file_col >= 0 and mp3_file_col < len(vals) and vals[mp3_file_col] else ""
        if mc:
            mp3_codes.append(mc)
            if not chapter_code:
                chapter_code = _parse_chapter_code_from_mp3_code(mc)
        if mf:
            mp3_files.append(mf)
    return (True, chapter_code or "", len(mp3_codes), mp3_codes, mp3_files)


def _source_num_to_file(mp3_files: List[str]) -> Dict[str, str]:
    """Map 001 -> filename, 002 -> filename, etc."""
    out: Dict[str, str] = {}
    for f in mp3_files or []:
        m = _SOURCE_MP3_NUMBER_RE.match(f)
        if m:
            out[m.group(1)] = f
    return out


def validate_chapter(
    book_slug: str,
    chapter_code: str,
    local_work_folder: str,
    local_mp3_folder: str,
    base_cdn_url: str = "",
    root_remote_folder: str = "books",
) -> Dict[str, Any]:
    """
    Validate a single chapter for Bunny upload readiness.
    Returns: {
        ok: bool,
        status: "READY" | "NOT_READY",
        errors: [...],
        warnings: [...],
        bunny_folder: str,
        bunny_base_url: str,
        row_count: int,
        mp3_count: int,
    }
    """
    errors: List[str] = []
    warnings: List[str] = []
    bunny_folder = build_bunny_folder_path(root_remote_folder, book_slug, chapter_code)
    bunny_base = (base_cdn_url or "").strip().rstrip("/")

    # 1. Work folder exists
    if not local_work_folder or not local_work_folder.strip():
        errors.append("Local work folder is empty.")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": 0,
            "mp3_count": 0,
        }
    if not os.path.isdir(local_work_folder):
        errors.append(f"Local work folder does not exist: {local_work_folder}")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": 0,
            "mp3_count": 0,
        }

    # 2. database.xlsx exists and is readable
    db_ok, db_chapter, row_count, mp3_codes, mp3_files_from_db = _read_database_info(local_work_folder)
    if not db_ok:
        errors.append("database.xlsx not found or unreadable in work folder.")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": 0,
            "mp3_count": 0,
        }
    if row_count == 0:
        errors.append("database.xlsx has no hymn rows.")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": 0,
            "mp3_count": 0,
        }

    # 3. Chapter code consistency (DB vs manifest)
    effective_code = db_chapter or chapter_code
    if db_chapter and chapter_code and db_chapter != chapter_code:
        warnings.append(f"Manifest chapter_code ({chapter_code}) differs from DB ({db_chapter}). Using DB.")

    # 4. MP3 folder exists
    if not local_mp3_folder or not local_mp3_folder.strip():
        errors.append("Local MP3 folder is empty.")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": row_count,
            "mp3_count": 0,
        }
    if not os.path.isdir(local_mp3_folder):
        errors.append(f"Local MP3 folder does not exist: {local_mp3_folder}")
        return {
            "ok": False,
            "status": "NOT_READY",
            "errors": errors,
            "warnings": warnings,
            "bunny_folder": bunny_folder,
            "bunny_base_url": bunny_base,
            "row_count": row_count,
            "mp3_count": 0,
        }

    # 5. MP3 files match rows (source numbering 001, 002, ...)
    mp3_file_list = _list_mp3_files(local_mp3_folder)
    mp3_count = len(mp3_file_list)
    num_to_file = _source_num_to_file(mp3_file_list)
    missing = []
    for i in range(1, row_count + 1):
        num_str = f"{i:03d}"
        if num_str not in num_to_file:
            missing.append(num_str)
    if missing:
        errors.append(f"Missing source MP3 numbers: {', '.join(missing[:10])}{'...' if len(missing) > 10 else ''}")

    # 6. Count consistency
    if mp3_count < row_count:
        errors.append(f"MP3 count ({mp3_count}) < hymn rows ({row_count}).")
    elif mp3_count > row_count:
        warnings.append(f"More MP3 files ({mp3_count}) than hymn rows ({row_count}).")

    # 7. Bunny config
    if not bunny_base:
        warnings.append("Base CDN URL not set.")
    elif not bunny_base.lower().startswith("https://"):
        errors.append("Base CDN URL must start with https://")
    if not book_slug or " " in book_slug:
        errors.append("Book slug required and must not contain spaces.")

    ok = len(errors) == 0
    status = "READY" if ok else "NOT_READY"
    return {
        "ok": ok,
        "status": status,
        "errors": errors,
        "warnings": warnings,
        "bunny_folder": bunny_folder,
        "bunny_base_url": bunny_base,
        "row_count": row_count,
        "mp3_count": mp3_count,
    }


# Manifest columns
MANIFEST_COLUMNS = ["book_slug", "chapter_code", "local_work_folder", "local_mp3_folder"]
MANIFEST_OPTIONAL = ["book_title", "chapter_name"]


def load_chapters_manifest(path: str) -> Tuple[bool, List[Dict[str, str]], List[str]]:
    """
    Load chapters manifest from CSV or XLSX.
    Returns: (ok, list of chapter records, errors)
    Each record: {book_slug, chapter_code, local_work_folder, local_mp3_folder, book_title?, chapter_name?}
    """
    errors: List[str] = []
    records: List[Dict[str, str]] = []
    if not path or not os.path.isfile(path):
        return (False, [], [f"File not found: {path}"])

    path_lower = path.lower()
    if path_lower.endswith(".csv"):
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for i, row in enumerate(reader, start=2):
                    rec = {k.strip(): (v.strip() if v else "") for k, v in row.items() if k}
                    slug = rec.get("book_slug", "").strip()
                    code = rec.get("chapter_code", "").strip()
                    work = rec.get("local_work_folder", "").strip()
                    mp3 = rec.get("local_mp3_folder", "").strip()
                    if not slug or not code:
                        errors.append(f"Row {i}: book_slug and chapter_code required.")
                        continue
                    rec["book_slug"] = slug
                    rec["chapter_code"] = code
                    rec["local_work_folder"] = work
                    rec["local_mp3_folder"] = mp3
                    records.append(rec)
        except Exception as e:
            return (False, [], [f"Failed to read CSV: {e}"])
    elif path_lower.endswith(".xlsx") and openpyxl:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                return (False, [], ["Manifest sheet is empty."])
            header = [str(c).strip() if c else "" for c in rows[0]]
            col_idx = {h: i for i, h in enumerate(header) if h}
            for req in ["book_slug", "chapter_code", "local_work_folder", "local_mp3_folder"]:
                if req not in col_idx:
                    return (False, [], [f"Required column '{req}' not found."])
            for i, row in enumerate(rows[1:], start=2):
                vals = list(row) + [None] * (len(header) - len(row)) if row else []
                slug = str(vals[col_idx["book_slug"]] or "").strip()
                code = str(vals[col_idx["chapter_code"]] or "").strip()
                work = str(vals[col_idx["local_work_folder"]] or "").strip()
                mp3 = str(vals[col_idx["local_mp3_folder"]] or "").strip()
                if not slug or not code:
                    errors.append(f"Row {i}: book_slug and chapter_code required.")
                    continue
                rec = {"book_slug": slug, "chapter_code": code, "local_work_folder": work, "local_mp3_folder": mp3}
                for opt in MANIFEST_OPTIONAL:
                    if opt in col_idx:
                        rec[opt] = str(vals[col_idx[opt]] or "").strip()
                records.append(rec)
        except Exception as e:
            return (False, [], [f"Failed to read XLSX: {e}"])
    else:
        return (False, [], ["Manifest must be CSV or XLSX. Install openpyxl for XLSX."])

    return (len(records) > 0, records, errors)


def run_chapter_preparation(
    manifest_path: str,
    base_cdn_url: str = "",
    root_remote_folder: str = "books",
) -> Tuple[bool, List[Dict[str, Any]], List[str]]:
    """
    Run validation for all chapters in manifest.
    Returns: (manifest_ok, list of report rows, manifest_errors)
    Each report row: {
        book_slug, chapter_code, book_title, chapter_name,
        local_work_folder, local_mp3_folder,
        bunny_target_folder, bunny_sample_url,
        status, row_count, mp3_count,
        errors, warnings,
    }
    """
    ok, records, manifest_errors = load_chapters_manifest(manifest_path)
    if not ok:
        return (False, [], manifest_errors)
    if not records:
        return (True, [], [])

    report: List[Dict[str, Any]] = []
    for rec in records:
        slug = rec.get("book_slug", "")
        code = rec.get("chapter_code", "")
        work = rec.get("local_work_folder", "")
        mp3 = rec.get("local_mp3_folder", "")
        v = validate_chapter(
            book_slug=slug,
            chapter_code=code,
            local_work_folder=work,
            local_mp3_folder=mp3,
            base_cdn_url=base_cdn_url,
            root_remote_folder=root_remote_folder,
        )
        sample_url = ""
        if v["bunny_base_url"] and v["bunny_folder"]:
            sample_url = build_bunny_public_url(
                base_cdn_url, root_remote_folder, slug, code, f"{code}-001.mp3"
            )
        report.append({
            "book_slug": slug,
            "chapter_code": code,
            "book_title": rec.get("book_title", ""),
            "chapter_name": rec.get("chapter_name", ""),
            "local_work_folder": work,
            "local_mp3_folder": mp3,
            "bunny_target_folder": v["bunny_folder"],
            "bunny_sample_url": sample_url,
            "status": v["status"],
            "row_count": v["row_count"],
            "mp3_count": v["mp3_count"],
            "errors": v["errors"],
            "warnings": v["warnings"],
        })
    return (True, report, [])


def format_preparation_report(
    report: List[Dict[str, Any]],
    base_cdn_url: str = "",
    manifest_warnings: List[str] = None,
) -> str:
    """Operator-friendly text report."""
    lines = [
        "BUNNY CHAPTER PREPARATION REPORT",
        "=================================",
        "",
        f"Base CDN URL: {base_cdn_url or 'N/A'}",
        f"Chapters: {len(report)}",
        "",
    ]
    if manifest_warnings:
        lines.append("Manifest warnings (skipped rows):")
        for w in manifest_warnings:
            lines.append(f"  - {w}")
        lines.append("")
    ready = sum(1 for r in report if r.get("status") == "READY")
    not_ready = len(report) - ready
    lines.append(f"READY: {ready}  |  NOT_READY: {not_ready}")
    lines.append("")
    lines.append("-" * 80)

    for i, r in enumerate(report, 1):
        status = r.get("status", "?")
        slug = r.get("book_slug", "")
        code = r.get("chapter_code", "")
        title = r.get("book_title", "") or slug
        ch_name = r.get("chapter_name", "") or code
        work = r.get("local_work_folder", "")
        mp3 = r.get("local_mp3_folder", "")
        bunny = r.get("bunny_target_folder", "")
        errs = r.get("errors", [])
        warns = r.get("warnings", [])
        rc = r.get("row_count", 0)
        mc = r.get("mp3_count", 0)

        lines.append("")
        lines.append(f"[{i}] {title} / {ch_name}")
        lines.append(f"    Book: {slug}  |  Chapter: {code}")
        lines.append(f"    Status: {status}")
        lines.append(f"    Local work folder: {work}")
        lines.append(f"    Local MP3 folder:  {mp3}")
        lines.append(f"    Bunny target:      {bunny}")
        lines.append(f"    Rows: {rc}  |  MP3 files: {mc}")
        if errs:
            lines.append("    ERRORS:")
            for e in errs:
                lines.append(f"      - {e}")
        if warns:
            lines.append("    WARNINGS:")
            for w in warns:
                lines.append(f"      - {w}")
        lines.append("")

    lines.append("-" * 80)
    lines.append("")
    return "\n".join(lines)


def create_manifest_template(path: str) -> bool:
    """Create a CSV template for chapters manifest."""
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(
                f,
                fieldnames=MANIFEST_COLUMNS + MANIFEST_OPTIONAL,
                extrasaction="ignore",
            )
            w.writeheader()
            w.writerow({
                "book_slug": "anastasimatarion",
                "chapter_code": "AN01",
                "local_work_folder": "C:\\path\\to\\AN01\\work",
                "local_mp3_folder": "C:\\path\\to\\AN01\\mp3",
                "book_title": "Αναστασιματάριο",
                "chapter_name": "Κύριε ἐκέκραξα",
            })
        return True
    except OSError:
        return False


def build_bunny_audio_url(base_cdn_url: str, filename: str) -> str:
    """
    Public pull-zone URL for one MP3: base + '/' + encoded basename.
    Delegates to marker_matching (single implementation for stage-2 enrichment).
    """
    from marker_matching import build_bunny_audio_url as _audio_url
    return _audio_url(base_cdn_url, filename)
