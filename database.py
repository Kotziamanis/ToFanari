# -*- coding: utf-8 -*-
"""ToFanari — Excel database export for Bunny.net / FlipBuilder / production."""

import os
from typing import List, Tuple
from urllib.parse import quote

from config import (
    DEFAULT_BOOK_CODE,
    DEFAULT_BUNNY_BASE_URL,
    USE_BOOK_SUBFOLDER,
)
from pdf_ops import Marker, format_number, get_active_markers

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# Production-ready columns (with hymn assignment)
HEADERS = [
    "#",
    "Page",
    "Y",
    "Song Title",
    "Echos",
    "Section",
    "MP3 Code",
    "MP3 File",
    "Bunny URL",
    "Status",
    "Notes",
]

# Extended columns for marker-based export (marker_id, mp3_filename, match status)
HEADERS_MARKER = [
    "#",
    "Marker ID",
    "MP3 File",
    "Status",
    "Song Title",
    "Echos",
    "Section",
    "Bunny URL",
    "Notes",
]


def get_mp3_code(code: str, index_1based: int) -> str:
    """Public: MP3 code e.g. AN01-001."""
    return _mp3_basename(code, index_1based)


def get_mp3_file(code: str, index_1based: int) -> str:
    """Public: MP3 filename e.g. AN01-001.mp3."""
    return _mp3_basename(code, index_1based) + ".mp3"


def get_mp3_url(
    base_url: str,
    code: str,
    index_1based: int,
) -> str:
    """Public: Bunny URL for MP3 file."""
    return _build_mp3_url(base_url, code, index_1based, USE_BOOK_SUBFOLDER)


def build_mp3_public_url_for_source_file(
    base_url: str,
    book_code: str,
    source_mp3_basename: str,
    use_book_subfolder: bool = USE_BOOK_SUBFOLDER,
) -> str:
    """
    Public CDN URL using the actual on-disk MP3 basename (e.g. '001 Kyrie.mp3').
    Path segments are percent-encoded (spaces, Greek, etc.).
    """
    base = (base_url or "").strip().rstrip("/")
    fn = os.path.basename((source_mp3_basename or "").strip())
    if not base or not fn:
        return ""
    enc = quote(fn, safe="")
    bc = (book_code or "").strip()
    if use_book_subfolder and bc:
        return f"{base}/{quote(bc, safe='')}/{enc}"
    return f"{base}/{enc}"


def get_mp3_url_for_source_file(
    base_url: str,
    book_code: str,
    source_mp3_basename: str,
) -> str:
    """Bunny URL for a source MP3 file by real filename (not BOOK-NNN placeholder)."""
    return build_mp3_public_url_for_source_file(
        base_url, book_code, source_mp3_basename, USE_BOOK_SUBFOLDER
    )


def _mp3_basename(code: str, index_1based: int) -> str:
    """
    MP3 naming: BOOKCODE-NNN (e.g. AN01-001).
    If code empty, fallback to 001, 002 for compatibility.
    """
    num = f"{index_1based:03d}"
    code_clean = (code or "").strip()
    return f"{code_clean}-{num}" if code_clean else num


def _build_mp3_url(
    base_url: str,
    code: str,
    index_1based: int,
    use_book_subfolder: bool = USE_BOOK_SUBFOLDER,
) -> str:
    """
    Bunny URL: base/BOOKCODE/BOOKCODE-NNN.mp3 when use_book_subfolder else base/BOOKCODE-NNN.mp3.
    """
    base = base_url.rstrip("/")
    mp3_name = _mp3_basename(code, index_1based) + ".mp3"
    if use_book_subfolder and (code or "").strip():
        return f"{base}/{(code or '').strip()}/{mp3_name}"
    return f"{base}/{mp3_name}"


def _all_markers_sorted(markers: List[Marker]) -> List[Marker]:
    """Active markers sorted by page, y (for export order and numbering)."""
    return get_active_markers(markers)


def validate_mp3_files(
    mp3_folder: str,
    markers: List[Marker],
    code: str = DEFAULT_BOOK_CODE,
) -> List[str]:
    """
    Check that expected BOOKCODE-NNN.mp3 files exist in mp3_folder (current book only).
    Returns list of missing filenames. Empty list = all exist.
    """
    if not (code or "").strip():
        return []
    ordered = _all_markers_sorted(markers)
    expected = [_mp3_basename(code, i + 1) + ".mp3" for i in range(len(ordered))]
    missing = []
    for name in expected:
        path = os.path.join(mp3_folder, name)
        if not os.path.isfile(path):
            missing.append(name)
    return missing


def build_database_xlsx(
    folder: str,
    markers: List[Marker],
    code: str = DEFAULT_BOOK_CODE,
    bunny_base_url: str = DEFAULT_BUNNY_BASE_URL,
    use_book_subfolder: bool = USE_BOOK_SUBFOLDER,
    assignments: List[dict] = None,
    source_mp3_files: List[str] = None,
) -> Tuple[str, int]:
    """
    Create production-ready database.xlsx with columns:
    #, Page, Y, Song Title, Echos, Section, MP3 Code, MP3 File, Bunny URL, Status, Notes.
    assignments: list of dicts with song_title, echos, section (optional status, notes). Uses "" and "TODO" if missing.
    source_mp3_files: list of source MP3 filenames (e.g. 001 Title.mp3) from the same scan as MP3 Check.
      Rows resolve MP3 File by leading numeric prefix (001 → first file starting with 001), not row index.
      MP3 Code remains BOOKCODE-NNN.
    """
    from marker_matching import build_mp3_prefix_map

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buttons"

    header_fill = PatternFill("solid", fgColor="800000")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ordered = _all_markers_sorted(markers)
    base = (bunny_base_url or DEFAULT_BUNNY_BASE_URL).rstrip("/")
    code_clean = (code or "").strip() or "BOOK"
    prefix_map = (
        build_mp3_prefix_map(source_mp3_files) if source_mp3_files else {}
    )

    for i, m in enumerate(ordered):
        num = m.number if m.number is not None and m.number > 0 else (i + 1)
        mp3_code = _mp3_basename(code_clean, num)
        wid = format_number(num)
        mp3_file = prefix_map.get(wid, "")
        if mp3_file:
            mp3_url = build_mp3_public_url_for_source_file(
                base, code_clean, mp3_file, use_book_subfolder
            )
        else:
            mp3_url = _build_mp3_url(base, code_clean, num, use_book_subfolder)
        a = (assignments or [{}])[i] if assignments and i < len(assignments) else {}
        song_title = (a.get("song_title") or "").strip() if isinstance(a.get("song_title"), str) else ""
        echos = (a.get("echos") or "").strip() if isinstance(a.get("echos"), str) else ""
        section = (a.get("section") or "").strip() if isinstance(a.get("section"), str) else ""
        status = (a.get("status") or "TODO").strip() if isinstance(a.get("status"), str) else "TODO"
        notes = (a.get("notes") or "").strip() if isinstance(a.get("notes"), str) else ""
        ws.append([
            format_number(num),
            m.page,
            round(m.y, 1),
            song_title,
            echos,
            section,
            mp3_code,
            mp3_file,
            mp3_url,
            status,
            notes,
        ])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    path = os.path.join(folder, "database.xlsx")
    wb.save(path)
    return path, len(ordered)


def build_database_xlsx_from_marker_matches(
    folder: str,
    matches: list,
    code: str = DEFAULT_BOOK_CODE,
    bunny_base_url: str = DEFAULT_BUNNY_BASE_URL,
    use_book_subfolder: bool = USE_BOOK_SUBFOLDER,
    assignments: list = None,
) -> Tuple[str, int]:
    """
    Create database.xlsx from marker matching results (no PDF markers).
    matches: list of objects with .marker_id, .mp3_filename, .status (e.g. MarkerMatch).
    Includes Marker ID, MP3 File, Status columns. Only OK matches get Bunny URL.
    assignments: optional list of dicts with song_title, echos, section, notes.
    """
    from marker_matching import MarkerMatch

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buttons"

    header_fill = PatternFill("solid", fgColor="800000")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(HEADERS_MARKER, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    base = (bunny_base_url or DEFAULT_BUNNY_BASE_URL).rstrip("/")
    code_clean = (code or "").strip() or "BOOK"
    assignments = assignments or []

    for i, m in enumerate(matches):
        if not isinstance(m, MarkerMatch):
            continue
        marker_id = (m.marker_id or "").strip()
        mp3_file = (m.mp3_filename or "").strip()
        status = (m.status or "").strip()
        a = assignments[i] if i < len(assignments) else {}
        song_title = (a.get("song_title") or "").strip() if isinstance(a.get("song_title"), str) else ""
        echos = (a.get("echos") or "").strip() if isinstance(a.get("echos"), str) else ""
        section = (a.get("section") or "").strip() if isinstance(a.get("section"), str) else ""
        notes = (a.get("notes") or "").strip() if isinstance(a.get("notes"), str) else ""

        # MP3 code: BOOK-001 style from marker_id
        try:
            num = int(marker_id) if marker_id else (i + 1)
        except ValueError:
            num = i + 1
        mp3_code = f"{code_clean}-{num:03d}" if code_clean and code_clean != "BOOK" else f"{num:03d}"
        if status == "OK" and mp3_file:
            mp3_url = build_mp3_public_url_for_source_file(
                base, code_clean, mp3_file, use_book_subfolder
            )
        else:
            mp3_url = ""

        ws.append([
            format_number(num),
            marker_id,
            mp3_file,
            status,
            song_title,
            echos,
            section,
            mp3_url,
            notes,
        ])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    path = os.path.join(folder, "database.xlsx")
    wb.save(path)
    return path, len(matches)


def preview_lines(
    markers: List[Marker],
    code: str = DEFAULT_BOOK_CODE,
    bunny_base_url: str = DEFAULT_BUNNY_BASE_URL,
    max_lines: int = 40,
    source_mp3_files: List[str] = None,
) -> List[str]:
    """Text preview of database rows for the UI. Uses source MP3 filenames when provided."""
    from marker_matching import build_mp3_prefix_map

    ordered = _all_markers_sorted(markers)
    base = bunny_base_url.rstrip("/")
    prefix_map = build_mp3_prefix_map(source_mp3_files or [])
    lines = [f"{'ID':<6}{'Page':<6}{'MP3_File':<24}MP3_URL", "─" * 72]
    for i, m in enumerate(ordered[: max_lines - 2]):
        num = m.number if m.number is not None and m.number > 0 else (i + 1)
        wid = format_number(num)
        mp3_file = prefix_map.get(wid, "") or (_mp3_basename(code, num) + ".mp3")
        if prefix_map.get(wid):
            url = build_mp3_public_url_for_source_file(
                base, code, mp3_file, USE_BOOK_SUBFOLDER
            )
        else:
            url = _build_mp3_url(base, code, num)
        lines.append(f"{format_number(num)}   {m.page:<6}{mp3_file:<24}{url}")
    return lines


def preview_lines_from_marker_matches(
    matches: list,
    code: str = DEFAULT_BOOK_CODE,
    bunny_base_url: str = DEFAULT_BUNNY_BASE_URL,
    max_lines: int = 40,
) -> List[str]:
    """Text preview of marker-match database rows for the UI."""
    from marker_matching import MarkerMatch

    base = (bunny_base_url or DEFAULT_BUNNY_BASE_URL).rstrip("/")
    code_clean = (code or "").strip() or "BOOK"
    lines = [f"{'#':<4}{'Marker':<8}{'MP3 File':<24}{'Status':<12}URL", "─" * 72]
    for i, m in enumerate(matches[: max_lines - 2]):
        if not isinstance(m, MarkerMatch):
            continue
        marker_id = (m.marker_id or "").strip()
        mp3_file = (m.mp3_filename or "").strip()
        status = (m.status or "").strip()
        try:
            num = int(marker_id) if marker_id else (i + 1)
        except ValueError:
            num = i + 1
        if status == "OK" and mp3_file:
            url = build_mp3_public_url_for_source_file(
                base, code_clean, mp3_file, USE_BOOK_SUBFOLDER
            )
        else:
            url = ""
        lines.append(f"{format_number(num):<4}{marker_id:<8}{mp3_file:<24}{status:<12}{url}")
    return lines
