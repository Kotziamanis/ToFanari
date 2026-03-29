# -*- coding: utf-8 -*-
"""ToFanari — Book Registry: load and validate book_registry.xlsx, map chapter codes to parent books.

Parameters structure:
- BOOKS sheet: book code, title, optional category/mode, etc.
- CHAPTERS sheet: Book_Code, Chapter_Order, Chapter_Code, Chapter_Title, Is_Active
  Defines expected chapters per book with order, title, and active flag.
- Chapter_List in BOOKS: fallback comma-separated codes if CHAPTERS sheet absent.
"""

import os
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

REQUIRED_COLUMNS = [
    "Book_Code",
    "Book_Title",
    "Book_Slug",
    "Thinkific_Course_Name",
    "Subscription_Group",
    "FlipBuilder_Book_Name",
    "Bookshelf_Name",
    "Bookshelf_Order",
    "Bunny_Root_Folder",
    "Chapter_List",
    "Is_Active",
    "Notes",
]

# Optional columns
OPTIONAL_COLUMNS = ["Category", "Mode", "Expected_Chapters_Count"]

SHEET_NAME = "BOOKS"
CHAPTERS_SHEET = "CHAPTERS"
CHAPTER_COLUMNS = ["Book_Code", "Chapter_Order", "Chapter_Code", "Chapter_Title", "Is_Active"]


@dataclass
class ChapterDef:
    """Single chapter definition from Parameters."""
    order: int
    code: str
    title: str
    active: bool


def _cell_str(val: Any) -> str:
    """Convert cell value to string, strip whitespace."""
    if val is None:
        return ""
    return str(val).strip()


def _norm_book_code(s: str) -> str:
    """Normalize book code to uppercase for consistency."""
    return (s or "").strip().upper()


def _norm_chapter_code(s: str) -> str:
    """Normalize chapter code to uppercase for consistency."""
    return (s or "").strip().upper()


def _parse_chapter_list(chapter_list_str: str) -> List[str]:
    """Parse Chapter_List (e.g. 'AN01,AN02,AN03') into list of chapter codes (uppercase)."""
    if not chapter_list_str:
        return []
    parts = [_norm_chapter_code(p) for p in chapter_list_str.split(",") if (p or "").strip()]
    return [p for p in parts if p]


def parse_chapters_with_titles(chapter_list_str: str) -> List[ChapterDef]:
    """
    Parse Chapter_List supporting:
    - Comma-separated codes: AN01,AN02,AN03
    - Multi-line: AN01\\nAN02 (order=line, title empty)
    - Multi-line with titles: AN01, Title 1\\nAN02, Title 2
    Returns list of ChapterDef.
    """
    if not (chapter_list_str or "").strip():
        return []
    s = chapter_list_str.strip()
    lines = [ln.strip() for ln in s.split("\n") if ln.strip()]
    if not lines:
        return []
    if len(lines) == 1:
        parts = [p.strip() for p in lines[0].split(",") if p.strip()]
        if len(parts) >= 2:
            second = parts[1]
            if len(second) <= 12 and second.replace("_", "").replace("-", "").isalnum():
                return [ChapterDef(order=i + 1, code=_norm_chapter_code(c), title="", active=True) for i, c in enumerate(parts)]
        if len(parts) >= 1:
            code = _norm_chapter_code(parts[0])
            title = parts[1] if len(parts) > 1 else ""
            return [ChapterDef(order=1, code=code, title=title, active=True)]
        return []
    result = []
    for i, ln in enumerate(lines):
        parts = [p.strip() for p in ln.split(",", 1) if p.strip()]
        code = _norm_chapter_code(parts[0]) if parts else ""
        title = parts[1] if len(parts) > 1 else ""
        if code:
            result.append(ChapterDef(order=i + 1, code=code, title=title, active=True))
    return result


def _parse_bool(val: Any) -> bool:
    """Parse Is_Active: 1, yes, true, y, active -> True; else False."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ("1", "yes", "true", "y", "active")


def load_chapters_from_registry(path: str) -> Tuple[Dict[str, List[ChapterDef]], List[str]]:
    """
    Load CHAPTERS sheet from book_registry.xlsx.
    Returns: (book_code -> [ChapterDef ordered by Chapter_Order], errors).
    If no CHAPTERS sheet, returns ({}, []).
    """
    result: Dict[str, List[ChapterDef]] = {}
    errors: List[str] = []

    if not openpyxl or not path or not os.path.isfile(path):
        return (result, errors)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        errors.append(str(e))
        return (result, errors)

    if CHAPTERS_SHEET not in wb.sheetnames:
        wb.close()
        return (result, errors)

    ws = wb[CHAPTERS_SHEET]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return (result, errors)

    header = [_cell_str(c) for c in header_row]
    col_idx = {h: i for i, h in enumerate(header) if h}
    for col in CHAPTER_COLUMNS:
        if col not in col_idx:
            wb.close()
            return (result, [f"CHAPTERS sheet missing column: {col}"])

    raw: Dict[str, List[tuple]] = {}
    for row in rows_iter:
        if not row:
            continue
        vals = list(row) + [None] * (len(header) - len(row))
        book_code = _norm_book_code(vals[col_idx["Book_Code"]])
        order_val = vals[col_idx["Chapter_Order"]]
        code = _norm_chapter_code(vals[col_idx["Chapter_Code"]])
        title = _cell_str(vals[col_idx["Chapter_Title"]])
        active = _parse_bool(vals[col_idx["Is_Active"]])
        if not book_code or not code:
            continue
        try:
            order = int(order_val) if order_val is not None else 0
        except (TypeError, ValueError):
            order = 0
        raw.setdefault(book_code, []).append((order, code, title, active))

    wb.close()

    for book_code, rows in raw.items():
        rows.sort(key=lambda r: r[0])
        result[book_code] = [
            ChapterDef(order=o, code=c, title=t, active=a) for o, c, t, a in rows
        ]
    return (result, errors)


def get_expected_chapters_for_book(
    book_code: str,
    books: List[Dict[str, Any]],
    chapters_data: Dict[str, List[ChapterDef]],
) -> List[ChapterDef]:
    """
    Get expected chapters for a book. Uses CHAPTERS sheet if available, else falls back to Chapter_List.
    """
    bc_norm = _norm_book_code(book_code)
    if chapters_data and bc_norm in chapters_data:
        return chapters_data[bc_norm]
    for b in books:
        if _norm_book_code(b.get("Book_Code") or "") == bc_norm:
            codes = _parse_chapter_list(b.get("Chapter_List", ""))
            return [
                ChapterDef(order=i + 1, code=c, title="", active=True)
                for i, c in enumerate(codes)
            ]
    return []


def compare_imported_vs_expected(
    book_code: str,
    expected: List[ChapterDef],
    imported: List[str],
    expected_count: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Compare imported chapter codes against expected.
    Returns: {
        "imported": [codes that are in imported and expected],
        "missing": [expected active codes not in imported],
        "missing_orders": [chapter order numbers that are missing],
        "duplicate": [codes imported more than once],
        "invalid": [imported codes not in expected],
        "merge_allowed": bool,
        "expected_active": [expected active chapter codes],
        "expected_count": int,
        "imported_count": int,
        "status": "COMPLETE" | "INCOMPLETE" | "ERROR",
    }
    """
    expected_codes = {ch.code for ch in expected}
    expected_active = [ch.code for ch in expected if ch.active]
    expected_count_val = expected_count if expected_count is not None else len(expected_active)
    imported_set = set(imported)
    count: Dict[str, int] = {}
    for c in imported:
        count[c] = count.get(c, 0) + 1

    imported_ok = [c for c in imported if c in expected_codes and count.get(c, 0) == 1]
    missing = [c for c in expected_active if c not in imported_set]
    code_to_order = {ch.code: ch.order for ch in expected if ch.active}
    missing_orders = sorted([code_to_order[c] for c in missing if c in code_to_order])
    duplicate = [c for c, n in count.items() if n > 1]
    invalid = [c for c in imported if c not in expected_codes]

    has_errors = len(duplicate) > 0 or len(invalid) > 0
    has_missing = len(missing) > 0

    if has_errors:
        status = "ERROR"
    elif has_missing:
        status = "INCOMPLETE"
    else:
        status = "COMPLETE"

    merge_allowed = (
        status == "COMPLETE"
        and all(c in imported_set for c in expected_active)
        and len(duplicate) == 0
        and len(invalid) == 0
    )

    return {
        "imported": imported_ok,
        "missing": missing,
        "missing_orders": missing_orders,
        "duplicate": duplicate,
        "invalid": invalid,
        "merge_allowed": merge_allowed,
        "expected_active": expected_active,
        "expected_count": expected_count_val,
        "imported_count": len(imported_ok),
        "status": status,
    }


def load_book_registry(path: str) -> Tuple[bool, List[Dict[str, Any]], List[str]]:
    """
    Load book_registry.xlsx from path.
    Returns: (ok, list of book rows, list of error messages).
    Each row is a dict with column names as keys.
    """
    errors: List[str] = []
    books: List[Dict[str, Any]] = []

    if not openpyxl:
        return (False, [], ["openpyxl is required. Install with: pip install openpyxl"])

    if not path or not os.path.isfile(path):
        return (False, [], [f"File not found: {path}"])

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return (False, [], [f"Failed to open file: {str(e)}"])

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return (False, [], [f"Sheet '{SHEET_NAME}' not found in workbook."])

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)

    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return (False, [], ["Sheet BOOKS is empty."])

    header = [_cell_str(c) for c in header_row]
    col_idx = {h: i for i, h in enumerate(header) if h}

    for req in REQUIRED_COLUMNS:
        if req not in col_idx:
            wb.close()
            return (False, [], [f"Required column '{req}' not found."])

    seen_book_codes: set = set()
    seen_book_slugs: set = set()
    chapter_to_books: Dict[str, List[str]] = {}

    for row_idx, row in enumerate(rows_iter, start=2):
        if not row:
            continue
        vals = list(row) + [None] * (len(header) - len(row))
        book: Dict[str, Any] = {}
        for col in REQUIRED_COLUMNS:
            idx = col_idx.get(col, -1)
            book[col] = _cell_str(vals[idx]) if idx >= 0 else ""
        for col in OPTIONAL_COLUMNS:
            idx = col_idx.get(col, -1)
            if idx >= 0:
                book[col] = _cell_str(vals[idx]) if vals[idx] is not None else ""

        book_code = _norm_book_code(book.get("Book_Code", ""))
        book_slug = _norm_book_code(book.get("Book_Slug", "").replace("-", "_")).replace("_", "-")
        book["Book_Code"] = book_code
        book["Book_Slug"] = book_slug
        chapter_list_str = book.get("Chapter_List", "")

        if not book_code:
            errors.append(f"Row {row_idx}: Book_Code is empty.")
        elif book_code in seen_book_codes:
            errors.append(f"Row {row_idx}: Duplicate Book_Code: {book_code}")
        else:
            seen_book_codes.add(book_code)

        if not book_slug:
            errors.append(f"Row {row_idx}: Book_Slug is empty.")
        elif book_slug in seen_book_slugs:
            errors.append(f"Row {row_idx}: Duplicate Book_Slug: {book_slug}")
        else:
            seen_book_slugs.add(book_slug)

        if not chapter_list_str:
            errors.append(f"Row {row_idx}: Chapter_List is empty.")

        chapters = _parse_chapter_list(chapter_list_str)
        for ch in chapters:
            if ch not in chapter_to_books:
                chapter_to_books[ch] = []
            chapter_to_books[ch].append(book_code)

        books.append(book)

    wb.close()

    for ch, codes in chapter_to_books.items():
        if len(codes) > 1:
            errors.append(
                f"Chapter code {ch} appears in more than one book in book_registry.xlsx"
            )

    # Load CHAPTERS sheet and attach to books
    chapters_data, ch_errors = load_chapters_from_registry(path)
    errors.extend(ch_errors)
    for book in books:
        bc = book.get("Book_Code", "")
        book["chapters"] = get_expected_chapters_for_book(bc, books, chapters_data)
        if not book.get("Chapter_List", "").strip() and book["chapters"]:
            book["Chapter_List"] = ",".join(ch.code for ch in book["chapters"])
        active_count = len([ch for ch in book["chapters"] if ch.active])
        raw = (book.get("Expected_Chapters_Count") or "").strip()
        if raw:
            try:
                book["expected_chapters_count"] = int(raw)
            except ValueError:
                book["expected_chapters_count"] = active_count
        else:
            book["expected_chapters_count"] = active_count

    ok = len(errors) == 0
    return (ok, books, errors)


def _ensure_registry_file(path: str) -> Tuple[bool, Any, List[str]]:
    """
    Open or create book_registry.xlsx. Returns (ok, workbook, errors).
    Caller must close workbook after use.
    """
    errors: List[str] = []
    if not openpyxl:
        return (False, None, ["openpyxl is required. Install with: pip install openpyxl"])

    if not path or not path.strip():
        return (False, None, ["No file path specified."])

    path = path.strip()
    if os.path.isfile(path):
        try:
            wb = openpyxl.load_workbook(path)
            return (True, wb, [])
        except Exception as e:
            return (False, None, [f"Failed to open file: {str(e)}"])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        for col, name in enumerate(REQUIRED_COLUMNS, 1):
            ws.cell(row=1, column=col, value=name)
        return (True, wb, [])


def _ensure_chapters_sheet(wb: Any) -> Any:
    """Ensure CHAPTERS sheet exists with headers. Returns the worksheet."""
    if CHAPTERS_SHEET in wb.sheetnames:
        ws = wb[CHAPTERS_SHEET]
    else:
        ws = wb.create_sheet(CHAPTERS_SHEET)
    if ws.max_row == 0:
        for col, name in enumerate(CHAPTER_COLUMNS, 1):
            ws.cell(row=1, column=col, value=name)
    return ws


def append_book_to_registry(path: str, book: Dict[str, Any]) -> Tuple[bool, List[str]]:
    """
    Append a new book row to book_registry.xlsx. Creates file if it does not exist.
    Also creates CHAPTERS rows from Chapter_List (or from book.get("chapters")).
    Returns: (ok, list of error messages).
    """
    errors: List[str] = []

    book_code = _norm_book_code(book.get("Book_Code") or "")
    book_title = (book.get("Book_Title") or "").strip()
    book_slug = _norm_book_code((book.get("Book_Slug") or "").replace("-", "_")).replace("_", "-")
    chapter_list = (book.get("Chapter_List") or "").strip()

    if not book_code:
        errors.append("Book_Code is required.")
    if not book_title:
        errors.append("Book_Title is required.")
    if not book_slug:
        errors.append("Book_Slug is required.")
    if not chapter_list and not book.get("chapters"):
        errors.append("Chapter_List or chapters is required.")

    expected_cnt = (book.get("Expected_Chapters_Count") or "").strip()
    if expected_cnt:
        try:
            int(expected_cnt)
        except ValueError:
            errors.append("Expected_Chapters_Count must be a number.")

    if errors:
        return (False, errors)

    ok, wb, load_errs = _ensure_registry_file(path)
    if not ok or not wb:
        return (False, load_errs)

    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        if ws.title != SHEET_NAME:
            ws = wb.create_sheet(SHEET_NAME)
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = {str(h).strip(): i for i, h in enumerate(header_row) if h}
        if "Expected_Chapters_Count" not in col_idx:
            next_col = ws.max_column + 1
            ws.cell(row=1, column=next_col, value="Expected_Chapters_Count")
            col_idx["Expected_Chapters_Count"] = next_col - 1
        next_row = ws.max_row + 1
        code_col = col_idx.get("Book_Code", 0) + 1
        slug_col = col_idx.get("Book_Slug", 2) + 1
        existing_codes = set()
        existing_slugs = set()
        for r in range(2, next_row):
            code_val = ws.cell(row=r, column=code_col).value
            slug_val = ws.cell(row=r, column=slug_col).value
            if code_val:
                existing_codes.add(_norm_book_code(str(code_val)))
            if slug_val:
                existing_slugs.add(_norm_book_code(str(slug_val).replace("-", "_")).replace("_", "-"))
        if book_code in existing_codes:
            return (False, [f"Duplicate Book_Code: {book_code}"])
        if book_slug in existing_slugs:
            return (False, [f"Duplicate Book_Slug: {book_slug}"])

        values = [
            book_code,
            book_title,
            book_slug,
            book.get("Thinkific_Course_Name", ""),
            book.get("Subscription_Group", ""),
            book.get("FlipBuilder_Book_Name", ""),
            book.get("Bookshelf_Name", ""),
            book.get("Bookshelf_Order", ""),
            book.get("Bunny_Root_Folder", ""),
            chapter_list or "",
            book.get("Is_Active", ""),
            book.get("Notes", ""),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=next_row, column=col, value=val if val is not None else "")

        raw_expected = (book.get("Expected_Chapters_Count") or "").strip()
        expected_cnt = None
        if raw_expected:
            try:
                expected_cnt = int(raw_expected)
            except ValueError:
                pass
        if expected_cnt is None and chapter_list:
            ch_list_pre = parse_chapters_with_titles(chapter_list)
            expected_cnt = len(ch_list_pre) if ch_list_pre else 0
        elif expected_cnt is None:
            expected_cnt = 0
        ecc_col = col_idx.get("Expected_Chapters_Count")
        if ecc_col is not None:
            ws.cell(row=next_row, column=ecc_col + 1, value=int(expected_cnt) if expected_cnt else "")

        # Write CHAPTERS rows
        ch_list = book.get("chapters")
        if not ch_list and chapter_list:
            ch_list = parse_chapters_with_titles(chapter_list)
        if ch_list:
            ch_ws = _ensure_chapters_sheet(wb)
            next_ch_row = ch_ws.max_row + 1
            for ch in ch_list:
                if isinstance(ch, ChapterDef):
                    ch_ws.cell(row=next_ch_row, column=1, value=book_code)
                    ch_ws.cell(row=next_ch_row, column=2, value=ch.order)
                    ch_ws.cell(row=next_ch_row, column=3, value=ch.code)
                    ch_ws.cell(row=next_ch_row, column=4, value=ch.title)
                    ch_ws.cell(row=next_ch_row, column=5, value="1" if ch.active else "0")
                    next_ch_row += 1

        wb.save(path)
        return (True, [])
    except Exception as e:
        return (False, [str(e)])


def append_chapter_to_registry(
    path: str,
    book_code: str,
    chapter_order: int,
    chapter_code: str,
    chapter_title: str = "",
    is_active: bool = True,
) -> Tuple[bool, List[str]]:
    """Append one chapter row to CHAPTERS sheet."""
    if not path or not os.path.isfile(path) or not openpyxl:
        return (False, ["Invalid path or openpyxl missing"])
    try:
        wb = openpyxl.load_workbook(path)
        ch_ws = _ensure_chapters_sheet(wb)
        next_row = ch_ws.max_row + 1
        ch_ws.cell(row=next_row, column=1, value=(book_code or "").strip())
        ch_ws.cell(row=next_row, column=2, value=chapter_order)
        ch_ws.cell(row=next_row, column=3, value=(chapter_code or "").strip())
        ch_ws.cell(row=next_row, column=4, value=(chapter_title or "").strip())
        ch_ws.cell(row=next_row, column=5, value="1" if is_active else "0")
        wb.save(path)
        return (True, [])
    except Exception as e:
        return (False, [str(e)])


def get_all_catalog_entries(
    books: List[Dict[str, Any]],
) -> List[Tuple[str, str, str]]:
    """
    Return list of (chapter_code, book_slug, book_title) for all active chapters in catalog.
    Used for dropdown/selection when operator must choose from predefined catalog only.
    """
    result: List[Tuple[str, str, str]] = []
    for b in books:
        slug = (b.get("Book_Slug") or "").strip()
        title = (b.get("Book_Title") or "").strip()
        for ch in b.get("chapters") or []:
            if isinstance(ch, ChapterDef) and ch.active:
                result.append((ch.code, slug, title))
    return result


def _norm_slug(s: str) -> str:
    """Normalize book slug for comparison (uppercase, preserve hyphen)."""
    return _norm_book_code((s or "").replace("-", "_")).replace("_", "-")


def validate_chapter_in_catalog(
    chapter_code: str,
    book_slug: str,
    books: List[Dict[str, Any]],
) -> Tuple[bool, Optional[Dict[str, Any]], List[str]]:
    """
    Check that chapter_code belongs to the given book in the catalog.
    Returns (ok, book_dict_or_none, list of error messages).
    """
    errors: List[str] = []
    ch = _norm_chapter_code(chapter_code or "")
    slug_norm = _norm_slug(book_slug or "")
    if not ch:
        return (False, None, ["Chapter code is required."])
    if not books:
        return (False, None, ["Master catalog not loaded. Load Parameters first."])
    book = None
    for b in books:
        if _norm_slug(b.get("Book_Slug") or "") == slug_norm:
            book = b
            break
    if not book:
        errors.append(f"Book '{book_slug}' not found in catalog.")
        return (False, None, errors)
    chapter_codes = [c.code for c in (book.get("chapters") or []) if isinstance(c, ChapterDef) and c.active]
    if ch not in chapter_codes:
        errors.append(f"Chapter '{ch}' is not in book '{slug}'. Expected: {', '.join(chapter_codes[:10])}{'...' if len(chapter_codes) > 10 else ''}")
        return (False, book, errors)
    return (True, book, [])


def find_book_for_chapter(
    chapter_code: str, books: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    """
    Given a chapter code (e.g. AN01), return the matching book registry row.
    Returns None if not found or if chapter appears in multiple books (caller should validate first).
    """
    if not chapter_code or not books:
        return None
    ch = _norm_chapter_code(chapter_code)
    if not ch:
        return None
    found = None
    for book in books:
        chapter_list_str = book.get("Chapter_List", "")
        chapters_from_list = _parse_chapter_list(chapter_list_str)
        chapters_from_def = [c.code for c in (book.get("chapters") or []) if isinstance(c, ChapterDef)]
        if ch in chapters_from_list or ch in chapters_from_def:
            if found is not None:
                return None
            found = book
    return found
